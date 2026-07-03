"""
Report Builder module for the Fusion Report Engine.

Defines the orchestrator class `ReportBuilder` and the `ReportRequest` dataclass.
All methods are configured with logging, timing decorators, docstrings, and placeholders.
"""

from dataclasses import dataclass
from datetime import date
import logging
from typing import Any, Optional
import polars as pl
import psycopg2

from .timer import timer

# Get logger for report builder
logger = logging.getLogger("fusion_report_engine.report_builder")

@dataclass
class ReportRequest:
    """
    Data transfer object encapsulating parameters for generating a report.
    """
    start_date: date
    end_date: date
    page: int = 1
    page_size: int = 100
    export: bool = False
    output_path: Optional[str] = None

class ReportBuilder:
    """
    Main orchestration class that guides the step-by-step assembly of reports.

    Acts as the entry point for the Django API.
    """

    def __init__(self) -> None:
        """
        Initializes the ReportBuilder system.
        """
        self.logger = logger
        self.conn: Any = None
        self.request: Optional[ReportRequest] = None

        # Source DataFrames
        self.loan_mst_df: Optional[pl.DataFrame] = None
        self.rawfile_df: Optional[pl.DataFrame] = None
        self.rawfile_backup_df: Optional[pl.DataFrame] = None
        self.voicebot_df: Optional[pl.DataFrame] = None
        self.whatsapp_df: Optional[pl.DataFrame] = None
        self.blaster_df: Optional[pl.DataFrame] = None

        # Target DataFrames
        self.base_accounts_df: Optional[pl.DataFrame] = None
        self.all_call_dispositions_df: Optional[pl.DataFrame] = None
        self.all_wa_dispositions_df: Optional[pl.DataFrame] = None
        self.all_blaster_dispositions_df: Optional[pl.DataFrame] = None
        self.best_disposition_df: Optional[pl.DataFrame] = None
        self.voicebot_summary_df: Optional[pl.DataFrame] = None
        self.whatsapp_summary_df: Optional[pl.DataFrame] = None
        self.blaster_summary_df: Optional[pl.DataFrame] = None
        self.communication_summary_df: Optional[pl.DataFrame] = None
        self.latest_master_df: Optional[pl.DataFrame] = None

        self.response_df: Optional[pl.DataFrame] = None
        self.whatsapp_messages_df: Optional[pl.DataFrame] = None
        self.voicebot_mtd_base_df: Optional[pl.DataFrame] = None
        self.voicebot_response_map_df: Optional[pl.DataFrame] = None
        self.whatsapp_mtd_base_df: Optional[pl.DataFrame] = None
        self.whatsapp_response_map_df: Optional[pl.DataFrame] = None
        self.blaster_mtd_base_df: Optional[pl.DataFrame] = None

        self.voicebot_attempt_totals_df: Optional[pl.DataFrame] = None
        self.latest_call_df: Optional[pl.DataFrame] = None
        self.latest_call_extra_df: Optional[pl.DataFrame] = None
        self.blaster_attempt_totals_df: Optional[pl.DataFrame] = None
        self.latest_blaster_df: Optional[pl.DataFrame] = None
        self.latest_whatsapp_df: Optional[pl.DataFrame] = None
        self.whatsapp_combined_df: Optional[pl.DataFrame] = None

        self.transactions_df: Optional[pl.DataFrame] = None
        self.latest_collection_df: Optional[pl.DataFrame] = None
        self.total_collection_df: Optional[pl.DataFrame] = None

        self.mtd_connection_flags_df: Optional[pl.DataFrame] = None
        self.mtd_wa_connection_flags_df: Optional[pl.DataFrame] = None
        self.last_connected_info_df: Optional[pl.DataFrame] = None
        self.last_connected_call_duration_df: Optional[pl.DataFrame] = None
        self.call_duration_stats_df: Optional[pl.DataFrame] = None
        self.voicebot_call_duration_stats_df: Optional[pl.DataFrame] = None

        self.voicebot_lm_df: Optional[pl.DataFrame] = None
        self.whatsapp_lm_df: Optional[pl.DataFrame] = None
        self.last_month_best_disposition_df: Optional[pl.DataFrame] = None
        self.latest_disposition_mtd_df: Optional[pl.DataFrame] = None
        self.total_ptp_responses_df: Optional[pl.DataFrame] = None

        self.latest_response_df: Optional[pl.DataFrame] = None
        self.latest_voicebot_response_df: Optional[pl.DataFrame] = None
        self.latest_whatsapp_response_df: Optional[pl.DataFrame] = None
        self.latest_voicebot_channel_disposition_df: Optional[pl.DataFrame] = None
        self.latest_whatsapp_channel_disposition_df: Optional[pl.DataFrame] = None
        self.latest_blaster_channel_disposition_df: Optional[pl.DataFrame] = None

        self.daily_call_dispositions_df: Optional[pl.DataFrame] = None
        self.month_dates_df: Optional[pl.DataFrame] = None
        self.daily_pivots_df: Optional[pl.DataFrame] = None
        self.final_report_df: Optional[pl.DataFrame] = None

    @timer
    def load_source_tables(self) -> None:
        """
        Loads the raw input tables (LoanMst, Rawfile, Rawfile_Backup) from database sources.
        """
        if not self.request:
            raise ValueError("ReportRequest reference is not set.")

        start_date = self.request.start_date
        end_date = self.request.end_date

        self.logger.info(
            "Loading raw source tables (LoanMst, Rawfile, Rawfile_Backup) for range %s to %s",
            start_date,
            end_date
        )

        from . import config
        db_config = config.DATABASE_CONFIG

        try:
            self.conn = psycopg2.connect(**db_config)
            self.logger.info("Successfully connected to PostgreSQL database.")
        except Exception as e:
            self.logger.error("Failed to connect to PostgreSQL database: %s", e)
            raise e

        try:
            with self.conn.cursor() as cursor:
                # 1. Rawfile
                self.logger.info("Loading Rawfile data...")
                rawfile_sql = """
                    SELECT *
                    FROM rawfile r
                    WHERE r.bankmstid = 53
                      AND r.inserted_date::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(rawfile_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.rawfile_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded rawfile rows: %d", len(self.rawfile_df))

                # 2. Rawfile_Backup
                self.logger.info("Loading Rawfile_Backup data...")
                rawfile_backup_sql = """
                    SELECT *
                    FROM rawfile_backup rb
                    WHERE rb.bankmstid = 53
                      AND rb.inserted_date::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(rawfile_backup_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.rawfile_backup_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded rawfile_backup rows: %d", len(self.rawfile_backup_df))

                # 3. LoanMst (filtered based on cohort loaded in rawfile)
                self.logger.info("Loading LoanMst data...")
                loan_mst_sql = """
                    SELECT DISTINCT lm.*
                    FROM "LoanMst" lm
                    JOIN rawfile r ON r.disbursementid = lm."DisbursementID"
                    WHERE lm."BankMstID" = 53
                      AND r.bankmstid = 53
                      AND r.inserted_date::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(loan_mst_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.loan_mst_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded LoanMst rows: %d", len(self.loan_mst_df))

                # 4. VoiceBotHistory
                self.logger.info("Loading VoiceBotHistory data...")
                voicebot_sql = """
                    SELECT *
                    FROM "VoiceBotHistory" vbh
                    WHERE vbh."BankMstID" = 53
                      AND vbh."CreatedDate"::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(voicebot_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.voicebot_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded VoiceBot rows: %d", len(self.voicebot_df))

                # 5. WhatsAppHistory
                self.logger.info("Loading WhatsAppHistory data...")
                whatsapp_sql = """
                    SELECT *
                    FROM "WhatsAppHistory" wah
                    WHERE wah."BankMstID" = 53
                      AND wah."CreatedDate"::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(whatsapp_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.whatsapp_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded WhatsApp rows: %d", len(self.whatsapp_df))

                # 6. BlasterHistory
                self.logger.info("Loading BlasterHistory data...")
                blaster_sql = """
                    SELECT *
                    FROM "BlasterHistory" bh
                    WHERE bh."BankMstID" = 53
                      AND bh."CreatedDate"::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(blaster_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.blaster_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded Blaster rows: %d", len(self.blaster_df))

                # 7. Response
                self.logger.info("Loading Response data...")
                response_sql = """
                    SELECT *
                    FROM "Response" r
                    WHERE r."BankMstID" = 53
                      AND r."Status" NOT IN ('', 'Terminated')
                """
                cursor.execute(response_sql)
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.response_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded Response rows: %d", len(self.response_df))

                # 8. WhatsApp_Messages
                self.logger.info("Loading WhatsApp_Messages data...")
                whatsapp_messages_sql = """
                    SELECT *
                    FROM "WhatsApp_Messages" wam
                    WHERE wam."MessageDate"::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(whatsapp_messages_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.whatsapp_messages_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded WhatsApp_Messages rows: %d", len(self.whatsapp_messages_df))

                # 9. Transactions
                self.logger.info("Loading Transactions data...")
                transactions_sql = """
                    SELECT *
                    FROM "Transactions" t
                    WHERE t."BankMstID" = 53
                      AND t."CollectedDate"::date BETWEEN %(start_date)s AND %(end_date)s
                """
                cursor.execute(transactions_sql, {"start_date": start_date, "end_date": end_date})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.transactions_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded Transactions rows: %d", len(self.transactions_df))

                # 10. VoiceBotHistory last month (lm)
                self.logger.info("Loading VoiceBotHistory last month data...")
                from dateutil.relativedelta import relativedelta
                from datetime import timedelta
                lm_start = (start_date - relativedelta(months=1)).replace(day=1)
                lm_end = start_date - timedelta(days=1)

                voicebot_lm_sql = """
                    SELECT *
                    FROM "VoiceBotHistory" vbh
                    WHERE vbh."BankMstID" = 53
                      AND vbh."CallTried" > 0
                      AND vbh."CallID" IS NOT NULL
                      AND vbh."CreatedDate"::date BETWEEN %(lm_start)s AND %(lm_end)s
                      AND vbh."CampaignMstID" IN (1282, 1441, 2304, 2312, 1979, 1973, 1972, 2010, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12)
                """
                cursor.execute(voicebot_lm_sql, {"lm_start": lm_start, "lm_end": lm_end})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.voicebot_lm_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded VoiceBot last month rows: %d", len(self.voicebot_lm_df))

                # 11. WhatsAppHistory last month (lm)
                self.logger.info("Loading WhatsAppHistory last month data...")
                whatsapp_lm_sql = """
                    SELECT *
                    FROM "WhatsAppHistory" wah
                    WHERE wah."BankMstID" = 53
                      AND wah."IsSent" = TRUE
                      AND wah."CreatedDate"::date BETWEEN %(lm_start)s AND %(lm_end)s
                      AND wah."CampaignMstID" IN (1283, 1442, 2305, 2313, 1974, 1975)
                """
                cursor.execute(whatsapp_lm_sql, {"lm_start": lm_start, "lm_end": lm_end})
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                self.whatsapp_lm_df = pl.DataFrame(
                    {col: [row[i] for row in rows] for i, col in enumerate(columns)}
                )
                self.logger.info("Loaded WhatsApp last month rows: %d", len(self.whatsapp_lm_df))

        except Exception as e:
            self.logger.error("Error executing database queries: %s", e)
            raise e
        finally:
            if self.conn and not self.conn.closed:
                self.conn.close()
                self.logger.info("Closed PostgreSQL connection.")

    @timer
    def build_base_accounts(self) -> None:
        """
        Builds the base account list from allocation tables.
        """
        self.logger.info("Building base accounts...")
        if self.rawfile_df is None or self.rawfile_backup_df is None or self.loan_mst_df is None:
            raise ValueError("Source dataframes not loaded. Call load_source_tables first.")

        # Step 1: Concatenate Rawfiles
        self.logger.info("Building combined_rawfile...")
        
        # Define expected rawfile columns to select
        columns_to_select = [
            "disbursementid", "customername", "customerid", "mobileno",
            "branchname", "branchcode", "bankmstid", "loanclassification",
            "overdueamount", "emiamount", "totaloutstanding", "inserted_date"
        ]
        
        existing_rf_cols = [col for col in columns_to_select if col in self.rawfile_df.columns]
        rf = self.rawfile_df.select(existing_rf_cols).with_columns(pl.lit(1).alias("source_priority"))
        
        existing_rf_backup_cols = [col for col in columns_to_select if col in self.rawfile_backup_df.columns]
        rf_backup = self.rawfile_backup_df.select(existing_rf_backup_cols).with_columns(pl.lit(2).alias("source_priority"))
        
        combined_rawfile = pl.concat([rf, rf_backup], how="diagonal")
        self.logger.info("Combined Rawfile rows: %d", len(combined_rawfile))

        # Step 2: Deduplicate (DISTINCT ON disbursementid equivalent)
        self.logger.info("Building latest_rawfile...")
        latest_rawfile = (
            combined_rawfile
            .sort(
                by=["disbursementid", "inserted_date", "source_priority"],
                descending=[False, True, False]
            )
            .unique(subset=["disbursementid"], keep="first")
        )
        self.logger.info("Latest Rawfile rows: %d", len(latest_rawfile))

        # Step 3: Join LoanMst and Apply Precedence
        self.logger.info("Joining LoanMst...")
        joined = latest_rawfile.join(
            self.loan_mst_df,
            left_on="disbursementid",
            right_on="DisbursementID",
            how="inner"
        )
        # Apply BankMstID = 53 filter
        joined = joined.filter(pl.col("BankMstID") == 53)
        
        # Sort by disbursementid ascending, IsActive descending, LoanMstID descending
        # and unique on disbursementid to match DISTINCT ON
        sorted_base = (
            joined
            .sort(
                by=["disbursementid", "IsActive", "LoanMstID"],
                descending=[False, True, True]
            )
            .unique(subset=["disbursementid"], keep="first")
        )

        # Step 4: Project and Rename Columns
        self.logger.info("Building base_accounts...")
        self.base_accounts_df = sorted_base.select([
            pl.col("LoanMstID").alias("LoanMstID"),
            pl.col("disbursementid").alias("DisbursementID"),
            pl.col("customername").alias("CustomerName"),
            pl.col("customerid").alias("CustomerNumber"),
            pl.col("mobileno").alias("Primary Mobile Number"),
            pl.col("branchname").alias("Branch"),
            pl.col("branchcode").alias("BranchCode"),
            pl.col("bankmstid").alias("BankMstID"),
            pl.col("loanclassification").alias("LoanClassification"),
            pl.when(pl.col("IsActive") == True).then(pl.lit("Active")).otherwise(pl.lit("Inactive")).alias("Loan Status"),
            pl.when(pl.col("DND") == True).then(pl.lit(1)).otherwise(pl.lit(0)).alias("DND"),
            pl.col("overdueamount").fill_null(0.0).alias("Default Amt"),
            pl.col("emiamount").fill_null(0.0).alias("EMI Amount"),
            pl.col("totaloutstanding").fill_null(0.0).alias("Total Outstanding")
        ])
        
        # Step 5: Validation, Profiling & Logging
        row_count = len(self.base_accounts_df)
        unique_loans = self.base_accounts_df["LoanMstID"].n_unique()
        unique_disb = self.base_accounts_df["DisbursementID"].n_unique()
        null_counts = self.base_accounts_df.null_count().to_dicts()[0]
        memory_bytes = self.base_accounts_df.estimated_size()
        
        self.logger.info("Base Accounts rows: %d", row_count)
        self.logger.info("Unique LoanMstID count: %d", unique_loans)
        self.logger.info("Unique DisbursementID count: %d", unique_disb)
        self.logger.info("Null counts per column: %s", null_counts)
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", memory_bytes, memory_bytes / 1024.0)

    @timer
    def build_all_call_dispositions(self) -> None:
        """
        Processes and maps VoiceBot call attempts into a unified disposition engine state.
        Stores the result as self.all_call_dispositions_df.
        """
        self.logger.info("Starting VoiceBot disposition processing (all_call_dispositions)...")
        if self.voicebot_df is None:
            raise ValueError("VoiceBot data is not loaded. Call load_source_tables first.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded. Call build_base_accounts first.")

        # Log initial VoiceBot count
        self.logger.info("Validation Log - Loaded VoiceBot rows: %d", len(self.voicebot_df))

        # Join VoiceBot with base_accounts (SQL: JOIN base_accounts ba ON ba."LoanMstID" = vbh."LoanMstID")
        # and filter (SQL: WHERE vbh."BankMstID" = 53 AND vbh."CallTried" > 0 AND vbh."CallID" IS NOT NULL AND vbh."CallID" <> '')
        joined = (
            self.voicebot_df
            .join(
                self.base_accounts_df.select(["LoanMstID"]),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (pl.col("BankMstID") == 53) &
                (pl.col("CallTried") > 0) &
                (pl.col("CallID").is_not_null()) &
                (pl.col("CallID") != "")
            )
        )
        self.logger.info("Validation Log - Rows after join with base_accounts and filtering: %d", len(joined))

        # Extract CreatedDate as call_date
        created_date_col = "CreatedDate" if "CreatedDate" in joined.columns else "createddate"
        joined = joined.with_columns(
            pl.col(created_date_col).cast(pl.Date).alias("call_date")
        )

        # Derive is_connected and call_duration
        recording_col = "Recording" if "Recording" in joined.columns else "recording"
        duration_col = "CallDuration" if "CallDuration" in joined.columns else "callduration"
        
        is_connected_expr = (
            pl.col(recording_col).is_not_null() &
            (pl.col(recording_col).str.len_bytes() > 1)
        )
        
        call_duration_expr = (
            pl.when(pl.col(duration_col) == 0)
            .then(pl.lit(1))
            .otherwise(pl.col(duration_col).fill_null(0))
        )

        classified = joined.with_columns([
            is_connected_expr.alias("is_connected"),
            call_duration_expr.alias("call_duration")
        ])
        
        # Log rows after connection classification
        self.logger.info("Validation Log - Rows after connected/non-connected classification: %d", len(classified))

        # Prep disposition source strings (handle potential case variation from DB)
        ext_col1 = "Extracolumn1" if "Extracolumn1" in classified.columns else "extracolumn1"
        dispo_col = "deposition" if "deposition" in classified.columns else "Disposition" if "Disposition" in classified.columns else "deposition"
        
        connected_disposition_source = pl.coalesce([
            pl.col(ext_col1).str.strip_chars(),
            pl.col(dispo_col).str.strip_chars()
        ]).fill_null("")
        
        connected_disposition_source_upper = connected_disposition_source.str.to_uppercase()
        connected_disposition_source_lower = connected_disposition_source.str.to_lowercase()
        connected_disposition_source_trimmed = connected_disposition_source

        # Extracolumn1 normalization for non-connected
        extracolumn1_trimmed = pl.col(ext_col1).str.strip_chars().fill_null("")
        extracolumn1_upper = extracolumn1_trimmed.str.to_uppercase()
        extracolumn1_compact = (
            extracolumn1_upper
            .str.replace_all(r"^[0-9]+[_ \t-]*", "")
            .str.replace_all(r"[^A-Z0-9]", "")
        )

        # ----------------------------------------------------
        # CONNECTED CASE EXPRESSION (excluding Response mappings)
        # ----------------------------------------------------
        ptp_settlement_cond = (
            connected_disposition_source_upper.str.contains("SETTLEMENT") & 
            (
                connected_disposition_source_upper.str.contains("PTP") |
                connected_disposition_source_upper.str.contains("PROMISE") |
                connected_disposition_source_upper.str.contains("PENDING") |
                connected_disposition_source_lower.is_in(["futurepay", "ready_to_pay", "willing_to_pay"])
            )
        )

        other_settlement_cond = (
            connected_disposition_source_upper.str.contains("SETTLEMENT") & 
            (pl.col("call_duration") < 60)
        )

        connected_expr = (
            pl.when(connected_disposition_source_trimmed == "")
            .then(pl.lit("No Response"))
            .when(
                connected_disposition_source_upper.str.contains("ALREADY PAID") |
                (connected_disposition_source_lower == "already_paid")
            )
            .then(pl.lit("Already Paid"))
            .when(connected_disposition_source_upper.str.contains("NO COMM"))
            .then(pl.lit("Hang Up / No Response"))
            .when(
                connected_disposition_source_upper.str.contains("DENIED") |
                connected_disposition_source_upper.str.contains("DENIES") |
                connected_disposition_source_upper.str.contains("DENIAL") |
                connected_disposition_source_upper.str.contains("REFUSED") |
                (connected_disposition_source_lower == "not_ready_to_pay")
            )
            .then(pl.lit("Refused to pay"))
            .when(connected_disposition_source_upper.str.contains("CONFUSION"))
            .then(pl.lit("Confusion"))
            .when(connected_disposition_source_upper.str.contains("PAYMENT DISPUTE"))
            .then(pl.lit("Payment Dispute"))
            .when(connected_disposition_source_upper.str.contains("INSURANCE"))
            .then(pl.lit("Insurance Dispute"))
            .when(
                connected_disposition_source_upper.str.contains("DISPUTE") |
                (connected_disposition_source_lower == "dispute_raised")
            )
            .then(pl.lit("Dispute"))
            .when(
                connected_disposition_source_upper.str.contains("FOLLOW_UP_REQUIRED") |
                (connected_disposition_source_lower == "follow_up_required")
            )
            .then(pl.lit("Follow-up Required"))
            .when(connected_disposition_source_lower.is_in(["callback", "call back"]))
            .then(pl.lit("Call Back"))
            .when(connected_disposition_source_lower == "inquiry")
            .then(pl.lit("Inquiry"))
            .when(connected_disposition_source_lower == "payment inquiry")
            .then(pl.lit("Payment Inquiry"))
            .when(connected_disposition_source_lower == "overdue inquiry")
            .then(pl.lit("Overdue Inquiry"))
            .when(connected_disposition_source_lower == "payment difficulty")
            .then(pl.lit("Payment Difficulty"))
            .when(connected_disposition_source_lower == "payment arrangement")
            .then(pl.lit("Payment Arrangement"))
            .when(connected_disposition_source_lower == "greeting/confirmation")
            .then(pl.lit("Greeting/Confirmation"))
            .when(connected_disposition_source_lower == "family/third party")
            .then(pl.lit("Family/Third Party"))
            .when(connected_disposition_source_lower == "contact verification")
            .then(pl.lit("Contact Verification"))
            .when(connected_disposition_source_lower == "no request")
            .then(pl.lit("No Request"))
            .when(connected_disposition_source_lower == "unclear intent")
            .then(pl.lit("Unclear Intent"))
            .when(connected_disposition_source_lower == "service request")
            .then(pl.lit("Service Request"))
            .when(connected_disposition_source_lower == "complaint/escalation")
            .then(pl.lit("Complaint/Escalation"))
            .when(connected_disposition_source_lower == "account issue")
            .then(pl.lit("Account Issue"))
            .when(
                connected_disposition_source_upper.str.contains("TOKEN") &
                (
                    connected_disposition_source_upper.str.contains("PTP") |
                    connected_disposition_source_upper.str.contains("PROMISE") |
                    connected_disposition_source_upper.str.contains("PENDING")
                )
            )
            .then(pl.lit("PTP - Token Amount"))
            .when(ptp_settlement_cond)
            .then(pl.lit("PTP (Promise to Pay) - Settlement"))
            .when(other_settlement_cond)
            .then(pl.lit("Other"))
            .when(
                connected_disposition_source_lower.is_in(["futurepay", "ready_to_pay", "willing_to_pay"]) |
                connected_disposition_source_upper.str.contains("WILL PAY AFTER SOMETIME")
            )
            .then(pl.lit("PTP"))
            .when(
                connected_disposition_source_upper.str.contains("PENDING") |
                connected_disposition_source_upper.str.contains("PTP") |
                connected_disposition_source_upper.str.contains("PROMISE") |
                connected_disposition_source_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                connected_disposition_source_upper.str.contains("NO EARLY PAYMENT") |
                connected_disposition_source_upper.str.contains("CONFIRM") |
                (connected_disposition_source_lower == "positive")
            )
            .then(pl.lit("PTP"))
            .when(connected_disposition_source_lower.is_in(["wrongnumber", "wrong number", "wrongnumber "]))
            .then(pl.lit("Wrong Contact"))
            .when(
                connected_disposition_source_upper.str.contains("WRONG CONTACT") |
                connected_disposition_source_upper.str.contains("WRONG NUMBER")
            )
            .then(pl.lit("Wrong Contact"))
            .when(
                connected_disposition_source_upper.str.contains("UNAWARE") |
                connected_disposition_source_upper.str.contains("COMMUNICATION ISSUE") |
                connected_disposition_source_upper.str.contains("LANGUAGE")
            )
            .then(pl.lit("Unaware / Communication Issue"))
            .when(
                connected_disposition_source_upper.str.contains("FAMILY ISSUE") |
                connected_disposition_source_upper.str.contains("HEALTH")
            )
            .then(pl.lit("Health / Family Issue"))
            .when(connected_disposition_source_lower.is_in(["collectionofficer missing", "collection officer missing"]))
            .then(pl.lit("Collection Officer Missing"))
            .when(connected_disposition_source_upper.str.contains("OTHER"))
            .then(pl.lit("Other"))
            .when(connected_disposition_source_upper.str.contains("NOISY ENV"))
            .then(pl.lit("Noisy Env"))
            .when(
                connected_disposition_source_upper.str.contains("INSUFFICIENT CALL DURATION") |
                connected_disposition_source_upper.str.contains("CALL TOO SHORT")
            )
            .then(pl.lit("Call too short for categorization"))
            .when(connected_disposition_source_lower == "neutral")
            .then(pl.lit("Neutral"))
            .otherwise(
                pl.when(connected_disposition_source_trimmed == "")
                .then(pl.lit("No Response"))
                .otherwise(connected_disposition_source_trimmed.str.replace_all("(?i)busy", "Busy"))
            )
        )

        # ----------------------------------------------------
        # NON-CONNECTED CASE EXPRESSION
        # ----------------------------------------------------
        non_connected_expr = (
            pl.when(
                extracolumn1_upper.str.contains("DISCONNECTED") |
                extracolumn1_upper.str.contains("HANG UP") |
                extracolumn1_upper.str.contains("HANGUP")
            )
            .then(pl.lit("Hang Up / No Response"))
            .when(
                extracolumn1_compact.is_in([
                    'NETWORKCONGESTION', 'NODOESNOTEXIST', 'ONLYRINGING',
                    'SWITCHEDOFF', 'SPEAKINGTOSOMEONEELSE', 'USEROUTOFFCOVERAGE',
                    'CALLREJECTED', 'OUTOFSERVICE', 'INCOMINGNOTAVAILABLE', 'BUSY'
                ])
            )
            .then(
                pl.when(extracolumn1_compact == 'NETWORKCONGESTION').then(pl.lit('Network Congestion'))
                .when(extracolumn1_compact == 'NODOESNOTEXIST').then(pl.lit('Number Does Not Exist'))
                .when(extracolumn1_compact == 'ONLYRINGING').then(pl.lit('Only Ringing'))
                .when(extracolumn1_compact == 'SWITCHEDOFF').then(pl.lit('Switched Off'))
                .when(extracolumn1_compact == 'SPEAKINGTOSOMEONEELSE').then(pl.lit('Speaking To Someone Else'))
                .when(extracolumn1_compact == 'USEROUTOFFCOVERAGE').then(pl.lit('User Out Of Coverage'))
                .when(extracolumn1_compact == 'CALLREJECTED').then(pl.lit('Call Rejected'))
                .when(extracolumn1_compact == 'OUTOFSERVICE').then(pl.lit('Out Of Service'))
                .when(extracolumn1_compact == 'INCOMINGNOTAVAILABLE').then(pl.lit('Incoming Not Available'))
                .when(extracolumn1_compact == 'BUSY').then(pl.lit('Busy'))
                .otherwise(pl.lit('Not Contactable'))
            )
            .when(
                extracolumn1_upper.str.contains("FAILED") |
                extracolumn1_upper.str.contains("CONGESTION")
            )
            .then(pl.lit("No Answer"))
            .when(
                extracolumn1_upper.str.contains("SWITCHED OFF") |
                extracolumn1_upper.str.contains("OUT OF SERVICE") |
                extracolumn1_upper.str.contains("NOT REACHABLE")
            )
            .then(pl.lit("Not Reachable / Out of Network"))
            .when(extracolumn1_upper.str.contains("BUSY"))
            .then(pl.lit("Busy"))
            .when(
                extracolumn1_upper.str.contains("NO COMM") |
                extracolumn1_upper.str.contains("ANSWERED")
            )
            .then(pl.lit("No Answer"))
            .when(
                extracolumn1_upper.str.contains("DENIED") |
                extracolumn1_upper.str.contains("DENIES") |
                extracolumn1_upper.str.contains("DENIAL") |
                extracolumn1_upper.str.contains("REFUSED") |
                extracolumn1_upper.str.contains("NOT READY TO PAY")
            )
            .then(pl.lit("Refused to pay"))
            .when(extracolumn1_upper == "INQUIRY")
            .then(pl.lit("Inquiry"))
            .when(extracolumn1_upper == "PAYMENT INQUIRY")
            .then(pl.lit("Payment Inquiry"))
            .when(extracolumn1_upper == "OVERDUE INQUIRY")
            .then(pl.lit("Overdue Inquiry"))
            .when(extracolumn1_upper == "PAYMENT DIFFICULTY")
            .then(pl.lit("Payment Difficulty"))
            .when(extracolumn1_upper == "PAYMENT ARRANGEMENT")
            .then(pl.lit("Payment Arrangement"))
            .when(extracolumn1_upper == "GREETING/CONFIRMATION")
            .then(pl.lit("Greeting/Confirmation"))
            .when(extracolumn1_upper == "FAMILY/THIRD PARTY")
            .then(pl.lit("Family/Third Party"))
            .when(extracolumn1_upper == "CONTACT VERIFICATION")
            .then(pl.lit("Contact Verification"))
            .when(extracolumn1_upper == "NO REQUEST")
            .then(pl.lit("No Request"))
            .when(extracolumn1_upper == "UNCLEAR INTENT")
            .then(pl.lit("Unclear Intent"))
            .when(extracolumn1_upper == "SERVICE REQUEST")
            .then(pl.lit("Service Request"))
            .when(extracolumn1_upper == "COMPLAINT/ESCALATION")
            .then(pl.lit("Complaint/Escalation"))
            .when(extracolumn1_upper == "ACCOUNT ISSUE")
            .then(pl.lit("Account Issue"))
            .when(
                extracolumn1_upper.str.contains("TOKEN") &
                (
                    extracolumn1_upper.str.contains("PTP") |
                    extracolumn1_upper.str.contains("PROMISE") |
                    extracolumn1_upper.str.contains("PENDING")
                )
            )
            .then(pl.lit("PTP - Token Amount"))
            .when(
                extracolumn1_upper.str.contains("SETTLEMENT") &
                (
                    extracolumn1_upper.str.contains("PTP") |
                    extracolumn1_upper.str.contains("PROMISE") |
                    extracolumn1_upper.str.contains("PENDING")
                )
            )
            .then(pl.lit("PTP (Promise to Pay) - Settlement"))
            .when(
                extracolumn1_upper.str.contains("PENDING") |
                extracolumn1_upper.str.contains("PTP") |
                extracolumn1_upper.str.contains("PROMISE") |
                extracolumn1_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                extracolumn1_upper.str.contains("NO EARLY PAYMENT") |
                extracolumn1_upper.str.contains("CONFIRM")
            )
            .then(pl.lit("PTP"))
            .when(
                extracolumn1_upper.str.contains("WRONG CONTACT") |
                extracolumn1_upper.str.contains("WRONG NUMBER")
            )
            .then(pl.lit("Wrong Contact"))
            .when(
                extracolumn1_upper.str.contains("UNAWARE") |
                extracolumn1_upper.str.contains("COMMUNICATION ISSUE") |
                extracolumn1_upper.str.contains("LANGUAGE")
            )
            .then(pl.lit("Unaware / Communication Issue"))
            .when(
                extracolumn1_upper.str.contains("FAMILY ISSUE") |
                extracolumn1_upper.str.contains("HEALTH")
            )
            .then(pl.lit("Health / Family Issue"))
            .when(extracolumn1_upper.str.contains("OTHER"))
            .then(pl.lit("Other"))
            .when(extracolumn1_upper.str.contains("NOISY ENV"))
            .then(pl.lit("Noisy Env"))
            .when(
                extracolumn1_upper.str.contains("INSUFFICIENT CALL DURATION") |
                extracolumn1_upper.str.contains("CALL TOO SHORT")
            )
            .then(pl.lit("Call too short for categorization"))
            .otherwise(
                pl.when(extracolumn1_trimmed == "")
                .then(pl.lit("No Answer"))
                .otherwise(extracolumn1_trimmed.str.replace_all("(?i)busy", "Busy"))
            )
        )

        # Apply connected / non-connected switch
        case_disposition = (
            pl.when(pl.col("is_connected"))
            .then(connected_expr)
            .otherwise(non_connected_expr)
        )

        # Outer COALESCE clean structure
        disposition_raw = pl.coalesce([
            case_disposition,
            extracolumn1_trimmed,
            pl.when(pl.col("is_connected"))
            .then(pl.lit("No Response"))
            .otherwise(pl.lit("No Communication"))
        ])

        # Clean trailing colons and whitespaces
        disposition_cleaned = disposition_raw.str.strip_chars().str.replace_all(r":+[ \t]*$", "")

        call_dp = classified.with_columns(
            disposition_cleaned.alias("disposition_cleaned")
        )
        self.logger.info("Validation Log - Rows after disposition mapping: %d", len(call_dp))

        # Join disposition alias map and ranking map using business_rules.py constants
        from . import business_rules
        
        alias_df = pl.DataFrame({
            "raw_disposition": [a[0] for a in business_rules.DISPOSITION_ALIASES],
            "normalized_disposition": [a[1] for a in business_rules.DISPOSITION_ALIASES],
            "ranking_disposition": [a[2] for a in business_rules.DISPOSITION_ALIASES],
        })
        alias_df = alias_df.with_columns(
            pl.col("raw_disposition").str.strip_chars().str.to_uppercase()
        ).unique(subset=["raw_disposition"], keep="first")

        ranking_df = pl.DataFrame({
            "disposition": [r[0] for r in business_rules.DISPOSITION_RANKING],
            "rank_val": [r[1] for r in business_rules.DISPOSITION_RANKING],
            "category": [r[2] for r in business_rules.DISPOSITION_RANKING],
        }, schema={"disposition": pl.String, "rank_val": pl.Float64, "category": pl.String})
        ranking_df = ranking_df.with_columns(
            pl.col("disposition").str.strip_chars().str.to_uppercase().alias("ranking_disp_upper")
        ).sort(by=["ranking_disp_upper", "rank_val", "category"]).unique(subset=["ranking_disp_upper"], keep="first")

        # Map to uppercase keys for robust joins
        call_dp = call_dp.with_columns(
            pl.col("disposition_cleaned").str.strip_chars().str.to_uppercase().alias("cleaned_upper")
        )
        
        # 1st join: Alias map
        call_dp = call_dp.join(alias_df, left_on="cleaned_upper", right_on="raw_disposition", how="left")

        # Prep ranking key
        ranking_key = (
            pl.when(pl.col("disposition_cleaned") == "PTP - Date")
            .then(pl.lit("PTP Date"))
            .otherwise(pl.coalesce([pl.col("ranking_disposition"), pl.col("disposition_cleaned")]))
            .str.strip_chars()
            .str.to_uppercase()
        )
        call_dp = call_dp.with_columns(ranking_key.alias("ranking_key_upper"))

        # 2nd join: Ranking map
        call_dp = call_dp.join(ranking_df, left_on="ranking_key_upper", right_on="ranking_disp_upper", how="left")
        self.logger.info("Validation Log - Rows after ranking mapping: %d", len(call_dp))

        # Final columns derivation
        final_disposition_expr = pl.coalesce([
            pl.when(pl.col("disposition_cleaned") == "PTP - Date")
            .then(pl.col("disposition_cleaned"))
            .otherwise(pl.col("normalized_disposition")),
            pl.col("disposition"),
            pl.when(pl.col("is_connected"))
            .then(pl.lit("Other"))
            .otherwise(pl.lit("No Answer"))
        ])

        ptp_list = [
            'PTP', 'PTP - Date', 'PTP - Token Amount',
            'PTP (Promise to Pay) - Settlement', 'Promise to Pay',
            'Pay Later', 'Pending', 'PTP Date'
        ]

        is_ptp_source_expr = (
            pl.when(pl.col("disposition_cleaned").is_in(ptp_list))
            .then(pl.lit(1))
            .otherwise(pl.lit(0))
        )

        # final projection and clean alias naming
        extracolumn1_col = "Extracolumn1" if "Extracolumn1" in call_dp.columns else "extracolumn1"
        queue_col = "VoiceBotQueueID" if "VoiceBotQueueID" in call_dp.columns else "voicebotqueueid"
        
        self.all_call_dispositions_df = call_dp.select([
            pl.col("LoanMstID"),
            pl.col("call_date"),
            pl.col(queue_col).alias("VoiceBotQueueID"),
            pl.col(extracolumn1_col).alias("source_extracolumn1"),
            final_disposition_expr.alias("disposition"),
            pl.lit(None).cast(pl.Date).alias("disposition_ptp_date"), # promise_datetime is not integrated yet
            is_ptp_source_expr.alias("is_ptp_source"),
            pl.col("is_connected").cast(pl.Int32).alias("connected_flag"),
            pl.col("rank_val").fill_null(999).alias("rank_val"),
            pl.coalesce([
                pl.col("category"),
                pl.when(pl.col("is_connected"))
                .then(pl.lit("Not Categorized"))
                .otherwise(pl.lit("Not Contactable"))
            ]).alias("category"),
            pl.lit(1).alias("source_priority")
        ])

        # Logging validation stats
        row_count = len(self.all_call_dispositions_df)
        self.logger.info("Validation Log - Final projection rows: %d", row_count)
        self.logger.info("Final DataFrame Shape: %s", self.all_call_dispositions_df.shape)
        self.logger.info("Final DataFrame Columns: %s", self.all_call_dispositions_df.columns)
        self.logger.info("Unique LoanMstID: %d", self.all_call_dispositions_df["LoanMstID"].n_unique())
        self.logger.info("Unique VoiceBotQueueID: %d", self.all_call_dispositions_df["VoiceBotQueueID"].n_unique())
        self.logger.info("Null counts per column: %s", self.all_call_dispositions_df.null_count().to_dicts()[0])
        
        # Store MTD base
        self.voicebot_mtd_base_df = classified
        
        # Build VoiceBot response map
        if self.response_df is None:
            self.response_df = pl.DataFrame(schema={
                "VoiceBotQueueID": pl.Int64, "WhatsappQueueID": pl.Int64, 
                "Status": pl.String, "PromiseDateTime": pl.Datetime, 
                "ResponseDateTime": pl.Datetime, "BankMstID": pl.Int64
            })
        
        duration_col_v = "CallDuration" if "CallDuration" in self.voicebot_mtd_base_df.columns else "callduration"
        max_dur_lf = (
            self.voicebot_mtd_base_df
            .lazy()
            .group_by("VoiceBotQueueID")
            .agg(pl.col(duration_col_v).fill_null(0).max().alias("max_call_duration"))
            .filter(pl.col("max_call_duration") > 15)
        )
        
        self.voicebot_response_map_df = (
            self.response_df
            .lazy()
            .join(max_dur_lf, on="VoiceBotQueueID", how="inner")
            .group_by("VoiceBotQueueID")
            .agg([
                pl.when(pl.col("Status").is_in(["Already Paid", "Claim_Expired"])).then(pl.lit(1)).otherwise(pl.lit(0)).max().alias("has_paid"),
                pl.when(pl.col("Status") == "Denied").then(pl.lit(1)).otherwise(pl.lit(0)).max().alias("has_denied"),
                pl.when(pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"])).then(pl.lit(1)).otherwise(pl.lit(0)).max().alias("has_ptp"),
                pl.col("PromiseDateTime").filter(pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"])).max().alias("promise_datetime")
            ])
            .collect()
        )

        # Memory profiling
        mem_bytes = self.all_call_dispositions_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)

    @timer
    def build_all_wa_dispositions(self) -> None:
        """
        Processes and maps WhatsApp messages into a unified disposition engine state.
        Stores the result as self.all_wa_dispositions_df.
        """
        self.logger.info("Starting WhatsApp disposition processing (all_wa_dispositions)...")
        if self.whatsapp_df is None:
            raise ValueError("WhatsApp data is not loaded. Call load_source_tables first.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded. Call build_base_accounts first.")

        # Log loaded WhatsApp count
        self.logger.info("Validation Log - Loaded WhatsApp rows: %d", len(self.whatsapp_df))

        # Join with base_accounts (SQL: JOIN base_accounts ba ON ba."LoanMstID" = wah."LoanMstID")
        # and filter (SQL: WHERE wah."BankMstID" = 53 AND wah."IsSent" = TRUE)
        is_sent_col = "IsSent" if "IsSent" in self.whatsapp_df.columns else "issent"
        bank_mst_col = "BankMstID" if "BankMstID" in self.whatsapp_df.columns else "bankmstid"
        
        joined = (
            self.whatsapp_df
            .join(
                self.base_accounts_df.select(["LoanMstID"]),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (pl.col(bank_mst_col) == 53) &
                (pl.col(is_sent_col) == True)
            )
        )
        self.logger.info("Validation Log - Rows after join with base_accounts: %d", len(joined))

        # Derive wa_date
        created_date_col = "CreatedDate" if "CreatedDate" in joined.columns else "createddate"
        joined = joined.with_columns(
            pl.col(created_date_col).cast(pl.Date).alias("wa_date")
        )

        # Define dummy/mock Response & Messages variables (out of scope for Phase 4B)
        wrm_has_paid = pl.lit(0)
        wrm_has_denied = pl.lit(0)
        wrm_has_ptp = pl.lit(0)
        wrm_has_any_response = pl.lit(0)
        wrm_promise_datetime = pl.lit(None).cast(pl.Datetime)
        received_after_send = pl.lit(False)

        ext_col2 = "ExtraColumn2" if "ExtraColumn2" in joined.columns else "extracolumn2"
        ext_col2_trimmed = pl.col(ext_col2).str.strip_chars().fill_null("")
        ext_col2_upper = ext_col2_trimmed.str.to_uppercase()
        
        # compact expression
        ext_col2_compact = (
            ext_col2_upper
            .str.replace_all(r"^[0-9]+[_ \t-]*", "")
            .str.replace_all(r"[^A-Z0-9]", "")
        )

        # connected_flag computation
        connected_flag_expr = (
            pl.when(
                (wrm_has_paid == 1) |
                (wrm_has_denied == 1) |
                (wrm_has_ptp == 1) |
                (wrm_has_any_response == 1) |
                received_after_send |
                (pl.col(ext_col2).is_not_null() & (pl.col(ext_col2).str.strip_chars() != ""))
            )
            .then(pl.lit(1))
            .otherwise(pl.lit(0))
        )

        # STEP 1: wa_dp CASE statement logic
        wa_dp_expr = (
            pl.when(wrm_has_paid == 1).then(pl.lit("Already Paid"))
            .when(wrm_has_denied == 1).then(pl.lit("Refused to pay"))
            .when(wrm_has_ptp == 1).then(pl.lit("PTP - WhatsApp"))
            .when(ext_col2_upper.str.contains("ALREADY PAID")).then(pl.lit("Already Paid"))
            .when(ext_col2_upper.str.contains("DENIED")).then(pl.lit("Refused to pay"))
            .when(ext_col2_upper.str.contains("NO COMM")).then(pl.lit("No Response"))
            .when(
                ext_col2_upper.str.contains("TOKEN") &
                (
                    ext_col2_upper.str.contains("PTP") |
                    ext_col2_upper.str.contains("PROMISE") |
                    ext_col2_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP Whatsapp - Free Text"))
            .when(
                ext_col2_upper.str.contains("SETTLEMENT") &
                (
                    ext_col2_upper.str.contains("PTP") |
                    ext_col2_upper.str.contains("PROMISE") |
                    ext_col2_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP Whatsapp - Free Text"))
            .when(
                ext_col2_upper.str.contains("PTP") |
                ext_col2_upper.str.contains("PENDING") |
                ext_col2_upper.str.contains("PROMISE") |
                ext_col2_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                ext_col2_upper.str.contains("NO EARLY PAYMENT") |
                ext_col2_upper.str.contains("CONFIRM")
            ).then(pl.lit("PTP Whatsapp - Free Text"))
            .when(ext_col2_upper.str.contains("CONFUSION")).then(pl.lit("Confusion"))
            .when(ext_col2_upper.str.contains("PAYMENT DISPUTE")).then(pl.lit("Payment Dispute"))
            .when(ext_col2_upper.str.contains("INSURANCE")).then(pl.lit("Insurance Dispute"))
            .when(ext_col2_upper.str.contains("DISPUTE")).then(pl.lit("Dispute"))
            .when(
                ext_col2_upper.str.contains("WRONG CONTACT") |
                ext_col2_upper.str.contains("WRONG NUMBER")
            ).then(pl.lit("Wrong Contact"))
            .when(
                ext_col2_upper.str.contains("UNAWARE") |
                ext_col2_upper.str.contains("COMMUNICATION ISSUE") |
                ext_col2_upper.str.contains("LANGUAGE")
            ).then(pl.lit("Unaware / Communication Issue"))
            .when(
                ext_col2_upper.str.contains("FAMILY ISSUE") |
                ext_col2_upper.str.contains("HEALTH")
            ).then(pl.lit("Health / Family Issue"))
            .when(ext_col2_upper.str.contains("NOISY ENV")).then(pl.lit("Noisy Env"))
            .when(
                ext_col2_upper.str.contains("INSUFFICIENT CALL DURATION") |
                ext_col2_upper.str.contains("CALL TOO SHORT")
            ).then(pl.lit("Call too short for categorization"))
            .when(ext_col2_upper.str.contains("OTHER")).then(pl.lit("Other"))
            .when(
                ext_col2_upper.str.contains("HANG UP") |
                ext_col2_upper.str.contains("HANGUP") |
                ext_col2_upper.str.contains("DISCONNECTED")
            ).then(pl.lit("Hang Up / No Response"))
            .when(
                ext_col2_compact.is_in([
                    'NETWORKCONGESTION', 'NODOESNOTEXIST', 'ONLYRINGING',
                    'SWITCHEDOFF', 'SPEAKINGTOSOMEONEELSE', 'USEROUTOFFCOVERAGE',
                    'CALLREJECTED', 'OUTOFSERVICE', 'INCOMINGNOTAVAILABLE', 'BUSY'
                ])
            ).then(
                pl.when(ext_col2_compact == 'NETWORKCONGESTION').then(pl.lit('Network Congestion'))
                .when(ext_col2_compact == 'NODOESNOTEXIST').then(pl.lit('Number Does Not Exist'))
                .when(ext_col2_compact == 'ONLYRINGING').then(pl.lit('Only Ringing'))
                .when(ext_col2_compact == 'SWITCHEDOFF').then(pl.lit('Switched Off'))
                .when(ext_col2_compact == 'SPEAKINGTOSOMEONEELSE').then(pl.lit('Speaking To Someone Else'))
                .when(ext_col2_compact == 'USEROUTOFFCOVERAGE').then(pl.lit('User Out Of Coverage'))
                .when(ext_col2_compact == 'CALLREJECTED').then(pl.lit('Call Rejected'))
                .when(ext_col2_compact == 'OUTOFSERVICE').then(pl.lit('Out Of Service'))
                .when(ext_col2_compact == 'INCOMINGNOTAVAILABLE').then(pl.lit('Incoming Not Available'))
                .when(ext_col2_compact == 'BUSY').then(pl.lit('Busy'))
                .otherwise(pl.lit('Not Contactable'))
            )
            .when(
                ext_col2_upper.str.contains("FAILED") |
                ext_col2_upper.str.contains("CONGESTION")
            ).then(pl.lit("No Answer"))
            .when(ext_col2_upper.str.contains("BUSY")).then(pl.lit("Busy"))
            .when(
                ext_col2_upper.str.contains("SWITCHED OFF") |
                ext_col2_upper.str.contains("OUT OF SERVICE") |
                ext_col2_upper.str.contains("NOT REACHABLE")
            ).then(pl.lit("Not Reachable / Out of Network"))
            .when(pl.col(ext_col2).is_not_null() & (pl.col(ext_col2).str.strip_chars() != "")).then(pl.col(ext_col2))
            .when((wrm_has_any_response == 1) | received_after_send).then(pl.lit("Not Categorized"))
            .otherwise(pl.lit("No Answer"))
        )

        wa_dp_df = joined.with_columns([
            connected_flag_expr.cast(pl.Int32).alias("connected_flag"),
            wa_dp_expr.alias("disposition")
        ])
        self.logger.info("Validation Log - Rows after wa_dp: %d", len(wa_dp_df))

        # STEP 2: wa_classified (clean disposition, join with alias & ranking maps)
        disposition_cleaned = pl.col("disposition").str.strip_chars().str.replace_all(r":+[ \t]*$", "")
        
        wa_classified_pre = wa_dp_df.with_columns(
            disposition_cleaned.alias("disposition_cleaned")
        )
        self.logger.info("Validation Log - Rows after wa_classified: %d", len(wa_classified_pre))

        from . import business_rules

        alias_df = pl.DataFrame({
            "raw_disposition": [a[0] for a in business_rules.DISPOSITION_ALIASES],
            "normalized_disposition": [a[1] for a in business_rules.DISPOSITION_ALIASES],
            "ranking_disposition": [a[2] for a in business_rules.DISPOSITION_ALIASES],
        })
        alias_df = alias_df.with_columns(
            pl.col("raw_disposition").str.strip_chars().str.to_uppercase()
        ).unique(subset=["raw_disposition"], keep="first")

        ranking_df = pl.DataFrame({
            "disposition": [r[0] for r in business_rules.DISPOSITION_RANKING],
            "rank_val": [r[1] for r in business_rules.DISPOSITION_RANKING],
            "category": [r[2] for r in business_rules.DISPOSITION_RANKING],
        }, schema={"disposition": pl.String, "rank_val": pl.Float64, "category": pl.String})
        ranking_df = ranking_df.with_columns(
            pl.col("disposition").str.strip_chars().str.to_uppercase().alias("ranking_disp_upper")
        ).sort(by=["ranking_disp_upper", "rank_val", "category"]).unique(subset=["ranking_disp_upper"], keep="first")

        # Join alias map
        wa_classified_pre = wa_classified_pre.with_columns(
            pl.col("disposition_cleaned").str.strip_chars().str.to_uppercase().alias("cleaned_upper")
        )
        classified_joined = wa_classified_pre.join(
            alias_df,
            left_on="cleaned_upper",
            right_on="raw_disposition",
            how="left"
        )

        # Derive ranking_disposition column: COALESCE(dam.ranking_disposition, disposition_cleaned)
        ranking_disposition_expr = pl.coalesce([
            pl.col("ranking_disposition"),
            pl.col("disposition_cleaned")
        ])
        classified_joined = classified_joined.with_columns(
            ranking_disposition_expr.alias("ranking_disposition")
        )

        # Join ranking map on ranking_disposition
        classified_joined = classified_joined.with_columns(
            pl.col("ranking_disposition").str.strip_chars().str.to_uppercase().alias("ranking_key_upper")
        )
        ranking_joined = classified_joined.join(
            ranking_df,
            left_on="ranking_key_upper",
            right_on="ranking_disp_upper",
            how="left"
        )
        
        # Rename disposition from ranking_df
        ranking_joined = ranking_joined.rename({"disposition": "ranked_disposition"})
        self.logger.info("Validation Log - Rows after Ranking: %d", len(ranking_joined))

        # STEP 3: Final Selection
        norm_or_ranked = pl.coalesce([
            pl.when(pl.col("normalized_disposition") != "").then(pl.col("normalized_disposition")).otherwise(None),
            pl.when(pl.col("ranked_disposition") != "").then(pl.col("ranked_disposition")).otherwise(None)
        ])

        final_disposition_expr = (
            pl.when(pl.col("ranking_disposition") == "Denied")
            .then(pl.lit("Refused to Pay"))
            .when(pl.col("disposition_cleaned") == "PTP - WhatsApp")
            .then(pl.lit("PTP - WhatsApp"))
            .when(pl.col("disposition_cleaned") == "PTP Whatsapp - Free Text")
            .then(pl.lit("PTP Whatsapp - Free Text"))
            .when(norm_or_ranked.is_not_null())
            .then(norm_or_ranked)
            .when(pl.col("disposition_cleaned") != "")
            .then(pl.col("disposition_cleaned"))
            .when(pl.col("connected_flag") == 1)
            .then(pl.lit("Not Categorized"))
            .otherwise(pl.lit("No Answer"))
        )

        final_ptp_date_expr = (
            pl.when(pl.col("disposition_cleaned") == "PTP - WhatsApp")
            .then(wrm_promise_datetime.cast(pl.Date))
            .otherwise(pl.lit(None).cast(pl.Date))
        )

        final_rank_val_expr = (
            pl.when(pl.col("ranking_disposition") == "Denied")
            .then(pl.lit(7.0))
            .when(pl.col("disposition_cleaned").is_in(['PTP - WhatsApp', 'PTP Whatsapp - Free Text']))
            .then(pl.lit(1.0))
            .when(pl.col("rank_val").is_not_null())
            .then(pl.col("rank_val"))
            .when(pl.col("connected_flag") == 1)
            .then(pl.lit(9.0))
            .otherwise(pl.lit(10.9))
        )

        final_category_expr = (
            pl.when(pl.col("ranking_disposition") == "Denied")
            .then(pl.lit("Positive"))
            .when(pl.col("disposition_cleaned").is_in(['PTP - WhatsApp', 'PTP Whatsapp - Free Text']))
            .then(pl.lit("Positive"))
            .when(pl.col("category").is_not_null())
            .then(pl.col("category"))
            .when(pl.col("connected_flag") == 1)
            .then(pl.lit("Not Categorized"))
            .otherwise(pl.lit("Positive"))
        )

        # final projection
        queue_col = "WhatsAppQueueID" if "WhatsAppQueueID" in ranking_joined.columns else "whatsappqueueid"
        self.all_wa_dispositions_df = ranking_joined.select([
            pl.col("LoanMstID"),
            pl.col("wa_date"),
            final_disposition_expr.alias("disposition"),
            final_ptp_date_expr.alias("disposition_ptp_date"),
            final_rank_val_expr.alias("rank_val"),
            final_category_expr.alias("category"),
            pl.lit(2).alias("source_priority"),
            pl.col("connected_flag"),
            pl.col(queue_col).alias("WhatsAppQueueID")
        ])

        # Validation, profiling & logging
        row_count = len(self.all_wa_dispositions_df)
        self.logger.info("Validation Log - Rows after Final projection: %d", row_count)
        self.logger.info("Final WhatsApp DataFrame Shape: %s", self.all_wa_dispositions_df.shape)
        self.logger.info("Final WhatsApp DataFrame Columns: %s", self.all_wa_dispositions_df.columns)
        self.logger.info("Unique LoanMstID count: %d", self.all_wa_dispositions_df["LoanMstID"].n_unique())
        self.logger.info("Null counts per column: %s", self.all_wa_dispositions_df.null_count().to_dicts()[0])
        
        # Store MTD base
        self.whatsapp_mtd_base_df = joined
        
        # Build WhatsApp response map
        if self.response_df is None:
            self.response_df = pl.DataFrame(schema={
                "VoiceBotQueueID": pl.Int64, "WhatsappQueueID": pl.Int64, 
                "Status": pl.String, "PromiseDateTime": pl.Datetime, 
                "ResponseDateTime": pl.Datetime, "BankMstID": pl.Int64
            })
            
        self.whatsapp_response_map_df = (
            self.response_df
            .lazy()
            .join(
                self.whatsapp_mtd_base_df.select(["WhatsAppQueueID"]).lazy(),
                left_on="WhatsappQueueID",
                right_on="WhatsAppQueueID",
                how="inner"
            )
            .group_by("WhatsappQueueID")
            .agg([
                pl.lit(1).alias("has_any_response"),
                pl.when(pl.col("Status").is_in(["Already Paid", "Claim_Expired"])).then(pl.lit(1)).otherwise(pl.lit(0)).max().alias("has_paid"),
                pl.when(pl.col("Status") == "Denied").then(pl.lit(1)).otherwise(pl.lit(0)).max().alias("has_denied"),
                pl.when(pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"])).then(pl.lit(1)).otherwise(pl.lit(0)).max().alias("has_ptp"),
                pl.col("PromiseDateTime").filter(pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"])).max().alias("promise_datetime"),
                pl.col("ResponseDateTime").max().alias("latest_response_datetime")
            ])
            .collect()
        )

        # Memory profiling
        mem_bytes = self.all_wa_dispositions_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)

    @timer
    def build_all_blaster_dispositions(self) -> None:
        """
        Processes and maps Blaster call attempts into a unified disposition engine state.
        Stores the result as self.all_blaster_dispositions_df.
        """
        self.logger.info("Starting Blaster disposition processing (all_blaster_dispositions)...")
        if self.blaster_df is None:
            raise ValueError("Blaster data is not loaded. Call load_source_tables first.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded. Call build_base_accounts first.")

        # Log loaded Blaster count
        self.logger.info("Validation Log - Loaded Blaster rows: %d", len(self.blaster_df))

        # Join with base_accounts
        joined = (
            self.blaster_df
            .join(
                self.base_accounts_df.select(["LoanMstID"]),
                on="LoanMstID",
                how="inner"
            )
        )
        self.logger.info("Validation Log - Rows after join with base_accounts: %d", len(joined))

        # Apply SQL filters: BankMstID = 53 and CallTried > 0
        bank_mst_col = "BankMstID" if "BankMstID" in joined.columns else "bankmstid"
        call_tried_col = "CallTried" if "CallTried" in joined.columns else "calltried"
        
        filtered = joined.filter(
            (pl.col(bank_mst_col) == 53) &
            (pl.col(call_tried_col) > 0)
        )
        self.logger.info("Validation Log - Rows after filtering: %d", len(filtered))

        # Derive blaster_date
        created_date_col = "CreatedDate" if "CreatedDate" in filtered.columns else "createddate"
        filtered = filtered.with_columns(
            pl.col(created_date_col).cast(pl.Date).alias("blaster_date")
        )

        # Connection Classification: is_connected & call_duration
        connected_col = "CallConnected" if "CallConnected" in filtered.columns else "callconnected"
        duration_col = "CallDuration" if "CallDuration" in filtered.columns else "callduration"
        
        is_connected_expr = pl.col(connected_col).fill_null(False)
        call_duration_expr = (
            pl.when(pl.col(duration_col) == 0)
            .then(pl.lit(1.0))
            .otherwise(pl.col(duration_col).fill_null(0.0))
        )

        classified = filtered.with_columns([
            is_connected_expr.alias("is_connected"),
            call_duration_expr.alias("call_duration")
        ])
        classified = classified.with_columns(
            pl.col("is_connected").cast(pl.Int32).alias("connected_flag")
        )
        self.logger.info("Validation Log - Rows after classification: %d", len(classified))

        # Disposition Mapping (Translating CASE statements)
        ext_col1_col = "ExtraColumn1" if "ExtraColumn1" in classified.columns else "extracolumn1"
        ext_col1_trimmed = pl.col(ext_col1_col).str.strip_chars().fill_null("")
        ext_col1_upper = ext_col1_trimmed.str.to_uppercase()

        # Compact expression for non-connected
        ext_col1_compact = (
            ext_col1_upper
            .str.replace_all(r"^[0-9]+[_ \t-]*", "")
            .str.replace_all(r"[^A-Z0-9]", "")
        )

        # Connected CASE Expression
        connected_expr = (
            pl.when(ext_col1_trimmed == "").then(pl.lit("No Response"))
            .when(ext_col1_upper.str.contains("BLASTER COMPLETED")).then(pl.lit("Blaster Completed"))
            .when(ext_col1_upper.str.contains("ALREADY PAID")).then(pl.lit("Already Paid"))
            .when(
                ext_col1_upper.str.contains("DENIED") |
                ext_col1_upper.str.contains("DENIES") |
                ext_col1_upper.str.contains("DENIAL") |
                ext_col1_upper.str.contains("REFUSED")
            ).then(pl.lit("Refused to pay"))
            .when(ext_col1_upper.str.contains("CONFUSION")).then(pl.lit("Confusion"))
            .when(ext_col1_upper.str.contains("PAYMENT DISPUTE")).then(pl.lit("Payment Dispute"))
            .when(ext_col1_upper.str.contains("INSURANCE")).then(pl.lit("Insurance Dispute"))
            .when(ext_col1_upper.str.contains("DISPUTE")).then(pl.lit("Dispute"))
            .when(ext_col1_upper.str.contains("FOLLOW_UP_REQUIRED")).then(pl.lit("Follow-up Required"))
            .when(ext_col1_upper.str.contains("CALL BACK")).then(pl.lit("Call Back"))
            .when(
                ext_col1_upper.str.contains("TOKEN") &
                (
                    ext_col1_upper.str.contains("PTP") |
                    ext_col1_upper.str.contains("PROMISE") |
                    ext_col1_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP - Token Amount"))
            .when(
                ext_col1_upper.str.contains("SETTLEMENT") &
                (
                    ext_col1_upper.str.contains("PTP") |
                    ext_col1_upper.str.contains("PROMISE") |
                    ext_col1_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP (Promise to Pay) - Settlement"))
            .when(
                ext_col1_upper.str.contains("SETTLEMENT") &
                (pl.col("call_duration") < 60)
            ).then(pl.lit("Other"))
            .when(
                ext_col1_upper.str.contains("PENDING") |
                ext_col1_upper.str.contains("PTP") |
                ext_col1_upper.str.contains("PROMISE") |
                ext_col1_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                ext_col1_upper.str.contains("NO EARLY PAYMENT") |
                ext_col1_upper.str.contains("CONFIRM")
            ).then(pl.lit("PTP"))
            .when(
                ext_col1_upper.str.contains("WRONG CONTACT") |
                ext_col1_upper.str.contains("WRONG NUMBER")
            ).then(pl.lit("Wrong Contact"))
            .when(
                ext_col1_upper.str.contains("UNAWARE") |
                ext_col1_upper.str.contains("COMMUNICATION ISSUE") |
                ext_col1_upper.str.contains("LANGUAGE")
            ).then(pl.lit("Unaware / Communication Issue"))
            .when(
                ext_col1_upper.str.contains("FAMILY ISSUE") |
                ext_col1_upper.str.contains("HEALTH")
            ).then(pl.lit("Health / Family Issue"))
            .when(ext_col1_upper.str.contains("OTHER")).then(pl.lit("Other"))
            .otherwise(ext_col1_trimmed.str.replace_all("(?i)busy", "Busy"))
        )

        # Non-connected CASE Expression
        non_connected_expr = (
            pl.when(ext_col1_trimmed == "").then(pl.lit("No Communication"))
            .when(
                ext_col1_upper.str.contains("DISCONNECTED") |
                ext_col1_upper.str.contains("HANG UP") |
                ext_col1_upper.str.contains("HANGUP")
            ).then(pl.lit("Hang Up / No Response"))
            .when(
                ext_col1_compact.is_in([
                    'NETWORKCONGESTION', 'NODOESNOTEXIST', 'ONLYRINGING',
                    'SWITCHEDOFF', 'SPEAKINGTOSOMEONEELSE', 'USEROUTOFFCOVERAGE',
                    'CALLREJECTED', 'OUTOFSERVICE', 'INCOMINGNOTAVAILABLE', 'BUSY'
                ])
            ).then(
                pl.when(ext_col1_compact == 'NETWORKCONGESTION').then(pl.lit('Network Congestion'))
                .when(ext_col1_compact == 'NODOESNOTEXIST').then(pl.lit('Number Does Not Exist'))
                .when(ext_col1_compact == 'ONLYRINGING').then(pl.lit('Only Ringing'))
                .when(ext_col1_compact == 'SWITCHEDOFF').then(pl.lit('Switched Off'))
                .when(ext_col1_compact == 'SPEAKINGTOSOMEONEELSE').then(pl.lit('Speaking To Someone Else'))
                .when(ext_col1_compact == 'USEROUTOFFCOVERAGE').then(pl.lit('User Out Of Coverage'))
                .when(ext_col1_compact == 'CALLREJECTED').then(pl.lit('Call Rejected'))
                .when(ext_col1_compact == 'OUTOFSERVICE').then(pl.lit('Out Of Service'))
                .when(ext_col1_compact == 'INCOMINGNOTAVAILABLE').then(pl.lit('Incoming Not Available'))
                .when(ext_col1_compact == 'BUSY').then(pl.lit('Busy'))
                .otherwise(pl.lit('Not Contactable'))
            )
            .when(
                ext_col1_upper.str.contains("FAILED") |
                ext_col1_upper.str.contains("CONGESTION")
            ).then(pl.lit("No Answer"))
            .when(
                ext_col1_upper.str.contains("SWITCHED OFF") |
                ext_col1_upper.str.contains("OUT OF SERVICE") |
                ext_col1_upper.str.contains("NOT REACHABLE")
            ).then(pl.lit("Not Reachable / Out of Network"))
            .when(ext_col1_upper.str.contains("BUSY")).then(pl.lit("Busy"))
            .when(
                ext_col1_upper.str.contains("NO COMM") |
                ext_col1_upper.str.contains("ANSWERED")
            ).then(pl.lit("No Answer"))
            .when(
                ext_col1_upper.str.contains("DENIED") |
                ext_col1_upper.str.contains("DENIES") |
                ext_col1_upper.str.contains("DENIAL") |
                ext_col1_upper.str.contains("REFUSED")
            ).then(pl.lit("Refused to pay"))
            .when(
                ext_col1_upper.str.contains("PENDING") |
                ext_col1_upper.str.contains("PTP") |
                ext_col1_upper.str.contains("PROMISE") |
                ext_col1_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                ext_col1_upper.str.contains("NO EARLY PAYMENT") |
                ext_col1_upper.str.contains("CONFIRM")
            ).then(pl.lit("PTP"))
            .when(
                ext_col1_upper.str.contains("WRONG CONTACT") |
                ext_col1_upper.str.contains("WRONG NUMBER")
            ).then(pl.lit("Wrong Contact"))
            .when(
                ext_col1_upper.str.contains("UNAWARE") |
                ext_col1_upper.str.contains("COMMUNICATION ISSUE") |
                ext_col1_upper.str.contains("LANGUAGE")
            ).then(pl.lit("Unaware / Communication Issue"))
            .when(ext_col1_upper.str.contains("OTHER")).then(pl.lit("Other"))
            .otherwise(
                pl.when(pl.col(ext_col1_col).is_null())
                .then(pl.lit("No Communication"))
                .otherwise(ext_col1_trimmed.str.replace_all("(?i)busy", "Busy"))
            )
        )

        # Switch CASE on is_connected
        case_disposition = (
            pl.when(pl.col("is_connected"))
            .then(connected_expr)
            .otherwise(non_connected_expr)
        )

        # Outer COALESCE
        disposition_raw = pl.coalesce([
            case_disposition,
            ext_col1_trimmed,
            pl.when(pl.col("is_connected"))
            .then(pl.lit("No Response"))
            .otherwise(pl.lit("No Communication"))
        ])

        # Clean trailing colons and whitespaces
        disposition_cleaned = disposition_raw.str.strip_chars().str.replace_all(r":+[ \t]*$", "")

        call_dp = classified.with_columns(
            disposition_cleaned.alias("disposition_cleaned")
        )
        self.logger.info("Validation Log - Rows after disposition cleaned: %d", len(call_dp))

        # Join disposition alias map and ranking map using business_rules.py constants
        from . import business_rules

        alias_df = pl.DataFrame({
            "raw_disposition": [a[0] for a in business_rules.DISPOSITION_ALIASES],
            "normalized_disposition": [a[1] for a in business_rules.DISPOSITION_ALIASES],
            "ranking_disposition": [a[2] for a in business_rules.DISPOSITION_ALIASES],
        })
        alias_df = alias_df.with_columns(
            pl.col("raw_disposition").str.strip_chars().str.to_uppercase()
        ).unique(subset=["raw_disposition"], keep="first")

        ranking_df = pl.DataFrame({
            "disposition": [r[0] for r in business_rules.DISPOSITION_RANKING],
            "rank_val": [r[1] for r in business_rules.DISPOSITION_RANKING],
            "category": [r[2] for r in business_rules.DISPOSITION_RANKING],
        }, schema={"disposition": pl.String, "rank_val": pl.Float64, "category": pl.String})
        ranking_df = ranking_df.with_columns(
            pl.col("disposition").str.strip_chars().str.to_uppercase().alias("ranking_disp_upper")
        ).sort(by=["ranking_disp_upper", "rank_val", "category"]).unique(subset=["ranking_disp_upper"], keep="first")

        # Map to uppercase keys for robust joins
        call_dp = call_dp.with_columns(
            pl.col("disposition_cleaned").str.strip_chars().str.to_uppercase().alias("cleaned_upper")
        )

        # 1st join: Alias map
        call_dp = call_dp.join(alias_df, left_on="cleaned_upper", right_on="raw_disposition", how="left")
        self.logger.info("Validation Log - Rows after alias mapping: %d", len(call_dp))

        # Prep ranking key
        ranking_key = (
            pl.coalesce([pl.col("ranking_disposition"), pl.col("disposition_cleaned")])
            .str.strip_chars()
            .str.to_uppercase()
        )
        call_dp = call_dp.with_columns(ranking_key.alias("ranking_key_upper"))

        # 2nd join: Ranking map
        call_dp = call_dp.join(ranking_df, left_on="ranking_key_upper", right_on="ranking_disp_upper", how="left")
        
        # Rename disposition from ranking_df
        call_dp = call_dp.rename({"disposition": "ranked_disposition"})
        self.logger.info("Validation Log - Rows after ranking: %d", len(call_dp))

        # Final disposition selection logic
        cleaned_str = pl.col("disposition_cleaned")
        is_ptp_prefix_rule = cleaned_str.str.starts_with("PTP - ") & (cleaned_str != "PTP - WhatsApp")

        case_ptp_expr = (
            pl.when(is_ptp_prefix_rule)
            .then(pl.lit("PTP"))
            .otherwise(pl.col("normalized_disposition"))
        )

        final_disposition_expr = pl.coalesce([
            case_ptp_expr,
            pl.col("ranked_disposition"),
            pl.when(pl.col("connected_flag") == 1)
            .then(pl.lit("Other"))
            .otherwise(pl.lit("No Answer"))
        ])

        final_rank_val_expr = pl.col("rank_val").fill_null(999.0)

        final_category_expr = pl.coalesce([
            pl.col("category"),
            pl.when(pl.col("connected_flag") == 1)
            .then(pl.lit("Not Categorized"))
            .otherwise(pl.lit("Not Contactable"))
        ])

        # final projection
        queue_col = "BlasterQueueID" if "BlasterQueueID" in call_dp.columns else "blasterqueueid"
        self.all_blaster_dispositions_df = call_dp.select([
            pl.col("LoanMstID"),
            pl.col("blaster_date"),
            pl.col(queue_col).alias("BlasterQueueID"),
            pl.col(ext_col1_col).alias("source_extracolumn1"),
            final_disposition_expr.alias("disposition"),
            pl.col("connected_flag"),
            final_rank_val_expr.alias("rank_val"),
            final_category_expr.alias("category"),
            pl.lit(1).alias("source_priority")
        ])

        # Validation stats logging
        row_count = len(self.all_blaster_dispositions_df)
        self.logger.info("Validation Log - Rows after final projection: %d", row_count)
        self.logger.info("Final Blaster DataFrame Shape: %s", self.all_blaster_dispositions_df.shape)
        self.logger.info("Final Blaster DataFrame Columns: %s", self.all_blaster_dispositions_df.columns)
        self.logger.info("Unique LoanMstID count: %d", self.all_blaster_dispositions_df["LoanMstID"].n_unique())
        self.logger.info("Unique BlasterQueueID count: %d", self.all_blaster_dispositions_df["BlasterQueueID"].n_unique())
        self.logger.info("Null counts per column: %s", self.all_blaster_dispositions_df.null_count().to_dicts()[0])
        
        # Store MTD base
        self.blaster_mtd_base_df = classified

        # Memory profiling
        mem_bytes = self.all_blaster_dispositions_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)

    @timer
    def build_best_disposition_per_loan(self) -> None:
        """
        Combines and aggregates normalized dispositions from VoiceBot, WhatsApp, and Blaster.
        Selects the single highest-priority disposition for every LoanMstID.
        Stores the result as self.best_disposition_df.
        """
        self.logger.info("Starting Best Disposition processing (build_best_disposition_per_loan)...")
        if self.all_call_dispositions_df is None:
            raise ValueError("VoiceBot dispositions data is not loaded.")
        if self.all_wa_dispositions_df is None:
            raise ValueError("WhatsApp dispositions data is not loaded.")
        if self.all_blaster_dispositions_df is None:
            raise ValueError("Blaster dispositions data is not loaded.")

        # Step 2: Standardize and cast all three schemas to prevent schema mismatch/implicit casting
        vb_standard = self.all_call_dispositions_df.select([
            pl.col("LoanMstID").cast(pl.Int64),
            pl.col("disposition").cast(pl.String),
            pl.col("disposition_ptp_date").cast(pl.Date),
            pl.col("is_ptp_source").cast(pl.Int32),
            pl.col("VoiceBotQueueID").cast(pl.Int64).alias("source_queue_id"),
            pl.col("source_extracolumn1").cast(pl.String).alias("source_call_disposition"),
            pl.col("rank_val").cast(pl.Float64),
            pl.col("category").cast(pl.String),
            pl.col("source_priority").cast(pl.Int32),
            pl.col("connected_flag").cast(pl.Int32),
            pl.lit("VoiceBot").cast(pl.String).alias("source"),
            pl.col("call_date").cast(pl.Date).alias("disposition_date")
        ])

        wa_standard = self.all_wa_dispositions_df.select([
            pl.col("LoanMstID").cast(pl.Int64),
            pl.col("disposition").cast(pl.String),
            pl.col("disposition_ptp_date").cast(pl.Date),
            pl.lit(0).cast(pl.Int32).alias("is_ptp_source"),
            pl.lit(None).cast(pl.Int64).alias("source_queue_id"),
            pl.lit(None).cast(pl.String).alias("source_call_disposition"),
            pl.col("rank_val").cast(pl.Float64),
            pl.col("category").cast(pl.String),
            pl.col("source_priority").cast(pl.Int32),
            pl.col("connected_flag").cast(pl.Int32),
            pl.lit("WhatsApp").cast(pl.String).alias("source"),
            pl.col("wa_date").cast(pl.Date).alias("disposition_date")
        ])

        bl_standard = self.all_blaster_dispositions_df.select([
            pl.col("LoanMstID").cast(pl.Int64),
            pl.col("disposition").cast(pl.String),
            pl.lit(None).cast(pl.Date).alias("disposition_ptp_date"),
            pl.lit(0).cast(pl.Int32).alias("is_ptp_source"),
            pl.col("BlasterQueueID").cast(pl.Int64).alias("source_queue_id"),
            pl.col("source_extracolumn1").cast(pl.String).alias("source_call_disposition"),
            pl.col("rank_val").cast(pl.Float64),
            pl.col("category").cast(pl.String),
            pl.col("source_priority").cast(pl.Int32),
            pl.col("connected_flag").cast(pl.Int32),
            pl.lit("Blaster").cast(pl.String).alias("source"),
            pl.col("blaster_date").cast(pl.Date).alias("disposition_date")
        ])

        # Step 3: Vertical Concatenation
        combined = pl.concat([vb_standard, wa_standard, bl_standard], how="vertical")
        
        # Log source row counts and contribution details
        self.logger.info("VoiceBot rows : %d", len(vb_standard))
        self.logger.info("WhatsApp rows : %d", len(wa_standard))
        self.logger.info("Blaster rows : %d", len(bl_standard))
        self.logger.info("Combined rows : %d", len(combined))
        self.logger.info("Combined columns: %s", combined.columns)

        # Step 4 & 5: Sort and select DISTINCT ON equivalent
        # Priority rules order:
        # LoanMstID ASC, rank_val ASC, connected_flag DESC, source_priority ASC, disposition_date DESC, source_queue_id DESC
        sorted_df = combined.sort(
            by=["LoanMstID", "rank_val", "connected_flag", "source_priority", "disposition_date", "source_queue_id"],
            descending=[False, False, True, False, True, True],
            nulls_last=True
        )
        self.logger.info("Validation Log - Row count after sorting: %d", len(sorted_df))

        # DISTINCT ON behavior: unique on LoanMstID keeping the first row
        deduplicated = sorted_df.unique(subset=["LoanMstID"], keep="first")

        # Step 6: Project only the target columns
        self.best_disposition_df = deduplicated.select([
            pl.col("LoanMstID"),
            pl.col("disposition").alias("BestDisposition"),
            pl.col("category").alias("BestDispositionCategory"),
            pl.col("source").alias("BestDispositionSource"),
            pl.col("disposition_date").alias("BestDispositionDate"),
            pl.col("disposition").alias("PTPType"),
            pl.col("disposition_ptp_date").alias("PTPDate"),
            pl.col("rank_val").alias("Rank"),
            pl.col("connected_flag").alias("ConnectedFlag"),
            pl.col("is_ptp_source"),
            pl.col("source_queue_id"),
            pl.col("source_call_disposition")
        ])

        # Step 7: Validation stats logging
        row_count = len(self.best_disposition_df)
        self.logger.info("Validation Log - Final rows: %d", row_count)
        self.logger.info("Final DataFrame Shape: %s", self.best_disposition_df.shape)
        self.logger.info("Unique LoanMstID count: %d", self.best_disposition_df["LoanMstID"].n_unique())
        self.logger.info("Duplicate LoanMstID count: %d", row_count - self.best_disposition_df["LoanMstID"].n_unique())
        self.logger.info("Null counts per column: %s", self.best_disposition_df.null_count().to_dicts()[0])
        
        # Memory profiling
        mem_bytes = self.best_disposition_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)

    @timer
    def build_voicebot_summary(self) -> None:
        """
        Aggregates VoiceBot history to produce one row per LoanMstID with attempt totals
        and latest mapped dispositions.
        Stores the result as self.voicebot_summary_df.
        """
        self.logger.info("Starting VoiceBot Summary processing (build_voicebot_summary)...")
        if self.voicebot_df is None:
            raise ValueError("VoiceBot data is not loaded.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded.")
        if self.all_call_dispositions_df is None:
            raise ValueError("All call dispositions data is not loaded.")

        # Log input counts
        self.logger.info("Validation Log - Loaded VoiceBot rows: %d", len(self.voicebot_df))
        self.logger.info("Validation Log - Base accounts rows: %d", len(self.base_accounts_df))
        self.logger.info("Validation Log - All call dispositions rows: %d", len(self.all_call_dispositions_df))
        self.logger.info("Unique LoanMstIDs in loaded VoiceBot: %s", self.voicebot_df.select("LoanMstID").unique().sort("LoanMstID")["LoanMstID"].to_list())

        # Build latest_call_dispo strictly chronologically (sort call_date DESC and VoiceBotQueueID DESC)
        latest_call_dispo = (
            self.all_call_dispositions_df
            .lazy()
            .sort(by=["LoanMstID", "call_date", "VoiceBotQueueID"], descending=[False, True, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("VoiceBot Latest DP"),
                pl.col("disposition").alias("LM Dispo."),
                pl.col("disposition_ptp_date").alias("PTP Date (From VoiceBot)")
            ])
        )

        # Build aggregation pipeline using LazyFrame
        created_date_col = "CreatedDate" if "CreatedDate" in self.voicebot_df.columns else "createddate"
        call_tried_col = "CallTried" if "CallTried" in self.voicebot_df.columns else "calltried"
        bank_mst_col = "BankMstID" if "BankMstID" in self.voicebot_df.columns else "bankmstid"
        recording_col = "Recording" if "Recording" in self.voicebot_df.columns else "recording"

        is_connected_col = (
            pl.col(recording_col).is_not_null() & 
            (pl.col(recording_col).str.len_bytes() > 1)
        )

        agg_lf = (
            self.voicebot_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (pl.col(bank_mst_col) == 53) &
                (pl.col(call_tried_col) > 0)
            )
            .group_by("LoanMstID")
            .agg([
                pl.col(created_date_col).max().cast(pl.Date).alias("VoiceBot Calling Date"),
                pl.len().alias("VoiceBot Call Attempts (MTD)"),
                is_connected_col.sum().alias("VoiceBot Call Connected Count (MTD)")
            ])
            .with_columns([
                pl.when(pl.col("VoiceBot Call Connected Count (MTD)") > 0)
                .then(pl.lit("Connected"))
                .otherwise(pl.lit("Not Connected"))
                .alias("VoiceBot CN/NC"),
                
                pl.when(pl.col("VoiceBot Call Connected Count (MTD)") > 0)
                .then(pl.lit(1))
                .otherwise(pl.lit(0))
                .alias("VoiceBot CN_Flag")
            ])
        )

        # Left join aggregation with latest call dispositions
        joined_lf = (
            agg_lf
            .join(
                latest_call_dispo,
                on="LoanMstID",
                how="left"
            )
        )

        # Project exactly the 9 required columns in the correct sequence
        final_lf = joined_lf.select([
            pl.col("LoanMstID"),
            pl.col("VoiceBot Calling Date"),
            pl.col("VoiceBot Call Attempts (MTD)"),
            pl.col("VoiceBot Call Connected Count (MTD)"),
            pl.col("VoiceBot CN/NC"),
            pl.col("VoiceBot CN_Flag"),
            pl.col("VoiceBot Latest DP"),
            pl.col("LM Dispo."),
            pl.col("PTP Date (From VoiceBot)")
        ])

        # Trigger execution EXACTLY once at the end
        self.voicebot_summary_df = final_lf.collect()

        # Step 2 & 3 logging: print filtered and summary unique LoanMstIDs
        filtered = self.voicebot_df.filter(
            (pl.col(bank_mst_col) == 53) & 
            (pl.col(call_tried_col) > 0)
        )
        self.logger.info("Unique LoanMstIDs in filtered VoiceBot: %s", filtered.select("LoanMstID").unique().sort("LoanMstID")["LoanMstID"].to_list())
        self.logger.info("LoanMstIDs in VoiceBot summary: %s", self.voicebot_summary_df.select("LoanMstID").sort("LoanMstID")["LoanMstID"].to_list())

        # Log final validation stats
        row_count = len(self.voicebot_summary_df)
        self.logger.info("Validation Log - Final rows: %d", row_count)
        self.logger.info("Final DataFrame Shape: %s", self.voicebot_summary_df.shape)
        self.logger.info("Unique LoanMstID count: %d", self.voicebot_summary_df["LoanMstID"].n_unique())
        self.logger.info("Duplicate LoanMstID count: %d", row_count - self.voicebot_summary_df["LoanMstID"].n_unique())
        self.logger.info("Null counts per column: %s", self.voicebot_summary_df.null_count().to_dicts()[0])
        
        # Memory profiling
        mem_bytes = self.voicebot_summary_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)


    @timer
    def build_whatsapp_summary(self) -> None:
        """
        Aggregates WhatsApp history to produce one row per LoanMstID with attempt totals
        and latest mapped WhatsApp disposition.
        Stores the result as self.whatsapp_summary_df.
        """
        self.logger.info("Starting WhatsApp Summary processing (build_whatsapp_summary)...")
        if self.whatsapp_df is None:
            raise ValueError("WhatsApp data is not loaded.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded.")
        if self.all_wa_dispositions_df is None:
            raise ValueError("All WhatsApp dispositions data is not loaded.")

        # Log input counts
        self.logger.info("Validation Log - Loaded WhatsApp rows: %d", len(self.whatsapp_df))
        self.logger.info("Validation Log - Base accounts rows: %d", len(self.base_accounts_df))
        self.logger.info("Validation Log - All WhatsApp dispositions rows: %d", len(self.all_wa_dispositions_df))

        # Sort strictly chronologically by wa_date DESC and WhatsAppQueueID DESC
        queue_col = "WhatsAppQueueID" if "WhatsAppQueueID" in self.all_wa_dispositions_df.columns else "whatsappqueueid"
        latest_wa_dispo = (
            self.all_wa_dispositions_df
            .lazy()
            .sort(by=["LoanMstID", "wa_date", queue_col], descending=[False, True, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("WhatsApp Latest DP"),
                pl.col("disposition").alias("LM Dispo."),
                pl.col("disposition_ptp_date").alias("PTP Date (From WhatsApp)")
            ])
        )

        # Build aggregation pipeline using LazyFrame
        is_sent_col = "IsSent" if "IsSent" in self.whatsapp_df.columns else "issent"
        bank_mst_col = "BankMstID" if "BankMstID" in self.whatsapp_df.columns else "bankmstid"
        created_date_col = "CreatedDate" if "CreatedDate" in self.whatsapp_df.columns else "createddate"
        is_read_col = "IsRead" if "IsRead" in self.whatsapp_df.columns else "isread"

        # Check if received_after_send is already in columns, otherwise default to False
        has_ras_col = any(c.lower() == "received_after_send" for c in self.whatsapp_df.columns)
        if has_ras_col:
            ras_col_name = [c for c in self.whatsapp_df.columns if c.lower() == "received_after_send"][0]
            received_after_send_expr = pl.col(ras_col_name)
        else:
            received_after_send_expr = pl.lit(False)

        is_connected_wa = (pl.col(is_read_col) == True) | (received_after_send_expr == True)

        agg_lf = (
            self.whatsapp_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (pl.col(bank_mst_col) == 53) &
                (pl.col(is_sent_col) == True)
            )
            .group_by("LoanMstID")
            .agg([
                pl.col(created_date_col).max().cast(pl.Date).alias("WhatsApp SMS sent date"),
                pl.len().alias("WhatsApp Attempts (MTD)"),
                is_connected_wa.sum().alias("WhatsApp Connected Count (MTD)")
            ])
            .with_columns([
                pl.when(pl.col("WhatsApp Connected Count (MTD)") > 0)
                .then(pl.lit("Connected"))
                .otherwise(pl.lit("Not Connected"))
                .alias("WhatsApp CN/NC"),
                
                pl.when(pl.col("WhatsApp Connected Count (MTD)") > 0)
                .then(pl.lit(1))
                .otherwise(pl.lit(0))
                .alias("WhatsApp CN_Flag")
            ])
        )

        # Left join aggregation with latest WhatsApp dispositions
        joined_lf = (
            agg_lf
            .join(
                latest_wa_dispo,
                on="LoanMstID",
                how="left"
            )
        )

        # Project exactly the required columns in the correct sequence
        final_lf = joined_lf.select([
            pl.col("LoanMstID"),
            pl.col("WhatsApp SMS sent date"),
            pl.col("WhatsApp Attempts (MTD)"),
            pl.col("WhatsApp Connected Count (MTD)"),
            pl.col("WhatsApp CN/NC"),
            pl.col("WhatsApp CN_Flag"),
            pl.col("WhatsApp Latest DP"),
            pl.col("LM Dispo."),
            pl.col("PTP Date (From WhatsApp)")
        ])

        # Trigger execution EXACTLY once at the end
        self.whatsapp_summary_df = final_lf.collect()

        # Log final validation stats
        row_count = len(self.whatsapp_summary_df)
        self.logger.info("Validation Log - Final rows: %d", row_count)
        self.logger.info("Final DataFrame Shape: %s", self.whatsapp_summary_df.shape)
        self.logger.info("Unique LoanMstID count: %d", self.whatsapp_summary_df["LoanMstID"].n_unique())
        self.logger.info("Duplicate LoanMstID count: %d", row_count - self.whatsapp_summary_df["LoanMstID"].n_unique())
        self.logger.info("Null counts per column: %s", self.whatsapp_summary_df.null_count().to_dicts()[0])
        
        # Memory profiling
        mem_bytes = self.whatsapp_summary_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)


    @timer
    def build_blaster_summary(self) -> None:
        """
        Aggregates Blaster history to produce one row per LoanMstID with attempt totals
        and latest mapped Blaster disposition.
        Stores the result as self.blaster_summary_df.
        """
        self.logger.info("Starting Blaster Summary processing (build_blaster_summary)...")
        if self.blaster_df is None:
            raise ValueError("Blaster data is not loaded.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded.")
        if self.all_blaster_dispositions_df is None:
            raise ValueError("All Blaster dispositions data is not loaded.")

        # Log input counts
        self.logger.info("Validation Log - Loaded Blaster rows: %d", len(self.blaster_df))
        self.logger.info("Validation Log - Base accounts rows: %d", len(self.base_accounts_df))
        self.logger.info("Validation Log - All Blaster dispositions rows: %d", len(self.all_blaster_dispositions_df))

        # Log unique LoanMstIDs in loaded Blaster
        self.logger.info("Unique LoanMstIDs in loaded Blaster: %s", self.blaster_df.select("LoanMstID").unique().sort("LoanMstID")["LoanMstID"].to_list())

        # Sort strictly chronologically: LoanMstID ASC, blaster_date DESC, BlasterQueueID DESC
        queue_col = "BlasterQueueID" if "BlasterQueueID" in self.all_blaster_dispositions_df.columns else "blasterqueueid"
        latest_blaster_dispo = (
            self.all_blaster_dispositions_df
            .lazy()
            .sort(by=["LoanMstID", "blaster_date", queue_col], descending=[False, True, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("Blaster Latest DP"),
                pl.col("disposition").alias("LM Dispo."),
                pl.lit(None).cast(pl.Date).alias("PTP Date (From Blaster)")
            ])
        )

        # Build aggregation pipeline using LazyFrame
        call_tried_col = "CallTried" if "CallTried" in self.blaster_df.columns else "calltried"
        bank_mst_col = "BankMstID" if "BankMstID" in self.blaster_df.columns else "bankmstid"
        created_date_col = "CreatedDate" if "CreatedDate" in self.blaster_df.columns else "createddate"
        connected_col = "CallConnected" if "CallConnected" in self.blaster_df.columns else "callconnected"

        # Unique LoanMstIDs check after filtering (using eager filter)
        filtered_eager = self.blaster_df.filter(
            (pl.col(bank_mst_col) == 53) &
            (pl.col(call_tried_col) > 0)
        )
        self.logger.info("Unique LoanMstIDs after filtering: %s", filtered_eager.select("LoanMstID").unique().sort("LoanMstID")["LoanMstID"].to_list())

        is_connected_blaster = pl.col(connected_col).fill_null(False)

        agg_lf = (
            self.blaster_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (pl.col(bank_mst_col) == 53) &
                (pl.col(call_tried_col) > 0)
            )
            .group_by("LoanMstID")
            .agg([
                pl.col(created_date_col).max().cast(pl.Date).alias("Blaster Calling Date"),
                pl.len().alias("Blaster Call Attempts (MTD)"),
                is_connected_blaster.sum().alias("Blaster Call Connected Count (MTD)")
            ])
            .with_columns([
                pl.when(pl.col("Blaster Call Connected Count (MTD)") > 0)
                .then(pl.lit("Connected"))
                .otherwise(pl.lit("Not Connected"))
                .alias("Blaster CN/NC"),
                
                pl.when(pl.col("Blaster Call Connected Count (MTD)") > 0)
                .then(pl.lit(1))
                .otherwise(pl.lit(0))
                .alias("Blaster CN_Flag")
            ])
        )

        # Left join aggregation with latest Blaster dispositions
        joined_lf = (
            agg_lf
            .join(
                latest_blaster_dispo,
                on="LoanMstID",
                how="left"
            )
        )

        # Project exactly the required columns in the correct sequence
        final_lf = joined_lf.select([
            pl.col("LoanMstID"),
            pl.col("Blaster Calling Date"),
            pl.col("Blaster Call Attempts (MTD)"),
            pl.col("Blaster Call Connected Count (MTD)"),
            pl.col("Blaster CN/NC"),
            pl.col("Blaster CN_Flag"),
            pl.col("Blaster Latest DP"),
            pl.col("LM Dispo."),
            pl.col("PTP Date (From Blaster)")
        ])

        # Trigger execution EXACTLY once at the end
        self.blaster_summary_df = final_lf.collect()

        # Log unique LoanMstIDs in Blaster summary
        self.logger.info("LoanMstIDs in Blaster summary: %s", self.blaster_summary_df.select("LoanMstID").sort("LoanMstID")["LoanMstID"].to_list())

        # Log final validation stats
        row_count = len(self.blaster_summary_df)
        self.logger.info("Validation Log - Final rows: %d", row_count)
        self.logger.info("Final DataFrame Shape: %s", self.blaster_summary_df.shape)
        self.logger.info("Unique LoanMstID count: %d", self.blaster_summary_df["LoanMstID"].n_unique())
        self.logger.info("Duplicate LoanMstID count: %d", row_count - self.blaster_summary_df["LoanMstID"].n_unique())
        self.logger.info("Null counts per column: %s", self.blaster_summary_df.null_count().to_dicts()[0])
        
        # Memory profiling
        mem_bytes = self.blaster_summary_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)


    @timer
    def build_communication_summary(self) -> None:
        """
        Integrates the previously generated intermediate DataFrames (base_accounts_df,
        best_disposition_df, voicebot_summary_df, whatsapp_summary_df, blaster_summary_df)
        into a single communication summary using LEFT joins on LoanMstID.
        Stores the result as self.communication_summary_df.
        """
        self.logger.info("Starting Communication Summary processing (build_communication_summary)...")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded.")
        if self.best_disposition_df is None:
            raise ValueError("Best disposition data is not loaded.")
        if self.voicebot_summary_df is None:
            raise ValueError("VoiceBot summary data is not loaded.")
        if self.whatsapp_summary_df is None:
            raise ValueError("WhatsApp summary data is not loaded.")
        if self.blaster_summary_df is None:
            raise ValueError("Blaster summary data is not loaded.")

        # Log input row counts
        self.logger.info("Validation Log - Loaded Base Accounts rows: %d", len(self.base_accounts_df))
        self.logger.info("Validation Log - Loaded Best Disposition rows: %d", len(self.best_disposition_df))
        self.logger.info("Validation Log - Loaded VoiceBot Summary rows: %d", len(self.voicebot_summary_df))
        self.logger.info("Validation Log - Loaded WhatsApp Summary rows: %d", len(self.whatsapp_summary_df))
        self.logger.info("Validation Log - Loaded Blaster Summary rows: %d", len(self.blaster_summary_df))

        # Convert to LazyFrame to perform all joins and trigger collection once
        base_lf = self.base_accounts_df.lazy()
        best_lf = self.best_disposition_df.lazy()
        voicebot_lf = self.voicebot_summary_df.lazy()
        whatsapp_lf = self.whatsapp_summary_df.lazy()
        blaster_lf = self.blaster_summary_df.lazy()

        # Step 1: Base Accounts
        self.logger.info("Validation Log - Base Accounts row count: %d, uniques: %d, duplicates: %d",
                         len(self.base_accounts_df),
                         self.base_accounts_df["LoanMstID"].n_unique(),
                         len(self.base_accounts_df) - self.base_accounts_df["LoanMstID"].n_unique())

        # Step 2: Join best_disposition_df
        join1_lf = base_lf.join(best_lf, on="LoanMstID", how="left")
        j1_ids = join1_lf.select("LoanMstID").collect()
        self.logger.info("Validation Log - After Best Disposition Join rows: %d, uniques: %d, duplicates: %d",
                         len(j1_ids),
                         j1_ids["LoanMstID"].n_unique(),
                         len(j1_ids) - j1_ids["LoanMstID"].n_unique())

        # Step 3: Join voicebot_summary_df
        join2_lf = join1_lf.join(voicebot_lf, on="LoanMstID", how="left")
        j2_ids = join2_lf.select("LoanMstID").collect()
        self.logger.info("Validation Log - After VoiceBot Summary Join rows: %d, uniques: %d, duplicates: %d",
                         len(j2_ids),
                         j2_ids["LoanMstID"].n_unique(),
                         len(j2_ids) - j2_ids["LoanMstID"].n_unique())

        # Step 4: Join whatsapp_summary_df
        join3_lf = join2_lf.join(whatsapp_lf, on="LoanMstID", how="left", suffix="_wa")
        j3_ids = join3_lf.select("LoanMstID").collect()
        self.logger.info("Validation Log - After WhatsApp Summary Join rows: %d, uniques: %d, duplicates: %d",
                         len(j3_ids),
                         j3_ids["LoanMstID"].n_unique(),
                         len(j3_ids) - j3_ids["LoanMstID"].n_unique())

        # Step 5: Join blaster_summary_df
        join4_lf = join3_lf.join(blaster_lf, on="LoanMstID", how="left", suffix="_bl")
        j4_ids = join4_lf.select("LoanMstID").collect()
        self.logger.info("Validation Log - After Blaster Summary Join rows: %d, uniques: %d, duplicates: %d",
                         len(j4_ids),
                         j4_ids["LoanMstID"].n_unique(),
                         len(j4_ids) - j4_ids["LoanMstID"].n_unique())

        # Trigger collection EXACTLY once
        self.communication_summary_df = join4_lf.collect()

        # Step 6: Assert that the final row count matches base_accounts_df
        assert len(self.communication_summary_df) == len(self.base_accounts_df), (
            f"Fidelity check failed: final summary row count ({len(self.communication_summary_df)}) "
            f"does not match base accounts row count ({len(self.base_accounts_df)})"
        )

        # Log final validation stats
        row_count = len(self.communication_summary_df)
        self.logger.info("Validation Log - Final rows: %d", row_count)
        self.logger.info("Final DataFrame Shape: %s", self.communication_summary_df.shape)
        self.logger.info("Final Column Names: %s", self.communication_summary_df.columns)
        self.logger.info("Unique LoanMstID count: %d", self.communication_summary_df["LoanMstID"].n_unique())
        self.logger.info("Duplicate LoanMstID count: %d", row_count - self.communication_summary_df["LoanMstID"].n_unique())
        self.logger.info("Null counts per column: %s", self.communication_summary_df.null_count().to_dicts()[0])
        
        # Memory profiling
        mem_bytes = self.communication_summary_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)


    @timer
    def build_latest_master(self) -> None:
        """
        Integrates the final communication summary into self.latest_master_df,
        acting as the single consolidated master source.
        """
        self.logger.info("Starting Latest Master processing (build_latest_master)...")
        
        # Verify required DataFrames exist
        deps = {
            "base_accounts_df": self.base_accounts_df,
            "latest_call_df": self.latest_call_df,
            "latest_blaster_df": self.latest_blaster_df,
            "latest_whatsapp_df": self.latest_whatsapp_df,
            "latest_response_df": self.latest_response_df,
            "latest_voicebot_response_df": self.latest_voicebot_response_df,
            "latest_whatsapp_response_df": self.latest_whatsapp_response_df,
            "latest_collection_df": self.latest_collection_df,
            "total_collection_df": self.total_collection_df,
            "whatsapp_combined_df": self.whatsapp_combined_df,
            "latest_call_extra_df": self.latest_call_extra_df,
            "mtd_connection_flags_df": self.mtd_connection_flags_df,
            "mtd_wa_connection_flags_df": self.mtd_wa_connection_flags_df,
            "best_disposition_df": self.best_disposition_df,
            "latest_voicebot_channel_disposition_df": self.latest_voicebot_channel_disposition_df,
            "latest_blaster_channel_disposition_df": self.latest_blaster_channel_disposition_df,
            "latest_whatsapp_channel_disposition_df": self.latest_whatsapp_channel_disposition_df,
            "last_month_best_disposition_df": self.last_month_best_disposition_df,
            "latest_disposition_mtd_df": self.latest_disposition_mtd_df,
            "last_connected_info_df": self.last_connected_info_df,
            "total_ptp_responses_df": self.total_ptp_responses_df,
            "last_connected_call_duration_df": self.last_connected_call_duration_df,
            "call_duration_stats_df": self.call_duration_stats_df,
            "voicebot_call_duration_stats_df": self.voicebot_call_duration_stats_df
        }
        for name, df in deps.items():
            if df is None:
                raise ValueError(f"Dependency DataFrame '{name}' is missing/None. Ensure previous phases have completed successfully.")

        # Select and alias columns to prevent suffixing/namespace conflicts
        base_lf = self.base_accounts_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("DisbursementID"),
            pl.col("DisbursementID").alias("Account Number"),
            pl.col("CustomerName").alias("Customer Name"),
            pl.col("CustomerNumber").alias("Customer Number"),
            pl.col("Primary Mobile Number"),
            pl.col("Branch").alias("Branch Name"),
            pl.col("BranchCode").alias("Branch Code"),
            pl.col("LoanClassification").alias("Pool Type"),
            pl.col("Loan Status"),
            pl.col("DND"),
            pl.col("Default Amt"),
            pl.col("EMI Amount"),
            pl.col("Total Outstanding")
        ])

        lc_lf = self.latest_call_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("AI Calling Date"),
            pl.col("call_attempts"),
            pl.col("calls_connected"),
            pl.col("calls_not_connected"),
            pl.col("DP_Code_Not_Connected"),
            pl.col("Extra_DP_Category").alias("Call_DP_Category"),
            pl.col("Extra_DP_Reason").alias("Call_DP_Reason"),
            pl.col("VoiceBotQueueID").alias("call_qid"),
            pl.col("connected_latest"),
            pl.col("not_connected_latest"),
            pl.col("call_has_paid"),
            pl.col("call_has_denied"),
            pl.col("call_has_ptp"),
            pl.col("call_promise_datetime")
        ])

        lb_lf = self.latest_blaster_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("Blaster Calling Date"),
            pl.col("blaster_attempts"),
            pl.col("blaster_connected_count"),
            pl.col("blaster_not_connected_count"),
            pl.col("Blaster_DP_Code"),
            pl.col("Blaster_DP_Category"),
            pl.col("BlasterQueueID").alias("blaster_qid"),
            pl.col("blaster_connected_latest"),
            pl.col("blaster_not_connected_latest")
        ])

        lw_lf = self.latest_whatsapp_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("Whatsapp Sent Date"),
            pl.col("whatsapp_attempts"),
            pl.col("whatsapp_delivered_count"),
            pl.col("whatsapp_read_count"),
            pl.col("whatsapp_connected_count"),
            pl.col("Extra_DP_Category").alias("WA_DP_Category"),
            pl.col("Extra_DP_Reason").alias("WA_DP_Reason"),
            pl.col("WhatsAppQueueID").alias("wa_qid"),
            pl.col("IsSent").alias("last_wa_sent"),
            pl.col("IsDelivered").alias("last_wa_delivered"),
            pl.col("IsRead").alias("last_wa_read")
        ])

        lr_lf = self.latest_response_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("WhatsappQueueID").alias("resp_whatsapp_qid")
        ])

        lvr_lf = self.latest_voicebot_response_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("Status").alias("Response Status"),
            pl.col("P2P_date"),
            pl.col("VoiceBotQueueID").alias("resp_voicebot_qid")
        ])

        lwr_lf = self.latest_whatsapp_response_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("Status").alias("WhatsApp Response Status"),
            pl.col("P2P_date").alias("WhatsApp_P2P_date"),
            pl.col("WhatsappQueueID").alias("latest_resp_whatsapp_qid")
        ])

        wc_lf = self.whatsapp_combined_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("all_messages").alias("All WhatsApp Messages")
        ])

        lce_lf = self.latest_call_extra_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("Conclusion"),
            pl.col("Recording")
        ])

        mcf_lf = self.mtd_connection_flags_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("unique_total_attempt"),
            pl.col("unique_total_connect"),
            pl.col("ai_connected_mtd"),
            pl.col("blaster_connected_mtd")
        ])

        mwcf_lf = self.mtd_wa_connection_flags_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("wa_connected_mtd")
        ])

        bd_lf = self.best_disposition_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("BestDisposition").alias("best_disposition"),
            pl.col("BestDispositionCategory").alias("best_category"),
            pl.col("BestDispositionSource").alias("best_disposition_channel"),
            pl.col("BestDispositionDate").alias("best_disposition_date"),
            pl.col("PTPDate").alias("best_disposition_ptp_date"),
            pl.col("is_ptp_source").alias("best_is_ptp_source"),
            pl.col("source_queue_id").alias("best_source_queue_id"),
            pl.col("source_call_disposition").alias("best_source_call_disposition")
        ])

        lvcd_lf = self.latest_voicebot_channel_disposition_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("voicebot_latest_disposition"),
            pl.col("voicebot_latest_disposition_category"),
            pl.col("voicebot_latest_disposition_date")
        ])

        lbcd_lf = self.latest_blaster_channel_disposition_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("blaster_latest_disposition"),
            pl.col("blaster_latest_disposition_category"),
            pl.col("blaster_latest_disposition_date")
        ])

        lwcd_lf = self.latest_whatsapp_channel_disposition_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("whatsapp_latest_disposition"),
            pl.col("whatsapp_latest_disposition_category"),
            pl.col("whatsapp_latest_disposition_date"),
            pl.col("whatsapp_latest_disposition_ptp_date")
        ])

        lmbd_lf = self.last_month_best_disposition_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("last_month_best_disposition"),
            pl.col("last_month_best_category")
        ])

        ldm_lf = self.latest_disposition_mtd_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("latest_disposition"),
            pl.col("latest_disposition_category"),
            pl.col("latest_disposition_channel"),
            pl.col("latest_disposition_date")
        ])

        lci_lf = self.last_connected_info_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("last_connected_date"),
            pl.col("last_connected_channel")
        ])

        tpr_lf = self.total_ptp_responses_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("ai_ptp_count"),
            pl.col("wa_ptp_count")
        ])

        lcd_lf = self.last_connected_call_duration_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("last_connected_call_duration")
        ])

        cds_lf = self.call_duration_stats_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("max_connected_call_duration")
        ])

        vcds_lf = self.voicebot_call_duration_stats_df.lazy().select([
            pl.col("LoanMstID"),
            pl.col("voicebot_max_connected_call_duration")
        ])

        col_lf = self.latest_collection_df.lazy().select([
            pl.col("DisbursementID"),
            pl.col("Last Collected Amount"),
            pl.col("Last Collection Date")
        ])

        tcol_lf = self.total_collection_df.lazy().select([
            pl.col("DisbursementID"),
            pl.col("total_collected_amt").alias("Total Collection Amt")
        ])

        # Step 2: Build the Lazy Join Pipeline
        pipeline_lf = (
            base_lf
            .join(lc_lf, on="LoanMstID", how="left")
            .join(lb_lf, on="LoanMstID", how="left")
            .join(lw_lf, on="LoanMstID", how="left")
            .join(lr_lf, on="LoanMstID", how="left")
            .join(lvr_lf, on="LoanMstID", how="left")
            .join(lwr_lf, on="LoanMstID", how="left")
            .join(wc_lf, on="LoanMstID", how="left")
            .join(lce_lf, on="LoanMstID", how="left")
            .join(mcf_lf, on="LoanMstID", how="left")
            .join(mwcf_lf, on="LoanMstID", how="left")
            .join(bd_lf, on="LoanMstID", how="left")
            .join(lvcd_lf, on="LoanMstID", how="left")
            .join(lbcd_lf, on="LoanMstID", how="left")
            .join(lwcd_lf, on="LoanMstID", how="left")
            .join(lmbd_lf, on="LoanMstID", how="left")
            .join(ldm_lf, on="LoanMstID", how="left")
            .join(lci_lf, on="LoanMstID", how="left")
            .join(tpr_lf, on="LoanMstID", how="left")
            .join(lcd_lf, on="LoanMstID", how="left")
            .join(cds_lf, on="LoanMstID", how="left")
            .join(vcds_lf, on="LoanMstID", how="left")
            .join(col_lf, on="DisbursementID", how="left")
            .join(tcol_lf, on="DisbursementID", how="left")
        )

        # Step 3: Compute SQL Derived Columns & Apply Filter
        calling_date_expr = (
            pl.when(pl.col("AI Calling Date").is_null()).then(pl.col("Blaster Calling Date"))
            .when(pl.col("Blaster Calling Date").is_null()).then(pl.col("AI Calling Date"))
            .otherwise(
                pl.max_horizontal(["AI Calling Date", "Blaster Calling Date"])
            )
        )

        dp_code_not_connected_expr = (
            pl.when(pl.col("DP_Code_Not_Connected").is_null()).then(pl.lit(None))
            .when(pl.col("DP_Code_Not_Connected").str.to_lowercase().str.contains("failed")).then(pl.lit("NO ANSWER"))
            .when(pl.col("DP_Code_Not_Connected").str.to_lowercase().str.contains("congestion")).then(pl.lit("NO ANSWER"))
            .otherwise(pl.col("DP_Code_Not_Connected"))
        )

        last_connected_call_duration_expr = (
            pl.when(pl.col("last_connected_channel").is_in(["VoiceBot", "Blaster"]))
            .then(pl.col("last_connected_call_duration"))
            .otherwise(pl.lit(None))
        )

        pipeline_lf = (
            pipeline_lf
            .with_columns([
                calling_date_expr.alias("Calling Date"),
                dp_code_not_connected_expr.alias("DP_Code_Not_Connected"),
                last_connected_call_duration_expr.alias("last_connected_call_duration")
            ])
            .filter(
                ~pl.col("Primary Mobile Number")
                  .cast(pl.Utf8)
                  .is_in(["0", "nan"])
            )
        )

        # Step 4: Project Columns in SQL Order
        self.latest_master_df = (
            pipeline_lf.select([
                pl.col("LoanMstID"),
                pl.col("Account Number"),
                pl.col("Customer Name"),
                pl.col("Customer Number"),
                pl.col("Primary Mobile Number"),
                pl.col("Branch Name"),
                pl.col("Branch Code"),
                pl.col("Pool Type"),
                pl.col("Loan Status"),
                pl.col("DND"),
                pl.col("Default Amt"),
                pl.col("EMI Amount"),
                pl.col("Total Outstanding"),
                pl.col("AI Calling Date"),
                pl.col("Blaster Calling Date"),
                pl.col("Calling Date"),
                pl.col("Whatsapp Sent Date"),
                pl.col("call_attempts"),
                pl.col("calls_connected"),
                pl.col("calls_not_connected"),
                pl.col("blaster_attempts"),
                pl.col("blaster_connected_count"),
                pl.col("blaster_not_connected_count"),
                pl.col("whatsapp_attempts"),
                pl.col("whatsapp_delivered_count"),
                pl.col("whatsapp_read_count"),
                pl.col("whatsapp_connected_count"),
                pl.col("DP_Code_Not_Connected"),
                pl.col("Blaster_DP_Code"),
                pl.col("Blaster_DP_Category"),
                pl.col("WA_DP_Category"),
                pl.col("WA_DP_Reason"),
                pl.col("Call_DP_Category"),
                pl.col("Call_DP_Reason"),
                pl.col("Response Status"),
                pl.col("P2P_date"),
                pl.col("resp_voicebot_qid"),
                pl.col("resp_whatsapp_qid"),
                pl.col("WhatsApp Response Status"),
                pl.col("WhatsApp_P2P_date"),
                pl.col("latest_resp_whatsapp_qid"),
                pl.col("call_qid"),
                pl.col("blaster_qid"),
                pl.col("wa_qid"),
                pl.col("Last Collected Amount"),
                pl.col("Last Collection Date"),
                pl.col("Total Collection Amt"),
                pl.col("All WhatsApp Messages"),
                pl.col("Conclusion"),
                pl.col("Recording"),
                pl.col("connected_latest"),
                pl.col("not_connected_latest"),
                pl.col("call_has_paid"),
                pl.col("call_has_denied"),
                pl.col("call_has_ptp"),
                pl.col("call_promise_datetime"),
                pl.col("blaster_connected_latest"),
                pl.col("blaster_not_connected_latest"),
                pl.col("last_wa_sent"),
                pl.col("last_wa_delivered"),
                pl.col("last_wa_read"),
                pl.col("unique_total_attempt"),
                pl.col("unique_total_connect"),
                pl.col("ai_connected_mtd"),
                pl.col("blaster_connected_mtd"),
                pl.col("wa_connected_mtd"),
                pl.col("best_disposition"),
                pl.col("best_category"),
                pl.col("best_disposition_channel"),
                pl.col("best_disposition_date"),
                pl.col("best_disposition_ptp_date"),
                pl.col("best_is_ptp_source"),
                pl.col("best_source_queue_id"),
                pl.col("best_source_call_disposition"),
                pl.col("voicebot_latest_disposition"),
                pl.col("voicebot_latest_disposition_category"),
                pl.col("voicebot_latest_disposition_date"),
                pl.col("blaster_latest_disposition"),
                pl.col("blaster_latest_disposition_category"),
                pl.col("blaster_latest_disposition_date"),
                pl.col("whatsapp_latest_disposition"),
                pl.col("whatsapp_latest_disposition_category"),
                pl.col("whatsapp_latest_disposition_date"),
                pl.col("whatsapp_latest_disposition_ptp_date"),
                pl.col("last_month_best_disposition"),
                pl.col("last_month_best_category"),
                pl.col("latest_disposition"),
                pl.col("latest_disposition_category"),
                pl.col("latest_disposition_channel"),
                pl.col("latest_disposition_date"),
                pl.col("last_connected_date"),
                pl.col("last_connected_channel"),
                pl.col("ai_ptp_count"),
                pl.col("wa_ptp_count"),
                pl.col("last_connected_call_duration"),
                pl.col("max_connected_call_duration"),
                pl.col("voicebot_max_connected_call_duration")
            ])
            .collect()
        )

        # Step 6: Validation & Assertions
        expected_rows = len(self.base_accounts_df.filter(~pl.col("Primary Mobile Number").cast(pl.Utf8).is_in(["0", "nan"])))
        assert len(self.latest_master_df) == expected_rows, (
            f"Fidelity check failed: final master row count ({len(self.latest_master_df)}) "
            f"does not match filtered base accounts row count ({expected_rows})"
        )

        row_count = len(self.latest_master_df)
        self.logger.info("Validation Log - Final rows: %d", row_count)
        self.logger.info("Final Shape: %s", self.latest_master_df.shape)
        self.logger.info("Final Column Names: %s", self.latest_master_df.columns)
        self.logger.info("Final Column Order: %s", self.latest_master_df.columns)
        self.logger.info("Unique LoanMstID count: %d", self.latest_master_df["LoanMstID"].n_unique())
        self.logger.info("Duplicate LoanMstID count: %d", row_count - self.latest_master_df["LoanMstID"].n_unique())
        self.logger.info("Null Counts: %s", self.latest_master_df.null_count().to_dicts()[0])
        
        mem_bytes = self.latest_master_df.estimated_size()
        self.logger.info("Estimated memory usage: %d bytes (%.4f KB)", mem_bytes, mem_bytes / 1024.0)

    @timer
    def build_daily_pivots(self) -> None:
        """
        Builds Phase 17 intermediate tables:
        - daily_call_dispositions_df
        - month_dates_df
        - daily_pivots_df
        """
        import datetime
        self.logger.info("Starting build_daily_pivots...")
        
        # Verify required DataFrames exist
        if self.voicebot_mtd_base_df is None:
            raise ValueError("voicebot_mtd_base_df is not available.")
        if self.all_call_dispositions_df is None:
            raise ValueError("all_call_dispositions_df is not available.")
        if self.latest_master_df is None:
            raise ValueError("latest_master_df is not available.")

        # Step 1: Build daily_call_dispositions_df
        created_date_col = "CreatedDate" if "CreatedDate" in self.voicebot_mtd_base_df.columns else "createddate"
        vb_queue_col = "VoiceBotQueueID" if "VoiceBotQueueID" in self.voicebot_mtd_base_df.columns else "voicebotqueueid"
        call_duration_col = "CallDuration" if "CallDuration" in self.voicebot_mtd_base_df.columns else "callduration"
        
        latest_daily_calls_lf = (
            self.voicebot_mtd_base_df
            .lazy()
            .sort(
                by=["LoanMstID", "call_date", created_date_col, vb_queue_col],
                descending=[False, False, True, True]
            )
            .unique(subset=["LoanMstID", "call_date"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("call_date"),
                pl.col(vb_queue_col).alias("VoiceBotQueueID"),
                pl.col(call_duration_col).alias("call_duration"),
                pl.col("is_connected")
            ])
        )

        # LEFT JOIN all_call_dispositions_df on VoiceBotQueueID
        daily_call_dispositions_lf = (
            latest_daily_calls_lf
            .join(
                self.all_call_dispositions_df.lazy().select(["VoiceBotQueueID", "disposition"]),
                on="VoiceBotQueueID",
                how="left"
            )
            .with_columns(
                pl.when(pl.col("disposition").is_not_null() & (pl.col("disposition").str.strip_chars() != ""))
                .then(pl.col("disposition"))
                .when(pl.col("is_connected") == True)
                .then(pl.lit("No Response"))
                .otherwise(pl.lit("No Communication"))
                .alias("disposition")
            )
            .select([
                pl.col("LoanMstID"),
                pl.col("call_date"),
                pl.col("call_duration"),
                pl.col("disposition")
            ])
        )
        self.daily_call_dispositions_df = daily_call_dispositions_lf.collect()

        # Step 2: Build month_dates_df
        dates = []
        curr = self.request.start_date
        while curr <= self.request.end_date:
            dates.append(curr)
            curr += datetime.timedelta(days=1)
        self.month_dates_df = pl.DataFrame({"date_val": dates}, schema={"date_val": pl.Date})

        # Step 3: Build daily_pivots_df
        lm_subset_lf = self.latest_master_df.lazy().select(["LoanMstID", "Account Number"])
        md_lf = self.month_dates_df.lazy()
        
        # CROSS JOIN lm and md
        cross_lf = lm_subset_lf.join(md_lf, how="cross")
        
        # LEFT JOIN daily_call_dispositions_df on LoanMstID and date_val = call_date
        joined_lf = (
            cross_lf
            .join(
                self.daily_call_dispositions_df.lazy(),
                left_on=["LoanMstID", "date_val"],
                right_on=["LoanMstID", "call_date"],
                how="left"
            )
        )

        # Build pivot expressions: format date as %d-%b-%y (e.g. 01-Jun-26)
        pivot_exprs = []
        curr_d = self.request.start_date
        while curr_d <= self.request.end_date:
            date_label = curr_d.strftime("%d-%b-%y")
            pivot_exprs.append(
                pl.col("disposition")
                .filter(pl.col("date_val") == curr_d)
                .max()
                .fill_null("")
                .alias(date_label)
            )
            curr_d += datetime.timedelta(days=1)

        daily_pivots_lf = (
            joined_lf
            .group_by(["LoanMstID", "Account Number"])
            .agg(pivot_exprs)
        )
        self.daily_pivots_df = daily_pivots_lf.collect()

        # Logging & Assertions for all three DataFrames
        dfs = [
            ("daily_call_dispositions_df", self.daily_call_dispositions_df),
            ("month_dates_df", self.month_dates_df),
            ("daily_pivots_df", self.daily_pivots_df),
        ]
        for name, df in dfs:
            self.logger.info("Validation Log - %s shape: %s", name, df.shape)
            self.logger.info("Validation Log - %s columns: %s", name, df.columns)
            self.logger.info("Validation Log - %s column order: %s", name, df.columns)
            self.logger.info("Validation Log - %s null counts: %s", name, df.null_count().to_dicts()[0])
            if "LoanMstID" in df.columns:
                self.logger.info("Validation Log - %s unique LoanMstID: %d", name, df["LoanMstID"].n_unique())
                self.logger.info("Validation Log - %s duplicate LoanMstID: %d", name, len(df) - df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s estimated memory: %d bytes (%.4f KB)", name, df.estimated_size(), df.estimated_size() / 1024.0)

        # Assert daily_pivots_df has unique LoanMstID
        assert len(self.daily_pivots_df) == self.daily_pivots_df["LoanMstID"].n_unique(), "Fidelity Check Failed: daily_pivots_df contains duplicate LoanMstID values!"

    @timer
    def build_final_report(self) -> None:
        """
        Builds Phase 18 final report table:
        - final_report_df
        """
        self.logger.info("Starting build_final_report...")
        
        # Verify required DataFrames exist
        if self.latest_master_df is None:
            raise ValueError("latest_master_df is not available.")
        if self.daily_pivots_df is None:
            raise ValueError("daily_pivots_df is not available.")
        if self.whatsapp_messages_df is None:
            raise ValueError("whatsapp_messages_df is not available.")
        if self.base_accounts_df is None:
            raise ValueError("base_accounts_df is not available.")

        # Step 1: Build wa_received_by_day_df, wa_received_mtd_loans_df, and wa_received_any_loans_df helpers
        assigned_to_col = "AssignedTo" if "AssignedTo" in self.whatsapp_messages_df.columns else "assignedto"
        msg_date_col = "MessageDate" if "MessageDate" in self.whatsapp_messages_df.columns else "messagedate"

        wa_received_by_day_lf = (
            self.whatsapp_messages_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (
                    pl.col(assigned_to_col).is_null() |
                    (pl.col(assigned_to_col) != "System")
                ) &
                (pl.col(msg_date_col).cast(pl.Date) >= self.request.start_date) &
                (pl.col(msg_date_col).cast(pl.Date) <= self.request.end_date)
            )
            .select([
                pl.col("LoanMstID"),
                pl.col(msg_date_col).cast(pl.Date).alias("received_date")
            ])
            .unique()
        )

        wa_received_mtd_loans_lf = (
            wa_received_by_day_lf
            .select("LoanMstID")
            .unique()
        )

        wa_received_any_loans_lf = (
            wa_received_mtd_loans_lf
            .select("LoanMstID")
            .with_columns(
                pl.lit(1).alias("has_wa_received")
            )
        )

        # Step 2: Build intermediate computations on latest_master_df
        pivot_cols = [c for c in self.daily_pivots_df.columns if c not in ["LoanMstID", "Account Number"]]
        
        # Blaster disposition mapping helper
        blaster_val = pl.coalesce([
            pl.when(pl.col("Blaster_DP_Category").fill_null("").str.strip_chars() != "").then(pl.col("Blaster_DP_Category")).otherwise(pl.lit(None)),
            pl.when(pl.col("Blaster_DP_Code").fill_null("").str.strip_chars() != "").then(pl.col("Blaster_DP_Code")).otherwise(pl.lit(None))
        ]).fill_null("")
        
        blaster_mapped = (
            pl.when(blaster_val.is_in(['PTP', 'Pending', 'Promise to Pay', 'PTP (Promise to Pay)', 'PTP - Token Amount', 'PTP (Promise to Pay) - Settlement', 'PTP - Settlement'])).then(pl.lit("Follow-up Required"))
            .when(blaster_val.is_in(['Refused to pay', 'Refused to Pay'])).then(pl.lit("Denied"))
            .when(blaster_val == 'Dispute').then(pl.lit("Payment Dispute"))
            .when(blaster_val == 'Call Back').then(pl.lit("Follow-up Required"))
            .when(blaster_val.is_in(['Answered', 'ANSWERED', 'Blaster Completed'])).then(pl.lit("No Response"))
            .when(blaster_val == 'Hang Up / No Response').then(pl.lit("Hang Up"))
            .when(blaster_val == 'Unaware / Communication Issue').then(pl.lit("Communication Issue"))
            .when(blaster_val == 'No Communication').then(pl.lit("No Answer"))
            .otherwise(blaster_val)
        )
        
        blaster_latest_dp_expr = (
            pl.when(blaster_mapped != "")
            .then(blaster_mapped)
            .otherwise(
                pl.when(pl.col("Blaster Calling Date").is_null()).then(pl.lit(""))
                .when(pl.col("blaster_connected_latest") == True).then(pl.lit("No Response"))
                .otherwise(pl.lit("No Answer"))
            )
        )

        pipeline_lf = (
            self.latest_master_df
            .lazy()
            .join(self.daily_pivots_df.lazy().select(["LoanMstID"] + pivot_cols), on="LoanMstID", how="left")
            .join(wa_received_any_loans_lf, on="LoanMstID", how="left")
            .with_columns([
                # DND
                pl.when(pl.col("DND").fill_null(0) == 1).then(pl.lit("Yes")).otherwise(pl.lit("No")).alias("DND"),
                
                # Attempts and Connects
                (pl.col("call_attempts").fill_null(0) + pl.col("blaster_attempts").fill_null(0) + pl.col("whatsapp_attempts").fill_null(0)).alias("Total VoiceBot+WhatsApp+Blaster"),
                (pl.col("calls_connected").fill_null(0) + pl.col("blaster_connected_count").fill_null(0) + pl.col("whatsapp_connected_count").fill_null(0)).alias("Total Connected VoiceBot+WhatsApp+Blaster"),
                (pl.col("call_attempts").fill_null(0) + pl.col("blaster_attempts").fill_null(0)).alias("Total Calls"),
                (pl.col("calls_connected").fill_null(0) + pl.col("blaster_connected_count").fill_null(0)).alias("Total Calls Connected"),
                pl.col("unique_total_attempt").alias("Unique Total Attempt"),
                pl.col("unique_total_connect").alias("Unique Total Connect"),
                pl.col("ai_connected_mtd").alias("VoiceBot Connected in this month"),
                pl.col("blaster_connected_mtd").alias("Connected in Blaster in this month"),
                pl.col("wa_connected_mtd").alias("Connected in WhatsApp in this month"),
                (pl.col("ai_ptp_count").fill_null(0) + pl.col("wa_ptp_count").fill_null(0)).alias("Total PTP"),
                
                # Last connected and durations
                pl.col("last_connected_date").alias("Last Connected Date (MTD)"),
                pl.col("last_connected_channel").alias("Last Connected Channel (MTD)"),
                pl.col("last_connected_call_duration").alias("Last Connected Voice Call Duration (Seconds)"),
                pl.col("max_connected_call_duration").alias("Max Connected Voice Call Duration (Seconds)"),
                
                # Best dispositions
                pl.col("best_disposition").fill_null("").alias("Best Disposition (MTD)"),
                pl.when(pl.col("best_disposition_ptp_date").is_not_null())
                .then(
                    pl.when(pl.col("best_source_call_disposition").fill_null("").str.strip_chars() == "").then(pl.lit("PTP (Promise To Pay)"))
                    .when(pl.col("best_source_call_disposition").str.to_lowercase().str.contains("promise")).then(pl.col("best_source_call_disposition"))
                    .otherwise(pl.lit("PTP (Promise To Pay)"))
                )
                .otherwise(pl.lit(""))
                .alias("PTP Type"),
                pl.col("best_category").fill_null("").alias("Best Disposition Category (MTD)"),
                pl.col("best_disposition_channel").fill_null("").alias("Best Disposition Channel (MTD)"),
                pl.col("best_disposition_date").alias("Best Disposition Date (MTD)"),
                pl.col("best_disposition_ptp_date").alias("Best Disposition MTD PTP Date"),
                pl.col("best_source_call_disposition").fill_null("").alias("Best VoiceBot Source ExtraColumn1"),
                
                # Latest dispositions
                pl.col("latest_disposition").fill_null("").alias("Latest Disposition (MTD)"),
                pl.col("latest_disposition_category").fill_null("").alias("Latest Disposition Category (MTD)"),
                pl.col("latest_disposition_channel").fill_null("").alias("Latest Disposition Channel (MTD)"),
                pl.col("latest_disposition_date").alias("Latest Disposition Date (MTD)"),
                
                # VoiceBot Call details
                pl.col("AI Calling Date").alias("VoiceBot Calling Date"),
                pl.col("call_attempts").fill_null(0).alias("VoiceBot Call Attempts (MTD)"),
                pl.col("calls_connected").fill_null(0).alias("VoiceBot Call Connected Count (MTD)"),
                pl.when(pl.col("call_attempts").fill_null(0) == 0).then(pl.lit(""))
                .when(pl.col("AI Calling Date").is_null()).then(pl.lit(""))
                .when(pl.col("connected_latest") == True).then(pl.lit("CN"))
                .when((pl.col("connected_latest") != True) & (pl.col("not_connected_latest") == True)).then(pl.lit("NC"))
                .otherwise(pl.lit(""))
                .alias("VoiceBot CN/NC"),
                pl.when(pl.col("connected_latest") == True).then(pl.lit(1)).otherwise(pl.lit(0)).alias("VoiceBot CN_Flag"),
                pl.when(pl.col("call_attempts").fill_null(0) == 0).then(pl.lit(""))
                .when(pl.col("AI Calling Date").is_null()).then(pl.lit(""))
                .otherwise(
                    pl.coalesce([
                        pl.col("voicebot_latest_disposition"),
                        pl.when(pl.col("connected_latest") == True).then(pl.lit("No Response")).otherwise(pl.lit("No Answer"))
                    ])
                )
                .alias("VoiceBot Latest DP"),
                
                # LM Best and call PTP
                pl.col("last_month_best_disposition").fill_null("").alias("LM Dispo."),
                pl.when(pl.col("call_has_ptp").fill_null(0) == 1)
                .then(pl.col("call_promise_datetime").cast(pl.Date))
                .otherwise(pl.lit(None))
                .alias("PTP Date (From VoiceBot)"),
                
                # Blaster
                pl.col("blaster_attempts").fill_null(0).alias("Blaster Call Attempts (MTD)"),
                pl.col("blaster_connected_count").fill_null(0).alias("Blaster Call Connected Count (MTD)"),
                pl.when(pl.col("Blaster Calling Date").is_null()).then(pl.lit(""))
                .when(pl.col("blaster_connected_latest") == True).then(pl.lit("CN"))
                .when((pl.col("blaster_connected_latest") != True) & (pl.col("blaster_not_connected_latest") == True)).then(pl.lit("NC"))
                .otherwise(pl.lit(""))
                .alias("Blaster CN/NC"),
                pl.when(pl.col("blaster_connected_latest") == True).then(pl.lit(1)).otherwise(pl.lit(0)).alias("Blaster CN_Flag"),
                blaster_latest_dp_expr.alias("Blaster Latest DP"),
                
                # WhatsApp
                pl.col("Whatsapp Sent Date").alias("WhatsApp SMS sent date"),
                pl.col("whatsapp_attempts").fill_null(0).alias("WhatsApp Attempts (MTD)"),
                pl.col("whatsapp_delivered_count").fill_null(0).alias("WhatsApp Delivered Count (MTD)"),
                pl.col("whatsapp_read_count").fill_null(0).alias("WhatsApp Read Count (MTD)"),
                pl.col("whatsapp_connected_count").fill_null(0).alias("WhatsApp Connected Count (MTD)"),
                
                pl.when(
                    (pl.col("last_wa_read") == True) |
                    (pl.col("last_wa_delivered") == True) |
                    (pl.col("last_wa_sent") == True)
                ).then(pl.lit(1)).otherwise(pl.lit(0)).alias("WA Sent"),
                
                pl.when(
                    (pl.col("last_wa_sent") == True) &
                    (
                        (pl.col("last_wa_read") == True) |
                        (pl.col("last_wa_delivered") == True) |
                        (pl.col("has_wa_received") == 1) |
                        pl.col("latest_resp_whatsapp_qid").is_not_null()
                    )
                ).then(pl.lit(1)).otherwise(pl.lit(0)).alias("WA Delivered"),
                
                pl.when(
                    (pl.col("last_wa_sent") == True) &
                    (
                        (pl.col("last_wa_read") == True) |
                        (pl.col("has_wa_received") == 1) |
                        pl.col("latest_resp_whatsapp_qid").is_not_null()
                    )
                ).then(pl.lit(1)).otherwise(pl.lit(0)).alias("WA Read"),
                
                pl.when(pl.col("whatsapp_attempts").fill_null(0) == 0).then(pl.lit("")).otherwise(pl.col("whatsapp_latest_disposition").fill_null("")).alias("DP Code (WhatsApp)"),
                pl.when(pl.col("whatsapp_latest_disposition") == 'PTP - WhatsApp').then(pl.col("whatsapp_latest_disposition_ptp_date")).otherwise(pl.lit(None)).alias("PTP Date (From WhatsApp)"),
                
                # Collections & EMI
                pl.col("Last Collected Amount").fill_null(0.0).alias("Last Collected Amount"),
                pl.when(pl.col("Total Collection Amt").fill_null(0.0) >= pl.col("EMI Amount").fill_null(0.0)).then(pl.col("Total Collection Amt").fill_null(0.0)).otherwise(pl.lit(0.0)).alias("Collection >=1 EMI"),
                pl.when(pl.col("Total Collection Amt").fill_null(0.0) >= pl.col("EMI Amount").fill_null(0.0)).then(pl.col("EMI Amount").fill_null(0.0)).otherwise(pl.lit(0.0)).alias("Resolve POS")
            ])
        )

        # Step 3: Project final columns
        projection_cols = [
            pl.col("Account Number"),
            pl.col("Customer Name"),
            pl.col("Customer Number"),
            pl.col("Primary Mobile Number"),
            pl.col("Branch Name"),
            pl.when(pl.col("Branch Code").is_null()).then(pl.lit(None))
              .otherwise(pl.concat_str([pl.lit("'"), pl.col("Branch Code"), pl.lit("'")]))
              .alias("Branch Code"),
            pl.col("Pool Type"),
            pl.col("Loan Status"),
            pl.col("DND"),
            pl.col("Default Amt"),
            pl.col("EMI Amount"),
            pl.col("Total Outstanding"),
            pl.col("Total VoiceBot+WhatsApp+Blaster"),
            pl.col("Total Connected VoiceBot+WhatsApp+Blaster"),
            pl.col("Total Calls"),
            pl.col("Total Calls Connected"),
            pl.col("Unique Total Attempt"),
            pl.col("Unique Total Connect"),
            pl.col("VoiceBot Connected in this month"),
            pl.col("Connected in Blaster in this month"),
            pl.col("Connected in WhatsApp in this month"),
            pl.col("Total PTP"),
            pl.col("Last Connected Date (MTD)"),
            pl.col("Last Connected Channel (MTD)"),
            pl.col("Last Connected Voice Call Duration (Seconds)"),
            pl.col("Max Connected Voice Call Duration (Seconds)"),
            pl.col("Best Disposition (MTD)"),
            pl.col("PTP Type"),
            pl.col("Best Disposition Category (MTD)"),
            pl.col("Best Disposition Channel (MTD)"),
            pl.col("Best Disposition Date (MTD)"),
            pl.col("Best Disposition MTD PTP Date"),
            pl.col("Best VoiceBot Source ExtraColumn1"),
            pl.col("Latest Disposition (MTD)"),
            pl.col("Latest Disposition Category (MTD)"),
            pl.col("Latest Disposition Channel (MTD)"),
            pl.col("Latest Disposition Date (MTD)"),
            pl.col("Calling Date"),
            pl.col("VoiceBot Calling Date"),
            pl.col("VoiceBot Call Attempts (MTD)"),
            pl.col("VoiceBot Call Connected Count (MTD)"),
            pl.col("VoiceBot CN/NC"),
            pl.col("VoiceBot CN_Flag"),
            pl.col("VoiceBot Latest DP"),
            pl.col("LM Dispo."),
            pl.col("PTP Date (From VoiceBot)"),
            pl.col("Blaster Calling Date"),
            pl.col("Blaster Call Attempts (MTD)"),
            pl.col("Blaster Call Connected Count (MTD)"),
            pl.col("Blaster CN/NC"),
            pl.col("Blaster CN_Flag"),
            pl.col("Blaster Latest DP"),
            pl.col("WhatsApp SMS sent date"),
            pl.col("WhatsApp Attempts (MTD)"),
            pl.col("WhatsApp Delivered Count (MTD)"),
            pl.col("WhatsApp Read Count (MTD)"),
            pl.col("WhatsApp Connected Count (MTD)"),
            pl.col("WA Sent"),
            pl.col("WA Delivered"),
            pl.col("WA Read"),
            pl.col("DP Code (WhatsApp)"),
            pl.col("PTP Date (From WhatsApp)"),
            pl.col("All WhatsApp Messages"),
            pl.col("Last Collected Amount"),
            pl.col("Last Collection Date"),
            pl.col("Total Collection Amt")
        ]
        
        # Add pivot columns
        projection_cols.extend([pl.col(c) for c in pivot_cols])
        
        # Add final Collection indicators
        projection_cols.extend([
            pl.col("Collection >=1 EMI"),
            pl.col("Resolve POS")
        ])

        self.final_report_df = (
            pipeline_lf
            .select(projection_cols)
            .sort(by="Account Number", descending=False)
            .collect()
        )

        # Logging & Assertions
        assert len(self.final_report_df) == len(self.latest_master_df), "Fidelity Check Failed: final_report_df row count does not match latest_master_df!"
        assert self.final_report_df["Account Number"].n_unique() == len(self.final_report_df), "Duplicate Account Number detected in final_report_df!"
        
        self.logger.info("Validation Log - final_report_df shape: %s", self.final_report_df.shape)
        self.logger.info("Validation Log - final_report_df columns: %s", self.final_report_df.columns)
        self.logger.info("Validation Log - final_report_df estimated memory: %d bytes (%.4f KB)", self.final_report_df.estimated_size(), self.final_report_df.estimated_size() / 1024.0)


    @timer
    def export_final_report(self) -> str:
        """
        Validates and exports the final_report_df to an Excel workbook with formatting.
        """
        import os
        import time
        self.logger.info("Starting export_final_report...")
        
        # Step 1: Validation
        if self.final_report_df is None:
            raise ValueError("final_report_df is not available. Call build_final_report first.")
        if len(self.final_report_df) == 0:
            raise ValueError("final_report_df is empty. Nothing to export.")
            
        # File naming & Output path resolution
        if not self.request:
            raise ValueError("ReportRequest context is missing.")
            
        if self.request.output_path:
            output_path = self.request.output_path
        else:
            start_str = self.request.start_date.strftime("%Y-%m-%d")
            end_str = self.request.end_date.strftime("%Y-%m-%d")
            output_path = f"Fusion_Report_{start_str}_to_{end_str}.xlsx"
            
        # Ensure output directories exist
        parent_dir = os.path.dirname(output_path)
        if parent_dir:
            os.makedirs(parent_dir, exist_ok=True)
            
        # Step 2: Excel Export using openpyxl
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import date, datetime
        
        self.logger.info("Creating Excel workbook...")
        start_time = time.perf_counter()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fusion Report"
        
        # Freeze first row
        ws.freeze_panes = "A2"
        
        headers = self.final_report_df.columns
        
        # Write bold & aligned header row
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            
        # Write values keeping data types (dates, datetimes, floats, ints, strings)
        row_idx = 2
        for row_data in self.final_report_df.iter_rows():
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Check value type
                if val is None:
                    cell.value = None
                elif isinstance(val, date) and not isinstance(val, datetime):
                    cell.value = val
                    cell.number_format = 'yyyy-mm-dd'
                elif isinstance(val, datetime):
                    cell.value = val
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                elif isinstance(val, (int, float)):
                    cell.value = val
                else:
                    cell.value = str(val)
            row_idx += 1
            
        # Enable auto-filter on the exact grid range
        last_col_letter = get_column_letter(len(headers))
        total_rows = len(self.final_report_df)
        ws.auto_filter.ref = f"A1:{last_col_letter}{total_rows + 1}"
        
        # Auto-adjust column widths based on cell length (cap at 50, pad by 3)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    if isinstance(cell.value, datetime):
                        val_str = cell.value.strftime("%Y-%m-%d %H:%M:%S")
                    elif isinstance(cell.value, date):
                        val_str = cell.value.strftime("%Y-%m-%d")
                    else:
                        val_str = str(cell.value)
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        # Save output workbook
        wb.save(output_path)
        write_duration = time.perf_counter() - start_time
        
        # Log metadata & size
        file_size = os.path.getsize(output_path)
        self.logger.info("Successfully exported report to %s", output_path)
        self.logger.info("Export metadata - Rows: %d, Columns: %d, File Size: %d bytes (%.2f KB), Duration: %.4f seconds", 
                         total_rows, len(headers), file_size, file_size / 1024.0, write_duration)
                         
        return output_path


    @timer
    def build_latest_channels(self) -> None:
        """
        Processes and aggregates the latest communication attempt details across channels.
        """
        self.logger.info("Starting build_latest_channels...")
        
        # Verify required tables exist
        if self.voicebot_mtd_base_df is None:
            raise ValueError("voicebot_mtd_base_df is not available. Ensure build_all_call_dispositions has executed.")
        if self.voicebot_response_map_df is None:
            raise ValueError("voicebot_response_map_df is not available. Ensure build_all_call_dispositions has executed.")
        if self.whatsapp_mtd_base_df is None:
            raise ValueError("whatsapp_mtd_base_df is not available. Ensure build_all_wa_dispositions has executed.")
        if self.whatsapp_response_map_df is None:
            raise ValueError("whatsapp_response_map_df is not available. Ensure build_all_wa_dispositions has executed.")
        if self.blaster_mtd_base_df is None:
            raise ValueError("blaster_mtd_base_df is not available. Ensure build_all_blaster_dispositions has executed.")
        if self.whatsapp_messages_df is None:
            raise ValueError("whatsapp_messages_df is not available. Ensure load_source_tables has executed.")
        if self.base_accounts_df is None:
            raise ValueError("base_accounts_df is not available.")

        # -------------------------------------------------------------------------
        # STEP 1: voicebot_attempt_totals_df
        # -------------------------------------------------------------------------
        call_tried_col_v = "CallTried" if "CallTried" in self.voicebot_mtd_base_df.columns else "calltried"
        self.voicebot_attempt_totals_df = (
            self.voicebot_mtd_base_df
            .lazy()
            .group_by("LoanMstID")
            .agg(pl.col(call_tried_col_v).fill_null(0).sum().alias("call_attempts"))
            .collect()
        )
        self.logger.info("voicebot_attempt_totals_df size: %d", len(self.voicebot_attempt_totals_df))

        # -------------------------------------------------------------------------
        # STEP 2: latest_call_df
        # -------------------------------------------------------------------------
        if "DisbursementID" in self.voicebot_mtd_base_df.columns:
            disb_col_v = "DisbursementID"
        elif "disbursementid" in self.voicebot_mtd_base_df.columns:
            disb_col_v = "disbursementid"
        else:
            disb_col_v = "DisbursementID"
            self.voicebot_mtd_base_df = (
                self.voicebot_mtd_base_df
                .join(
                    self.base_accounts_df.select(["LoanMstID", "DisbursementID"]),
                    on="LoanMstID",
                    how="left"
                )
            )

        ext_col1 = "Extracolumn1" if "Extracolumn1" in self.voicebot_mtd_base_df.columns else "extracolumn1"
        dispo_col = "deposition" if "deposition" in self.voicebot_mtd_base_df.columns else "Disposition" if "Disposition" in self.voicebot_mtd_base_df.columns else "deposition"
        
        conn_disp_src = pl.coalesce([
            pl.col(ext_col1).str.strip_chars(),
            pl.col(dispo_col).str.strip_chars()
        ]).fill_null("")
        
        conn_disp_src_upper = conn_disp_src.str.to_uppercase()
        conn_disp_src_lower = conn_disp_src.str.to_lowercase()

        dp_code_connected = (
            pl.when(conn_disp_src.str.strip_chars() == "").then(pl.lit("No Response"))
            .when(conn_disp_src_upper.str.contains("ALREADY PAID") | (conn_disp_src_lower == "already_paid")).then(pl.lit("Already Paid"))
            .when(conn_disp_src_upper.str.contains("DENIED") | conn_disp_src_upper.str.contains("DENIES") | conn_disp_src_upper.str.contains("DENIAL") | conn_disp_src_upper.str.contains("REFUSED")).then(pl.lit("Refused to pay"))
            .when(conn_disp_src_lower == "not_ready_to_pay").then(pl.lit("Refused to pay"))
            .when(conn_disp_src_upper.str.contains("CONFUSION")).then(pl.lit("Confusion"))
            .when(conn_disp_src_upper.str.contains("PAYMENT DISPUTE")).then(pl.lit("Payment Dispute"))
            .when(conn_disp_src_upper.str.contains("INSURANCE")).then(pl.lit("Insurance Dispute"))
            .when(conn_disp_src_upper.str.contains("DISPUTE") | (conn_disp_src_lower == "dispute_raised")).then(pl.lit("Dispute"))
            .when(conn_disp_src_upper.str.contains("FOLLOW_UP_REQUIRED") | (conn_disp_src_lower == "follow_up_required")).then(pl.lit("Follow-up Required"))
            .when(conn_disp_src_lower.is_in(["callback", "call back"])).then(pl.lit("Call Back"))
            .when(conn_disp_src_lower == "inquiry").then(pl.lit("Inquiry"))
            .when(conn_disp_src_lower == "payment inquiry").then(pl.lit("Payment Inquiry"))
            .when(conn_disp_src_lower == "overdue inquiry").then(pl.lit("Overdue Inquiry"))
            .when(conn_disp_src_lower == "payment difficulty").then(pl.lit("Payment Difficulty"))
            .when(conn_disp_src_lower == "payment arrangement").then(pl.lit("Payment Arrangement"))
            .when(conn_disp_src_lower == "greeting/confirmation").then(pl.lit("Greeting/Confirmation"))
            .when(conn_disp_src_lower == "family/third party").then(pl.lit("Family/Third Party"))
            .when(conn_disp_src_lower == "contact verification").then(pl.lit("Contact Verification"))
            .when(conn_disp_src_lower == "no request").then(pl.lit("No Request"))
            .when(conn_disp_src_lower == "unclear intent").then(pl.lit("Unclear Intent"))
            .when(conn_disp_src_lower == "service request").then(pl.lit("Service Request"))
            .when(conn_disp_src_lower == "complaint/escalation").then(pl.lit("Complaint/Escalation"))
            .when(conn_disp_src_lower == "account issue").then(pl.lit("Account Issue"))
            .when(conn_disp_src_upper.str.contains("TOKEN_AMOUNT")).then(pl.lit("PTP - Token Amount"))
            .when(
                conn_disp_src_upper.str.contains("SETTLEMENT") &
                (
                    conn_disp_src_upper.str.contains("PTP") |
                    conn_disp_src_upper.str.contains("PROMISE") |
                    conn_disp_src_upper.str.contains("PENDING") |
                    conn_disp_src_lower.is_in(["futurepay", "ready_to_pay", "willing_to_pay"])
                )
            ).then(pl.lit("PTP - Settlement"))
            .when(conn_disp_src_upper.str.contains("SETTLEMENT") & (pl.col("CallDuration") < 60)).then(pl.lit("Other"))
            .when(conn_disp_src_lower.is_in(["futurepay", "ready_to_pay", "willing_to_pay"]) | conn_disp_src_upper.str.contains("WILL PAY AFTER SOMETIME")).then(
                pl.when(pl.col("promise_datetime").is_not_null()).then(pl.lit("PTP - Date")).otherwise(pl.lit("PTP"))
            )
            .when(
                conn_disp_src_upper.str.contains("PENDING") |
                conn_disp_src_upper.str.contains("PTP") |
                conn_disp_src_upper.str.contains("PROMISE") |
                conn_disp_src_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                conn_disp_src_upper.str.contains("NO EARLY PAYMENT") |
                conn_disp_src_upper.str.contains("CONFIRM") |
                (conn_disp_src_lower == "positive")
            ).then(
                pl.when(pl.col("promise_datetime").is_not_null()).then(pl.lit("PTP - Date")).otherwise(pl.lit("PTP"))
            )
            .when(conn_disp_src_lower.is_in(["wrongnumber", "wrong number"])).then(pl.lit("Wrong Contact"))
            .when(conn_disp_src_upper.str.contains("WRONG CONTACT") | conn_disp_src_upper.str.contains("WRONG NUMBER")).then(pl.lit("Wrong Contact"))
            .when(conn_disp_src_upper.str.contains("UNAWARE") | conn_disp_src_upper.str.contains("COMMUNICATION ISSUE") | conn_disp_src_upper.str.contains("LANGUAGE")).then(pl.lit("Unaware / Communication Issue"))
            .when(conn_disp_src_upper.str.contains("FAMILY ISSUE") | conn_disp_src_upper.str.contains("HEALTH")).then(pl.lit("Health / Family Issue"))
            .when(conn_disp_src_lower.is_in(["collectionofficer missing", "collection officer missing"])).then(pl.lit("Collection Officer Missing"))
            .when(conn_disp_src_upper.str.contains("OTHER")).then(pl.lit("Other"))
            .when(conn_disp_src_lower == "neutral").then(pl.lit("Neutral"))
            .otherwise(conn_disp_src.str.replace_all("(?i)busy", "Busy"))
        )

        ext_col1_val = pl.col(ext_col1).fill_null("")
        ext_col1_upper = ext_col1_val.str.to_uppercase()
        ext_col1_compact = (
            ext_col1_upper
            .str.replace_all(r"^[0-9]+[_ \t-]*", "")
            .str.replace_all(r"[^A-Z0-9]", "")
        )
        
        not_conn_list = [
            'NETWORKCONGESTION', 'NODOESNOTEXIST', 'ONLYRINGING',
            'SWITCHEDOFF', 'SPEAKINGTOSOMEONEELSE', 'USEROUTOFFCOVERAGE',
            'CALLREJECTED', 'OUTOFSERVICE', 'INCOMINGNOTAVAILABLE', 'BUSY'
        ]
        
        dp_code_not_connected_expr = (
            pl.when(ext_col1_compact.is_in(not_conn_list)).then(
                pl.when(ext_col1_compact == 'NETWORKCONGESTION').then(pl.lit('Network Congestion'))
                .when(ext_col1_compact == 'NODOESNOTEXIST').then(pl.lit('Number Does Not Exist'))
                .when(ext_col1_compact == 'ONLYRINGING').then(pl.lit('Only Ringing'))
                .when(ext_col1_compact == 'SWITCHEDOFF').then(pl.lit('Switched Off'))
                .when(ext_col1_compact == 'SPEAKINGTOSOMEONEELSE').then(pl.lit('Speaking To Someone Else'))
                .when(ext_col1_compact == 'USEROUTOFFCOVERAGE').then(pl.lit('User Out Of Coverage'))
                .when(ext_col1_compact == 'CALLREJECTED').then(pl.lit('Call Rejected'))
                .when(ext_col1_compact == 'OUTOFSERVICE').then(pl.lit('Out Of Service'))
                .when(ext_col1_compact == 'INCOMINGNOTAVAILABLE').then(pl.lit('Incoming Not Available'))
                .when(ext_col1_compact == 'BUSY').then(pl.lit('Busy'))
                .otherwise(pl.lit('Not Contactable'))
            )
            .otherwise(ext_col1_val.str.replace_all("(?i)busy", "Busy"))
        )

        dp_code_raw = (
            pl.when(pl.col("is_connected") == True)
            .then(dp_code_connected)
            .otherwise(dp_code_not_connected_expr)
        )
        
        dp_code_cleaned = dp_code_raw.str.strip_chars().str.replace_all(r":+[ \t]*$", "")

        extra_category_connected = (
            pl.when(conn_disp_src.str.strip_chars() == "").then(pl.lit("No Response"))
            .when(conn_disp_src_upper.str.contains("ALREADY PAID") | (conn_disp_src_lower == "already_paid")).then(pl.lit("Already Paid"))
            .when(conn_disp_src_upper.str.contains("DENIED") | conn_disp_src_upper.str.contains("DENIES") | conn_disp_src_upper.str.contains("DENIAL") | conn_disp_src_upper.str.contains("REFUSED")).then(pl.lit("Refused to pay"))
            .when(conn_disp_src_lower == "not_ready_to_pay").then(pl.lit("Refused to pay"))
            .when(conn_disp_src_upper.str.contains("CONFUSION")).then(pl.lit("Confusion"))
            .when(conn_disp_src_upper.str.contains("PAYMENT DISPUTE")).then(pl.lit("Payment Dispute"))
            .when(conn_disp_src_upper.str.contains("INSURANCE")).then(pl.lit("Insurance Dispute"))
            .when(conn_disp_src_upper.str.contains("DISPUTE") | (conn_disp_src_lower == "dispute_raised")).then(pl.lit("Dispute"))
            .when(conn_disp_src_upper.str.contains("FOLLOW_UP_REQUIRED") | (conn_disp_src_lower == "follow_up_required")).then(pl.lit("Follow-up Required"))
            .when(conn_disp_src_lower.is_in(["callback", "call back"])).then(pl.lit("Call Back"))
            .when(conn_disp_src_lower == "inquiry").then(pl.lit("Inquiry"))
            .when(conn_disp_src_lower == "payment inquiry").then(pl.lit("Payment Inquiry"))
            .when(conn_disp_src_lower == "overdue inquiry").then(pl.lit("Overdue Inquiry"))
            .when(conn_disp_src_lower == "payment difficulty").then(pl.lit("Payment Difficulty"))
            .when(conn_disp_src_lower == "payment arrangement").then(pl.lit("Payment Arrangement"))
            .when(conn_disp_src_lower == "greeting/confirmation").then(pl.lit("Greeting/Confirmation"))
            .when(conn_disp_src_lower == "family/third party").then(pl.lit("Family/Third Party"))
            .when(conn_disp_src_lower == "contact verification").then(pl.lit("Contact Verification"))
            .when(conn_disp_src_lower == "no request").then(pl.lit("No Request"))
            .when(conn_disp_src_lower == "unclear intent").then(pl.lit("Unclear Intent"))
            .when(conn_disp_src_lower == "service request").then(pl.lit("Service Request"))
            .when(conn_disp_src_lower == "complaint/escalation").then(pl.lit("Complaint/Escalation"))
            .when(conn_disp_src_lower == "account issue").then(pl.lit("Account Issue"))
            .when(
                conn_disp_src_upper.str.contains("TOKEN") &
                (
                    conn_disp_src_upper.str.contains("PTP") |
                    conn_disp_src_upper.str.contains("PROMISE") |
                    conn_disp_src_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP - Token Amount"))
            .when(
                conn_disp_src_upper.str.contains("SETTLEMENT") &
                (
                    conn_disp_src_upper.str.contains("PTP") |
                    conn_disp_src_upper.str.contains("PROMISE") |
                    conn_disp_src_upper.str.contains("PENDING") |
                    conn_disp_src_lower.is_in(["futurepay", "ready_to_pay", "willing_to_pay"])
                )
            ).then(pl.lit("PTP (Promise to Pay) - Settlement"))
            .when(conn_disp_src_upper.str.contains("SETTLEMENT") & (pl.col("CallDuration") < 60)).then(pl.lit("Other"))
            .when(conn_disp_src_lower.is_in(["futurepay", "ready_to_pay", "willing_to_pay"]) | conn_disp_src_upper.str.contains("WILL PAY AFTER SOMETIME")).then(
                pl.when(pl.col("promise_datetime").is_not_null()).then(pl.lit("PTP - Date")).otherwise(pl.lit("PTP"))
            )
            .when(
                conn_disp_src_upper.str.contains("PENDING") |
                conn_disp_src_upper.str.contains("PTP") |
                conn_disp_src_upper.str.contains("PROMISE") |
                conn_disp_src_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                conn_disp_src_upper.str.contains("NO EARLY PAYMENT") |
                conn_disp_src_upper.str.contains("CONFIRM") |
                (conn_disp_src_lower == "positive")
            ).then(
                pl.when(pl.col("promise_datetime").is_not_null()).then(pl.lit("PTP - Date")).otherwise(pl.lit("PTP"))
            )
            .when(conn_disp_src_lower.is_in(["wrongnumber", "wrong number"])).then(pl.lit("Wrong Contact"))
            .when(conn_disp_src_upper.str.contains("WRONG CONTACT") | conn_disp_src_upper.str.contains("WRONG NUMBER")).then(pl.lit("Wrong Contact"))
            .when(conn_disp_src_upper.str.contains("UNAWARE") | conn_disp_src_upper.str.contains("COMMUNICATION ISSUE") | conn_disp_src_upper.str.contains("LANGUAGE")).then(pl.lit("Unaware / Communication Issue"))
            .when(conn_disp_src_upper.str.contains("FAMILY ISSUE") | conn_disp_src_upper.str.contains("HEALTH")).then(pl.lit("Health / Family Issue"))
            .when(conn_disp_src_lower.is_in(["collectionofficer missing", "collection officer missing"])).then(pl.lit("Collection Officer Missing"))
            .when(conn_disp_src_upper.str.contains("OTHER")).then(pl.lit("Other"))
            .when(conn_disp_src_lower == "neutral").then(pl.lit("Neutral"))
            .otherwise(conn_disp_src.str.replace_all("(?i)busy", "Busy"))
        )

        extra_category_raw = (
            pl.when(pl.col("is_connected") == True)
            .then(extra_category_connected)
            .otherwise(dp_code_not_connected_expr)
        )
        extra_category_cleaned = extra_category_raw.str.strip_chars().str.replace_all(r":+[ \t]*$", "")

        created_date_col = "CreatedDate" if "CreatedDate" in self.voicebot_mtd_base_df.columns else "createddate"
        recording_col = "Recording" if "Recording" in self.voicebot_mtd_base_df.columns else "recording"
        queue_col_v = "VoiceBotQueueID" if "VoiceBotQueueID" in self.voicebot_mtd_base_df.columns else "voicebotqueueid"
        disb_col_v = "DisbursementID" if "DisbursementID" in self.voicebot_mtd_base_df.columns else "disbursementid"

        calls_not_connected_expr = (pl.col("is_connected") == False).sum().over("LoanMstID")
        calls_connected_expr = (pl.col("is_connected") == True).sum().over("LoanMstID")

        filtered_calls = (
            self.voicebot_mtd_base_df
            .lazy()
            .join(
                self.voicebot_response_map_df.lazy(),
                on="VoiceBotQueueID",
                how="left"
            )
            .with_columns([
                dp_code_cleaned.alias("DP_Code_Not_Connected"),
                extra_category_cleaned.alias("Extra_DP_Category"),
                pl.col("ExtraColumn2").alias("Extra_DP_Reason"),
                pl.col("has_paid").fill_null(0),
                pl.col("has_denied").fill_null(0),
                pl.col("has_ptp").fill_null(0),
                calls_not_connected_expr.alias("calls_not_connected"),
                calls_connected_expr.alias("calls_connected")
            ])
            .sort(by=["LoanMstID", created_date_col], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
        )

        rec_col = pl.col(recording_col)
        connected_latest_expr = pl.when(rec_col.is_not_null() & (rec_col.str.len_bytes() > 1)).then(pl.lit(True)).otherwise(pl.lit(False))
        not_connected_latest_expr = pl.when(rec_col.is_null() | (rec_col == "")).then(pl.lit(True)).otherwise(pl.lit(False))

        self.latest_call_df = (
            filtered_calls
            .join(
                self.voicebot_attempt_totals_df.lazy(),
                on="LoanMstID",
                how="left"
            )
            .select([
                pl.col("LoanMstID"),
                pl.col(disb_col_v).alias("DisbursementID"),
                pl.col("call_date").alias("AI Calling Date"),
                rec_col.alias("Recording"),
                pl.col("VoiceBotQueueID"),
                pl.col("DP_Code_Not_Connected"),
                pl.col("Extra_DP_Category"),
                pl.col("Extra_DP_Reason"),
                connected_latest_expr.alias("connected_latest"),
                not_connected_latest_expr.alias("not_connected_latest"),
                pl.col("has_paid").fill_null(0).alias("call_has_paid"),
                pl.col("has_denied").fill_null(0).alias("call_has_denied"),
                pl.col("has_ptp").fill_null(0).alias("call_has_ptp"),
                pl.col("promise_datetime").alias("call_promise_datetime"),
                pl.col("call_attempts").fill_null(0).alias("call_attempts"),
                pl.col("calls_not_connected").fill_null(0).alias("calls_not_connected"),
                pl.col("calls_connected").fill_null(0).alias("calls_connected")
            ])
            .collect()
        )
        self.logger.info("latest_call_df size: %d", len(self.latest_call_df))

        # -------------------------------------------------------------------------
        # STEP 3: latest_call_extra_df
        # -------------------------------------------------------------------------
        self.latest_call_extra_df = (
            self.voicebot_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .sort(by=["LoanMstID", created_date_col], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("Conclusion"),
                pl.col(recording_col).alias("Recording")
            ])
            .collect()
        )
        self.logger.info("latest_call_extra_df size: %d", len(self.latest_call_extra_df))

        # -------------------------------------------------------------------------
        # STEP 4: blaster_attempt_totals_df
        # -------------------------------------------------------------------------
        call_tried_col_bl = "CallTried" if "CallTried" in self.blaster_mtd_base_df.columns else "calltried"
        self.blaster_attempt_totals_df = (
            self.blaster_mtd_base_df
            .lazy()
            .group_by("LoanMstID")
            .agg(pl.col(call_tried_col_bl).fill_null(0).sum().alias("blaster_attempts"))
            .collect()
        )
        self.logger.info("blaster_attempt_totals_df size: %d", len(self.blaster_attempt_totals_df))

        # -------------------------------------------------------------------------
        # STEP 5: latest_blaster_df
        # -------------------------------------------------------------------------
        if "DisbursementID" in self.blaster_mtd_base_df.columns:
            disb_col_bl = "DisbursementID"
        elif "disbursementid" in self.blaster_mtd_base_df.columns:
            disb_col_bl = "disbursementid"
        else:
            disb_col_bl = "DisbursementID"
            self.blaster_mtd_base_df = (
                self.blaster_mtd_base_df
                .join(
                    self.base_accounts_df.select(["LoanMstID", "DisbursementID"]),
                    on="LoanMstID",
                    how="left"
                )
            )

        ext_col1_bl = "ExtraColumn1" if "ExtraColumn1" in self.blaster_mtd_base_df.columns else "extracolumn1"
        ext_col1_bl_val = pl.col(ext_col1_bl).str.strip_chars().fill_null("")
        ext_col1_bl_upper = ext_col1_bl_val.str.to_uppercase()
        
        ext_col1_bl_compact = (
            ext_col1_bl_upper
            .str.replace_all(r"^[0-9]+[_ \t-]*", "")
            .str.replace_all(r"[^A-Z0-9]", "")
        )

        blaster_dp_connected = (
            pl.when(
                ext_col1_bl_upper.str.contains("SETTLEMENT") &
                (ext_col1_bl_upper.str.contains("PTP") == False) &
                (ext_col1_bl_upper.str.contains("PROMISE") == False) &
                (ext_col1_bl_upper.str.contains("PENDING") == False) &
                (pl.col("CallDuration") < 60)
            ).then(pl.lit("Other"))
            .otherwise(
                pl.when(ext_col1_bl_val != "").then(ext_col1_bl_val).otherwise(pl.lit("No Response"))
            )
        )

        blaster_dp_not_connected = (
            pl.when(ext_col1_bl_compact.is_in(not_conn_list)).then(
                pl.when(ext_col1_bl_compact == 'NETWORKCONGESTION').then(pl.lit('Network Congestion'))
                .when(ext_col1_bl_compact == 'NODOESNOTEXIST').then(pl.lit('Number Does Not Exist'))
                .when(ext_col1_bl_compact == 'ONLYRINGING').then(pl.lit('Only Ringing'))
                .when(ext_col1_bl_compact == 'SWITCHEDOFF').then(pl.lit('Switched Off'))
                .when(ext_col1_bl_compact == 'SPEAKINGTOSOMEONEELSE').then(pl.lit('Speaking To Someone Else'))
                .when(ext_col1_bl_compact == 'USEROUTOFFCOVERAGE').then(pl.lit('User Out Of Coverage'))
                .when(ext_col1_bl_compact == 'CALLREJECTED').then(pl.lit('Call Rejected'))
                .when(ext_col1_bl_compact == 'OUTOFSERVICE').then(pl.lit('Out Of Service'))
                .when(ext_col1_bl_compact == 'INCOMINGNOTAVAILABLE').then(pl.lit('Incoming Not Available'))
                .when(ext_col1_bl_compact == 'BUSY').then(pl.lit('Busy'))
                .otherwise(pl.lit('Not Contactable'))
            )
            .otherwise(
                pl.when(ext_col1_bl_val != "").then(ext_col1_bl_val).otherwise(pl.lit("No Communication"))
            )
        )

        blaster_dp_raw = (
            pl.when(pl.col("is_connected") == True)
            .then(blaster_dp_connected)
            .otherwise(blaster_dp_not_connected)
        )
        
        blaster_dp_cleaned = blaster_dp_raw.str.strip_chars().str.replace_all(r":+[ \t]*$", "")

        blaster_not_connected_count_expr = (pl.col("is_connected") == False).sum().over("LoanMstID")
        blaster_connected_count_expr = (pl.col("is_connected") == True).sum().over("LoanMstID")

        created_date_col_bl = "CreatedDate" if "CreatedDate" in self.blaster_mtd_base_df.columns else "createddate"
        recording_col_bl = "Recording" if "Recording" in self.blaster_mtd_base_df.columns else "recording"
        queue_col_bl = "BlasterQueueID" if "BlasterQueueID" in self.blaster_mtd_base_df.columns else "blasterqueueid"
        disb_col_bl = "DisbursementID" if "DisbursementID" in self.blaster_mtd_base_df.columns else "disbursementid"

        filtered_blaster = (
            self.blaster_mtd_base_df
            .lazy()
            .with_columns([
                blaster_dp_cleaned.alias("Blaster_DP_Code"),
                blaster_dp_cleaned.alias("Blaster_DP_Category"),
                pl.col("Conclusion").alias("Blaster_Conclusion"),
                pl.col("CallDuration").alias("Blaster_CallDuration"),
                blaster_not_connected_count_expr.alias("blaster_not_connected_count"),
                blaster_connected_count_expr.alias("blaster_connected_count")
            ])
            .sort(by=["LoanMstID", created_date_col_bl], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
        )

        self.latest_blaster_df = (
            filtered_blaster
            .join(
                self.blaster_attempt_totals_df.lazy(),
                on="LoanMstID",
                how="left"
            )
            .select([
                pl.col("LoanMstID"),
                pl.col(disb_col_bl).alias("DisbursementID"),
                pl.col("blaster_date").alias("Blaster Calling Date"),
                pl.col(recording_col_bl).alias("Recording"),
                pl.col(queue_col_bl).alias("BlasterQueueID"),
                pl.col("Blaster_DP_Code"),
                pl.col("Blaster_DP_Category"),
                pl.col("Blaster_Conclusion"),
                pl.col("Blaster_CallDuration"),
                pl.col("is_connected").alias("blaster_connected_latest"),
                pl.when(pl.col("is_connected") == True).then(pl.lit(False)).otherwise(pl.lit(True)).alias("blaster_not_connected_latest"),
                pl.col("blaster_attempts").fill_null(0).alias("blaster_attempts"),
                pl.col("blaster_not_connected_count").fill_null(0).alias("blaster_not_connected_count"),
                pl.col("blaster_connected_count").fill_null(0).alias("blaster_connected_count")
            ])
            .collect()
        )
        self.logger.info("latest_blaster_df size: %d", len(self.latest_blaster_df))

        # -------------------------------------------------------------------------
        # STEP 6: latest_whatsapp_df
        # -------------------------------------------------------------------------
        if "received_after_send" not in self.whatsapp_mtd_base_df.columns:
            self.whatsapp_mtd_base_df = self.whatsapp_mtd_base_df.with_columns(
                pl.lit(False).alias("received_after_send")
            )

        if "DisbursementID" in self.whatsapp_mtd_base_df.columns:
            disb_col_wa = "DisbursementID"
        elif "disbursementid" in self.whatsapp_mtd_base_df.columns:
            disb_col_wa = "disbursementid"
        else:
            disb_col_wa = "DisbursementID"
            self.whatsapp_mtd_base_df = (
                self.whatsapp_mtd_base_df
                .join(
                    self.base_accounts_df.select(["LoanMstID", "DisbursementID"]),
                    on="LoanMstID",
                    how="left"
                )
            )

        ext_col2 = "ExtraColumn2" if "ExtraColumn2" in self.whatsapp_mtd_base_df.columns else "extracolumn2"
        ext_col2_val = pl.col(ext_col2).fill_null("")
        ext_col2_upper = ext_col2_val.str.to_uppercase()
        
        ext_col2_compact = (
            ext_col2_upper
            .str.replace_all(r"^[0-9]+[_ \t-]*", "")
            .str.replace_all(r"[^A-Z0-9]", "")
        )

        mapped_wa_disposition_expr = (
            pl.when(pl.col("has_paid") == 1).then(pl.lit("Already Paid"))
            .when(pl.col("has_denied") == 1).then(pl.lit("Refused to pay"))
            .when(pl.col("has_ptp") == 1).then(pl.lit("PTP - WhatsApp"))
            .when(ext_col2_upper.str.contains("ALREADY PAID")).then(pl.lit("Already Paid"))
            .when(ext_col2_upper.str.contains("DENIED")).then(pl.lit("Refused to pay"))
            .when(ext_col2_upper.str.contains("NO COMM")).then(pl.lit("No Response"))
            .when(
                ext_col2_upper.str.contains("TOKEN") &
                (
                    ext_col2_upper.str.contains("PTP") |
                    ext_col2_upper.str.contains("PROMISE") |
                    ext_col2_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP Whatsapp - Free Text"))
            .when(
                ext_col2_upper.str.contains("SETTLEMENT") &
                (
                    ext_col2_upper.str.contains("PTP") |
                    ext_col2_upper.str.contains("PROMISE") |
                    ext_col2_upper.str.contains("PENDING")
                )
            ).then(pl.lit("PTP Whatsapp - Free Text"))
            .when(
                ext_col2_upper.str.contains("PTP") |
                ext_col2_upper.str.contains("PENDING") |
                ext_col2_upper.str.contains("PROMISE") |
                ext_col2_upper.str.contains("DUE DATE PAYMENT INTENDED") |
                ext_col2_upper.str.contains("NO EARLY PAYMENT") |
                ext_col2_upper.str.contains("CONFIRM")
            ).then(pl.lit("PTP Whatsapp - Free Text"))
            .when(ext_col2_upper.str.contains("CONFUSION")).then(pl.lit("Confusion"))
            .when(ext_col2_upper.str.contains("PAYMENT DISPUTE")).then(pl.lit("Payment Dispute"))
            .when(ext_col2_upper.str.contains("INSURANCE")).then(pl.lit("Insurance Dispute"))
            .when(ext_col2_upper.str.contains("DISPUTE")).then(pl.lit("Dispute"))
            .when(ext_col2_upper.str.contains("WRONG CONTACT") | ext_col2_upper.str.contains("WRONG NUMBER")).then(pl.lit("Wrong Contact"))
            .when(ext_col2_upper.str.contains("UNAWARE") | ext_col2_upper.str.contains("COMMUNICATION ISSUE") | ext_col2_upper.str.contains("LANGUAGE")).then(pl.lit("Unaware / Communication Issue"))
            .when(ext_col2_upper.str.contains("FAMILY ISSUE") | ext_col2_upper.str.contains("HEALTH")).then(pl.lit("Health / Family Issue"))
            .when(ext_col2_upper.str.contains("NOISY ENV")).then(pl.lit("Noisy Env"))
            .when(ext_col2_upper.str.contains("INSUFFICIENT CALL DURATION") | ext_col2_upper.str.contains("CALL TOO SHORT")).then(pl.lit("Call too short for categorization"))
            .when(ext_col2_upper.str.contains("OTHER")).then(pl.lit("Other"))
            .when(ext_col2_upper.str.contains("HANG UP") | ext_col2_upper.str.contains("HANGUP") | ext_col2_upper.str.contains("DISCONNECTED")).then(pl.lit("Hang Up / No Response"))
            .when(ext_col2_compact.is_in(not_conn_list)).then(
                pl.when(ext_col2_compact == 'NETWORKCONGESTION').then(pl.lit('Network Congestion'))
                .when(ext_col2_compact == 'NODOESNOTEXIST').then(pl.lit('Number Does Not Exist'))
                .when(ext_col2_compact == 'ONLYRINGING').then(pl.lit('Only Ringing'))
                .when(ext_col2_compact == 'SWITCHEDOFF').then(pl.lit('Switched Off'))
                .when(ext_col2_compact == 'SPEAKINGTOSOMEONEELSE').then(pl.lit('Speaking To Someone Else'))
                .when(ext_col2_compact == 'USEROUTOFFCOVERAGE').then(pl.lit('User Out Of Coverage'))
                .when(ext_col2_compact == 'CALLREJECTED').then(pl.lit('Call Rejected'))
                .when(ext_col2_compact == 'OUTOFSERVICE').then(pl.lit('Out Of Service'))
                .when(ext_col2_compact == 'INCOMINGNOTAVAILABLE').then(pl.lit('Incoming Not Available'))
                .when(ext_col2_compact == 'BUSY').then(pl.lit('Busy'))
                .otherwise(pl.lit('Not Contactable'))
            )
            .when(ext_col2_upper.str.contains("FAILED") | ext_col2_upper.str.contains("CONGESTION")).then(pl.lit("No Answer"))
            .when(ext_col2_upper.str.contains("BUSY")).then(pl.lit("Busy"))
            .when(ext_col2_upper.str.contains("SWITCHED OFF") | ext_col2_upper.str.contains("OUT OF SERVICE") | ext_col2_upper.str.contains("NOT REACHABLE")).then(pl.lit("Not Reachable / Out of Network"))
            .when(pl.col(ext_col2).is_not_null() & (pl.col(ext_col2).str.strip_chars() != "")).then(pl.col(ext_col2))
            .when((pl.col("has_any_response") == 1) | pl.col("received_after_send")).then(pl.lit("Not Categorized"))
            .otherwise(pl.lit("No Answer"))
        )
        
        mapped_wa_disposition_cleaned = mapped_wa_disposition_expr.str.strip_chars().str.replace_all(r":+[ \t]*$", "")

        is_sent_col = "IsSent" if "IsSent" in self.whatsapp_mtd_base_df.columns else "issent"
        is_delivered_col = "IsDelivered" if "IsDelivered" in self.whatsapp_mtd_base_df.columns else "isdelivered"
        is_read_col = "IsRead" if "IsRead" in self.whatsapp_mtd_base_df.columns else "isread"
        created_date_col_wa = "CreatedDate" if "CreatedDate" in self.whatsapp_mtd_base_df.columns else "createddate"
        queue_col_wa = "WhatsAppQueueID" if "WhatsAppQueueID" in self.whatsapp_mtd_base_df.columns else "whatsappqueueid"
        disb_col_wa = "DisbursementID" if "DisbursementID" in self.whatsapp_mtd_base_df.columns else "disbursementid"

        ord_case_1 = (
            pl.when(
                (pl.col("has_paid") == 1) |
                (pl.col("has_denied") == 1) |
                (pl.col("has_ptp") == 1) |
                (pl.col("has_any_response") == 1)
            ).then(pl.lit(0))
            .when(
                pl.col(ext_col2).is_not_null() & (pl.col(ext_col2).str.strip_chars() != "")
            ).then(pl.lit(1))
            .when(
                (pl.col(is_read_col) == True) |
                pl.col("received_after_send") |
                (pl.col(is_delivered_col) == True)
            ).then(pl.lit(1))
            .otherwise(pl.lit(2))
        )

        ord_case_2 = (
            pl.when(
                (pl.col("has_paid") == 1) |
                (pl.col("has_denied") == 1) |
                (pl.col("has_ptp") == 1) |
                (pl.col("has_any_response") == 1)
            ).then(pl.col(created_date_col_wa))
            .otherwise(pl.lit(None).cast(pl.Datetime))
        )

        ord_case_3 = (
            pl.when(
                pl.col(ext_col2).is_not_null() & (pl.col(ext_col2).str.strip_chars() != "")
            ).then(pl.col(created_date_col_wa))
            .otherwise(pl.lit(None).cast(pl.Datetime))
        )

        ord_case_4 = (
            pl.when(
                (pl.col(is_read_col) == True) |
                pl.col("received_after_send") |
                (pl.col(is_delivered_col) == True)
            ).then(pl.col(created_date_col_wa))
            .otherwise(pl.lit(None).cast(pl.Datetime))
        )

        base_lf = (
            self.whatsapp_mtd_base_df
            .lazy()
            .join(
                self.whatsapp_response_map_df.lazy(),
                left_on="WhatsAppQueueID",
                right_on="WhatsappQueueID",
                how="left"
            )
            .with_columns([
                pl.col("has_paid").fill_null(0),
                pl.col("has_denied").fill_null(0),
                pl.col("has_ptp").fill_null(0),
                pl.col("has_any_response").fill_null(0)
            ])
            .with_columns([
                mapped_wa_disposition_cleaned.alias("mapped_wa_disposition"),
                ord_case_1.alias("ord_case_1"),
                ord_case_2.alias("ord_case_2"),
                ord_case_3.alias("ord_case_3"),
                ord_case_4.alias("ord_case_4")
            ])
            .sort(
                by=["LoanMstID", "ord_case_1", "ord_case_2", "ord_case_3", "ord_case_4", created_date_col_wa],
                descending=[False, False, True, True, True, True]
            )
            .unique(subset=["LoanMstID"], keep="first")
        )

        attempts_lf = (
            self.whatsapp_mtd_base_df
            .lazy()
            .join(
                self.whatsapp_response_map_df.lazy(),
                left_on="WhatsAppQueueID",
                right_on="WhatsappQueueID",
                how="left"
            )
            .with_columns(pl.col("has_any_response").fill_null(0))
            .group_by("LoanMstID")
            .agg([
                pl.len().alias("whatsapp_attempts"),
                (
                    (pl.col(is_delivered_col) == True) |
                    (pl.col(is_read_col) == True) |
                    pl.col("received_after_send") |
                    (pl.col("has_any_response") == 1)
                ).sum().alias("whatsapp_delivered_count"),
                (
                    (pl.col(is_read_col) == True) |
                    pl.col("received_after_send") |
                    (pl.col("has_any_response") == 1)
                ).sum().alias("whatsapp_read_count"),
                (
                    (pl.col(is_read_col) == True) |
                    pl.col("received_after_send") |
                    (pl.col("has_any_response") == 1)
                ).sum().alias("whatsapp_connected_count")
            ])
        )

        from . import business_rules
        disposition_ranking_map_df = pl.DataFrame({
            "disposition": [r[0] for r in business_rules.DISPOSITION_RANKING],
        })
        disposition_ranking_map_df = (
            disposition_ranking_map_df
            .with_columns(
                pl.col("disposition").str.strip_chars().str.to_uppercase().alias("disp_upper")
            )
            .unique(subset=["disp_upper"], keep="first")
        )

        self.latest_whatsapp_df = (
            base_lf
            .join(attempts_lf, on="LoanMstID", how="left")
            .with_columns(
                pl.col("mapped_wa_disposition").str.to_uppercase().alias("mapped_wa_disp_upper")
            )
            .join(
                disposition_ranking_map_df.lazy(),
                left_on="mapped_wa_disp_upper",
                right_on="disp_upper",
                how="left"
            )
            .select([
                pl.col("LoanMstID"),
                pl.col(disb_col_wa).alias("DisbursementID"),
                pl.col("wa_date").alias("Whatsapp Sent Date"),
                pl.col("WhatsAppQueueID"),
                pl.coalesce([pl.col("disposition"), pl.col("mapped_wa_disposition")]).alias("Extra_DP_Category"),
                pl.col("ExtraColumn3").alias("Extra_DP_Reason"),
                pl.col(is_sent_col).alias("IsSent"),
                pl.col(is_delivered_col).alias("IsDelivered"),
                pl.col(is_read_col).alias("IsRead"),
                pl.col("whatsapp_attempts").fill_null(0).alias("whatsapp_attempts"),
                pl.col("whatsapp_delivered_count").fill_null(0).alias("whatsapp_delivered_count"),
                pl.col("whatsapp_read_count").fill_null(0).alias("whatsapp_read_count"),
                pl.col("whatsapp_connected_count").fill_null(0).alias("whatsapp_connected_count")
            ])
            .collect()
        )
        self.logger.info("latest_whatsapp_df size: %d", len(self.latest_whatsapp_df))

        # -------------------------------------------------------------------------
        # STEP 7: whatsapp_combined_df
        # -------------------------------------------------------------------------
        assigned_to_col = "AssignedTo" if "AssignedTo" in self.whatsapp_messages_df.columns else "assignedto"
        msg_date_col = "MessageDate" if "MessageDate" in self.whatsapp_messages_df.columns else "messagedate"
        content_col = "Content" if "Content" in self.whatsapp_messages_df.columns else "content"
        
        self.whatsapp_combined_df = (
            self.whatsapp_messages_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .filter(
                (pl.col(assigned_to_col) != "System") |
                (pl.col(assigned_to_col).is_null())
            )
            .sort(by=["LoanMstID", msg_date_col], descending=[False, False])
            .group_by("LoanMstID")
            .agg(
                pl.col(content_col).fill_null("").str.concat(" | ").alias("all_messages")
            )
            .collect()
        )
        self.logger.info("whatsapp_combined_df size: %d", len(self.whatsapp_combined_df))

    @timer
    def build_latest_responses(self) -> Any:
        """
        Extracts and merges user responses across conversational channel logs.
        """
        logger.info("Extracting and building latest responses.")
        # TODO: Implement latest user message processing.
        raise NotImplementedError("build_latest_responses is not yet implemented.")

    @timer
    def build_dispositions(self) -> Any:
        """
        Extracts and consolidates system and agent call dispositions.
        """
        logger.info("Building dispositions.")
        # TODO: Apply disposition aggregation logic.
        raise NotImplementedError("build_dispositions is not yet implemented.")

    @timer
    def build_collections(self) -> None:
        """
        Builds collections/payments data associated with targets.
        """
        self.logger.info("Starting collections processing (build_collections)...")
        if self.transactions_df is None:
            raise ValueError("Transactions data is not loaded.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded.")

        # Log initial stats
        loaded_count = len(self.transactions_df)
        self.logger.info("Validation Log - Loaded Transactions rows: %d", loaded_count)

        # Filter: BankMstID == 53
        bank_mst_col = "BankMstID" if "BankMstID" in self.transactions_df.columns else "bankmstid"
        bank_filtered = self.transactions_df.filter(pl.col(bank_mst_col) == 53)
        self.logger.info("Validation Log - Rows after Bank filter: %d", len(bank_filtered))

        # Filter: Collection date within report month
        coll_date_col = "CollectedDate" if "CollectedDate" in bank_filtered.columns else "collecteddate"
        date_filtered = bank_filtered.filter(
            (pl.col(coll_date_col).cast(pl.Date) >= self.request.start_date) &
            (pl.col(coll_date_col).cast(pl.Date) <= self.request.end_date)
        )
        self.logger.info("Validation Log - Rows after Date filter: %d", len(date_filtered))
        
        # Filter: LoanMstID exists in base_accounts_df
        loan_mst_col = "LoanMstID" if "LoanMstID" in date_filtered.columns else "loanmstid"
        loan_filtered = (
            date_filtered
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                left_on=loan_mst_col,
                right_on="LoanMstID",
                how="inner"
            )
            .collect()
        )
        self.logger.info("Validation Log - Rows after Loan filter: %d", len(loan_filtered))

        # Identify DisbursementID and CollectedAmt columns
        disb_col = "DisbursementID" if "DisbursementID" in loan_filtered.columns else "disbursementid"
        coll_amt_col = "CollectedAmt" if "CollectedAmt" in loan_filtered.columns else "collectedamt"

        # -------------------------------------------------------------------------
        # STEP 1: latest_collection_df
        # -------------------------------------------------------------------------
        latest_lf = (
            loan_filtered
            .lazy()
            .sort(by=[disb_col, coll_date_col], descending=[False, True])
            .unique(subset=[disb_col], keep="first")
            .select([
                pl.col(disb_col).alias("DisbursementID"),
                pl.col(coll_date_col).cast(pl.Date).alias("Last Collection Date"),
                pl.col(coll_amt_col).alias("Last Collected Amount")
            ])
        )
        self.latest_collection_df = latest_lf.collect()
        self.logger.info("Validation Log - Rows after latest selection: %d", len(self.latest_collection_df))

        # -------------------------------------------------------------------------
        # STEP 2: total_collection_df
        # -------------------------------------------------------------------------
        total_lf = (
            loan_filtered
            .lazy()
            .group_by(disb_col)
            .agg(
                pl.col(coll_amt_col).sum().alias("total_collected_amt")
            )
            .select([
                pl.col(disb_col).alias("DisbursementID"),
                pl.col("total_collected_amt")
            ])
        )
        self.total_collection_df = total_lf.collect()
        self.logger.info("Validation Log - Rows after aggregation: %d", len(self.total_collection_df))

        # -------------------------------------------------------------------------
        # FINAL VALIDATION
        # -------------------------------------------------------------------------
        for name, df in [("latest_collection_df", self.latest_collection_df), ("total_collection_df", self.total_collection_df)]:
            self.logger.info("Validation Log - %s shape: %s", name, df.shape)
            self.logger.info("Validation Log - %s columns: %s", name, df.columns)
            self.logger.info("Validation Log - %s column order: %s", name, df.columns)
            self.logger.info("Validation Log - %s null counts: %s", name, df.null_count().to_dicts()[0])
            self.logger.info("Validation Log - %s unique DisbursementID: %d", name, df["DisbursementID"].n_unique())
            self.logger.info("Validation Log - %s duplicate DisbursementID: %d", name, len(df) - df["DisbursementID"].n_unique())
            self.logger.info("Validation Log - %s estimated memory: %d bytes (%.4f KB)", name, df.estimated_size(), df.estimated_size() / 1024.0)

    @timer
    def build_connection_statistics(self) -> None:
        """
        Calculates and maps call connection ratios and other dialer/channel metrics.
        """
        self.logger.info("Starting connection statistics processing (build_connection_statistics)...")
        
        # Verify required tables exist
        if self.voicebot_mtd_base_df is None:
            raise ValueError("voicebot_mtd_base_df is not available.")
        if self.blaster_mtd_base_df is None:
            raise ValueError("blaster_mtd_base_df is not available.")
        if self.whatsapp_mtd_base_df is None:
            raise ValueError("whatsapp_mtd_base_df is not available.")
        if self.whatsapp_response_map_df is None:
            raise ValueError("whatsapp_response_map_df is not available.")

        # Fallback column names
        created_date_col_v = "CreatedDate" if "CreatedDate" in self.voicebot_mtd_base_df.columns else "createddate"
        created_date_col_bl = "CreatedDate" if "CreatedDate" in self.blaster_mtd_base_df.columns else "createddate"
        duration_col_v = "CallDuration" if "CallDuration" in self.voicebot_mtd_base_df.columns else "callduration"
        duration_col_bl = "CallDuration" if "CallDuration" in self.blaster_mtd_base_df.columns else "callduration"

        # -------------------------------------------------------------------------
        # STEP 1: mtd_connection_flags_df
        # -------------------------------------------------------------------------
        call_attempts = (
            self.voicebot_mtd_base_df
            .lazy()
            .group_by("LoanMstID")
            .agg(
                pl.col("is_connected").any().cast(pl.Int32).alias("call_connected")
            )
        )
        self.logger.info("Validation Log - call_attempts size: %d", len(call_attempts.collect()))

        blaster_attempts = (
            self.blaster_mtd_base_df
            .lazy()
            .group_by("LoanMstID")
            .agg(
                pl.col("is_connected").any().cast(pl.Int32).alias("blaster_connected")
            )
        )
        self.logger.info("Validation Log - blaster_attempts size: %d", len(blaster_attempts.collect()))

        wa_conn_expr = (
            (pl.col("IsRead") == True) |
            (pl.col("received_after_send") == True) |
            (pl.col("has_any_response").fill_null(0) == 1)
        )
        
        wa_attempts = (
            self.whatsapp_mtd_base_df
            .lazy()
            .join(
                self.whatsapp_response_map_df.lazy(),
                left_on="WhatsAppQueueID",
                right_on="WhatsappQueueID",
                how="left"
            )
            .with_columns(
                pl.col("has_any_response").fill_null(0)
            )
            .group_by("LoanMstID")
            .agg(
                wa_conn_expr.any().cast(pl.Int32).alias("wa_connected")
            )
        )
        self.logger.info("Validation Log - wa_attempts size: %d", len(wa_attempts.collect()))

        all_loan_ids = (
            pl.concat([
                call_attempts.select("LoanMstID"),
                blaster_attempts.select("LoanMstID"),
                wa_attempts.select("LoanMstID")
            ])
            .unique()
        )
        self.logger.info("Validation Log - all_loan_ids size: %d", len(all_loan_ids.collect()))

        mtd_connection_flags_df = (
            all_loan_ids
            .join(call_attempts, on="LoanMstID", how="left")
            .join(blaster_attempts, on="LoanMstID", how="left")
            .join(wa_attempts, on="LoanMstID", how="left")
            .with_columns([
                pl.col("call_connected").fill_null(0),
                pl.col("blaster_connected").fill_null(0),
                pl.col("wa_connected").fill_null(0),
            ])
            .select([
                pl.col("LoanMstID"),
                pl.lit(1).alias("unique_total_attempt"),
                pl.when(
                    (pl.col("call_connected") == 1) |
                    (pl.col("blaster_connected") == 1) |
                    (pl.col("wa_connected") == 1)
                ).then(pl.lit(1)).otherwise(pl.lit(0)).alias("unique_total_connect"),
                pl.col("call_connected").alias("ai_connected_mtd"),
                pl.col("blaster_connected").alias("blaster_connected_mtd")
            ])
        )
        self.mtd_connection_flags_df = mtd_connection_flags_df.collect()
        self.logger.info("Validation Log - mtd_connection_flags_df size: %d", len(self.mtd_connection_flags_df))

        # -------------------------------------------------------------------------
        # STEP 2: mtd_wa_connection_flags_df
        # -------------------------------------------------------------------------
        self.mtd_wa_connection_flags_df = (
            self.whatsapp_mtd_base_df
            .lazy()
            .join(
                self.whatsapp_response_map_df.lazy(),
                left_on="WhatsAppQueueID",
                right_on="WhatsappQueueID",
                how="left"
            )
            .with_columns(
                pl.col("has_any_response").fill_null(0)
            )
            .group_by("LoanMstID")
            .agg(
                wa_conn_expr.any().cast(pl.Int32).alias("wa_connected_mtd")
            )
            .collect()
        )
        self.logger.info("Validation Log - mtd_wa_connection_flags_df size: %d", len(self.mtd_wa_connection_flags_df))

        # -------------------------------------------------------------------------
        # STEP 3: last_connected_info_df
        # -------------------------------------------------------------------------
        call_connections = (
            self.voicebot_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .select([
                pl.col("LoanMstID"),
                pl.col("call_date").cast(pl.Date).alias("connection_date"),
                pl.lit("VoiceBot").alias("channel")
            ])
        )

        wa_conn_1 = (
            self.whatsapp_mtd_base_df
            .lazy()
            .filter(
                (pl.col("IsRead") == True) |
                (pl.col("received_after_send") == True)
            )
            .select([
                pl.col("LoanMstID"),
                pl.col("wa_date").cast(pl.Date).alias("connection_date"),
                pl.lit("WhatsApp").alias("channel")
            ])
        )

        wa_conn_2 = (
            self.whatsapp_mtd_base_df
            .lazy()
            .join(
                self.whatsapp_response_map_df.lazy(),
                left_on="WhatsAppQueueID",
                right_on="WhatsappQueueID" if "WhatsAppQueueID" in self.whatsapp_response_map_df.columns else "WhatsappQueueID",
                how="inner"
            )
            .filter(
                (pl.col("has_any_response").fill_null(0) == 1) &
                pl.col("latest_response_datetime").is_not_null()
            )
            .select([
                pl.col("LoanMstID"),
                pl.col("latest_response_datetime").cast(pl.Date).alias("connection_date"),
                pl.lit("WhatsApp").alias("channel")
            ])
        )

        wa_connections = pl.concat([wa_conn_1, wa_conn_2])

        blaster_connections = (
            self.blaster_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .select([
                pl.col("LoanMstID"),
                pl.col("blaster_date").cast(pl.Date).alias("connection_date"),
                pl.lit("Blaster").alias("channel")
            ])
        )

        all_connections = pl.concat([
            call_connections,
            wa_connections,
            blaster_connections
        ])
        self.logger.info("Validation Log - all_connections size: %d", len(all_connections.collect()))

        ranked_connections = (
            all_connections
            .sort(by=["LoanMstID", "connection_date", "LoanMstID"], descending=[False, True, False])
            .with_columns(
                pl.int_range(1, pl.len() + 1).over("LoanMstID").alias("rn")
            )
            .filter(pl.col("rn") == 1)
        )
        self.logger.info("Validation Log - ranked_connections size: %d", len(ranked_connections.collect()))

        self.last_connected_info_df = (
            ranked_connections
            .select([
                pl.col("LoanMstID"),
                pl.col("connection_date").alias("last_connected_date"),
                pl.col("channel").alias("last_connected_channel")
            ])
            .collect()
        )
        self.logger.info("Validation Log - last_connected_info_df size: %d", len(self.last_connected_info_df))

        # -------------------------------------------------------------------------
        # STEP 4: last_connected_call_duration_df
        # -------------------------------------------------------------------------
        conn_v_dur = (
            self.voicebot_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .select([
                pl.col("LoanMstID"),
                pl.col(created_date_col_v).alias("CreatedDate"),
                pl.col(duration_col_v).alias("CallDuration")
            ])
        )

        conn_bl_dur = (
            self.blaster_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .select([
                pl.col("LoanMstID"),
                pl.col(created_date_col_bl).alias("CreatedDate"),
                pl.col(duration_col_bl).alias("CallDuration")
            ])
        )

        connected_call_durations = pl.concat([conn_v_dur, conn_bl_dur])

        self.last_connected_call_duration_df = (
            connected_call_durations
            .sort(by=["LoanMstID", "CreatedDate"], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("CallDuration").alias("last_connected_call_duration")
            ])
            .collect()
        )
        self.logger.info("Validation Log - last_connected_call_duration_df size: %d", len(self.last_connected_call_duration_df))

        # -------------------------------------------------------------------------
        # STEP 5: call_duration_stats_df
        # -------------------------------------------------------------------------
        ccd_v = (
            self.voicebot_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .select([
                pl.col("LoanMstID"),
                pl.col(duration_col_v).alias("CallDuration")
            ])
        )
        
        ccd_bl = (
            self.blaster_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .select([
                pl.col("LoanMstID"),
                pl.col(duration_col_bl).alias("CallDuration")
            ])
        )
        
        ccd = pl.concat([ccd_v, ccd_bl])
        
        self.call_duration_stats_df = (
            ccd
            .group_by("LoanMstID")
            .agg(
                pl.col("CallDuration").fill_null(0).max().alias("max_connected_call_duration")
            )
            .collect()
        )
        self.logger.info("Validation Log - call_duration_stats_df size: %d", len(self.call_duration_stats_df))

        # -------------------------------------------------------------------------
        # STEP 6: voicebot_call_duration_stats_df
        # -------------------------------------------------------------------------
        self.voicebot_call_duration_stats_df = (
            self.voicebot_mtd_base_df
            .lazy()
            .filter(pl.col("is_connected") == True)
            .group_by("LoanMstID")
            .agg(
                pl.col(duration_col_v).fill_null(0).max().alias("voicebot_max_connected_call_duration")
            )
            .collect()
        )
        self.logger.info("Validation Log - voicebot_call_duration_stats_df size: %d", len(self.voicebot_call_duration_stats_df))

        # -------------------------------------------------------------------------
        # FINAL VALIDATION FOR EACH
        # -------------------------------------------------------------------------
        dfs = [
            ("mtd_connection_flags_df", self.mtd_connection_flags_df),
            ("mtd_wa_connection_flags_df", self.mtd_wa_connection_flags_df),
            ("last_connected_info_df", self.last_connected_info_df),
            ("last_connected_call_duration_df", self.last_connected_call_duration_df),
            ("call_duration_stats_df", self.call_duration_stats_df),
            ("voicebot_call_duration_stats_df", self.voicebot_call_duration_stats_df),
        ]
        for name, df in dfs:
            self.logger.info("Validation Log - %s shape: %s", name, df.shape)
            self.logger.info("Validation Log - %s columns: %s", name, df.columns)
            self.logger.info("Validation Log - %s column order: %s", name, df.columns)
            self.logger.info("Validation Log - %s null counts: %s", name, df.null_count().to_dicts()[0])
            self.logger.info("Validation Log - %s unique LoanMstID: %d", name, df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s duplicate LoanMstID: %d", name, len(df) - df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s estimated memory: %d bytes (%.4f KB)", name, df.estimated_size(), df.estimated_size() / 1024.0)

    @timer
    def build_historical_dispositions(self) -> None:
        """
        Builds Phase 14 Historical Disposition tables:
        - last_month_best_disposition_df
        - latest_disposition_mtd_df
        - total_ptp_responses_df
        """
        self.logger.info("Starting Historical Dispositions processing (build_historical_dispositions)...")
        if self.voicebot_lm_df is None or self.whatsapp_lm_df is None:
            raise ValueError("Last month VoiceBot or WhatsApp raw data is not loaded.")
        if self.base_accounts_df is None:
            raise ValueError("Base accounts data is not loaded.")
        if self.response_df is None:
            raise ValueError("Response data is not loaded.")
        if self.all_call_dispositions_df is None or self.all_wa_dispositions_df is None or self.all_blaster_dispositions_df is None:
            raise ValueError("MTD call/wa/blaster dispositions data are not loaded.")

        # -------------------------------------------------------------------------
        # STEP 1: last_month_best_disposition_df
        # -------------------------------------------------------------------------
        base_loans = self.base_accounts_df.select(["LoanMstID"]).unique()

        # Inner join to LoanMstIDs in base_accounts
        voicebot_lm_filtered = self.voicebot_lm_df.join(base_loans, on="LoanMstID", how="inner")
        whatsapp_lm_filtered = self.whatsapp_lm_df.join(base_loans, on="LoanMstID", how="inner")

        self.logger.info("Validation Log - voicebot_lm_filtered count: %d", len(voicebot_lm_filtered))
        self.logger.info("Validation Log - whatsapp_lm_filtered count: %d", len(whatsapp_lm_filtered))

        from dateutil.relativedelta import relativedelta
        from datetime import timedelta
        lm_start = (self.request.start_date - relativedelta(months=1)).replace(day=1)
        this_m_start = self.request.start_date

        # Column fallbacks for VoiceBot History
        v_cols = voicebot_lm_filtered.columns
        created_date_col_v = "CreatedDate" if "CreatedDate" in v_cols else "createddate"
        duration_col_v = "CallDuration" if "CallDuration" in v_cols else "callduration"
        recording_col_v = "Recording" if "Recording" in v_cols else "recording"
        ext_col1 = "Extracolumn1" if "Extracolumn1" in v_cols else "extracolumn1"
        dispo_col = "deposition" if "deposition" in v_cols else "Disposition" if "Disposition" in v_cols else "deposition"

        raw_disposition = pl.coalesce([
            pl.col(ext_col1).str.strip_chars(),
            pl.col(dispo_col).str.strip_chars()
        ]).fill_null("")

        raw_disposition_trimmed = raw_disposition.str.strip_chars()
        raw_disposition_upper = raw_disposition_trimmed.str.to_uppercase()
        raw_disposition_lower = raw_disposition_trimmed.str.to_lowercase()

        call_duration_expr = (
            pl.when(pl.col(duration_col_v) == 0)
            .then(pl.lit(1))
            .otherwise(pl.col(duration_col_v).fill_null(0))
        )

        voicebot_lm_base_lf = (
            voicebot_lm_filtered
            .lazy()
            .with_columns([
                call_duration_expr.alias("call_duration"),
                pl.col(recording_col_v).alias("Recording"),
                raw_disposition.alias("raw_disposition"),
                raw_disposition_trimmed.alias("raw_disposition_trimmed"),
                raw_disposition_upper.alias("raw_disposition_upper"),
                raw_disposition_lower.alias("raw_disposition_lower"),
            ])
            .filter(
                (pl.col(created_date_col_v).cast(pl.Date) >= lm_start) &
                (pl.col(created_date_col_v).cast(pl.Date) < this_m_start)
            )
        )

        # Build voicebot_lm_response from response_df
        vb_queue_col = "VoiceBotQueueID" if "VoiceBotQueueID" in self.response_df.columns else "voicebotqueueid"
        vb_queue_ids_lf = (
            voicebot_lm_base_lf
            .select(["VoiceBotQueueID"])
            .filter(pl.col("VoiceBotQueueID").is_not_null())
            .unique()
        )

        voicebot_lm_response_lf = (
            self.response_df
            .lazy()
            .filter(
                (pl.col("BankMstID") == 53) &
                (pl.col(vb_queue_col).is_not_null())
            )
            .group_by(vb_queue_col)
            .agg([
                pl.col("Status").is_in(["Already Paid", "Claim_Expired"]).any().alias("has_already_paid"),
                (pl.col("Status") == "Denied").any().alias("has_denied"),
                pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"]).any().alias("has_ptp")
            ])
            .join(vb_queue_ids_lf, left_on=vb_queue_col, right_on="VoiceBotQueueID", how="inner")
        )

        # Build whatsapp_lm_base
        wa_cols = whatsapp_lm_filtered.columns
        created_date_col_wa = "CreatedDate" if "CreatedDate" in wa_cols else "createddate"
        whatsapp_lm_base_lf = (
            whatsapp_lm_filtered
            .lazy()
            .filter(
                (pl.col(created_date_col_wa).cast(pl.Date) >= lm_start) &
                (pl.col(created_date_col_wa).cast(pl.Date) < this_m_start)
            )
        )

        # Build whatsapp_lm_response
        wa_queue_col = "WhatsappQueueID" if "WhatsappQueueID" in self.response_df.columns else "whatsappqueueid"
        wa_queue_col_hist = "WhatsAppQueueID" if "WhatsAppQueueID" in wa_cols else "whatsappqueueid"

        wa_queue_ids_lf = (
            whatsapp_lm_base_lf
            .select([wa_queue_col_hist])
            .filter(pl.col(wa_queue_col_hist).is_not_null())
            .unique()
        )

        whatsapp_lm_response_lf = (
            self.response_df
            .lazy()
            .filter(
                (pl.col("BankMstID") == 53) &
                (pl.col(wa_queue_col).is_not_null())
            )
            .group_by(wa_queue_col)
            .agg([
                pl.col("Status").is_in(["Already Paid", "Claim_Expired"]).any().alias("has_already_paid"),
                (pl.col("Status") == "Denied").any().alias("has_denied"),
                pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"]).any().alias("has_ptp")
            ])
            .join(wa_queue_ids_lf, left_on=wa_queue_col, right_on=wa_queue_col_hist, how="inner")
        )

        # Union attempts
        vbr_has_already_paid = pl.col("has_already_paid").fill_null(False)
        vbr_has_denied = pl.col("has_denied").fill_null(False)
        vbr_has_ptp = pl.col("has_ptp").fill_null(False)

        vb_disposition_expr = (
            pl.when((pl.col("call_duration") > 15) & vbr_has_already_paid).then(pl.lit("Already Paid"))
            .when((pl.col("call_duration") > 15) & vbr_has_denied).then(pl.lit("Refused to pay"))
            .when((pl.col("call_duration") > 15) & vbr_has_ptp).then(pl.lit("PTP"))
            .when(pl.col("Recording").is_not_null() & (pl.col("Recording").str.len_bytes() > 1) & (pl.col("raw_disposition_trimmed") == "")).then(pl.lit("No Response"))
            .when(pl.col("raw_disposition_upper").str.contains("ALREADY PAID") | (pl.col("raw_disposition_lower") == "already_paid")).then(pl.lit("Already Paid"))
            .when(pl.col("raw_disposition_upper").str.contains("DENIED") | pl.col("raw_disposition_upper").str.contains("DENIES") | pl.col("raw_disposition_upper").str.contains("DENIAL") | pl.col("raw_disposition_upper").str.contains("REFUSED")).then(pl.lit("Refused to pay"))
            .when(pl.col("raw_disposition_lower") == "not_ready_to_pay").then(pl.lit("Refused to pay"))
            .when(pl.col("raw_disposition_upper").str.contains("CONFUSION")).then(pl.lit("Confusion"))
            .when(pl.col("raw_disposition_upper").str.contains("PAYMENT DISPUTE")).then(pl.lit("Payment Dispute"))
            .when(pl.col("raw_disposition_upper").str.contains("INSURANCE")).then(pl.lit("Insurance Dispute"))
            .when(pl.col("raw_disposition_upper").str.contains("DISPUTE") | (pl.col("raw_disposition_lower") == "dispute_raised")).then(pl.lit("Dispute"))
            .when(pl.col("raw_disposition_lower") == "follow_up_required").then(pl.lit("Follow-up Required"))
            .when(pl.col("raw_disposition_lower").is_in(["callback", "call back"])).then(pl.lit("Call Back"))
            .when(pl.col("raw_disposition_lower") == "inquiry").then(pl.lit("Inquiry"))
            .when(pl.col("raw_disposition_lower") == "payment inquiry").then(pl.lit("Payment Inquiry"))
            .when(pl.col("raw_disposition_lower") == "overdue inquiry").then(pl.lit("Overdue Inquiry"))
            .when(pl.col("raw_disposition_lower") == "payment difficulty").then(pl.lit("Payment Difficulty"))
            .when(pl.col("raw_disposition_lower") == "payment arrangement").then(pl.lit("Payment Arrangement"))
            .when(pl.col("raw_disposition_lower") == "greeting/confirmation").then(pl.lit("Greeting/Confirmation"))
            .when(pl.col("raw_disposition_lower") == "family/third party").then(pl.lit("Family/Third Party"))
            .when(pl.col("raw_disposition_lower") == "contact verification").then(pl.lit("Contact Verification"))
            .when(pl.col("raw_disposition_lower") == "no request").then(pl.lit("No Request"))
            .when(pl.col("raw_disposition_lower") == "unclear intent").then(pl.lit("Unclear Intent"))
            .when(pl.col("raw_disposition_lower") == "service request").then(pl.lit("Service Request"))
            .when(pl.col("raw_disposition_lower") == "complaint/escalation").then(pl.lit("Complaint/Escalation"))
            .when(pl.col("raw_disposition_lower") == "account issue").then(pl.lit("Account Issue"))
            .when(
                pl.col("raw_disposition_upper").str.contains("TOKEN") &
                (
                    pl.col("raw_disposition_upper").str.contains("PTP") |
                    pl.col("raw_disposition_upper").str.contains("PROMISE") |
                    pl.col("raw_disposition_upper").str.contains("PENDING")
                )
            ).then(pl.lit("PTP - Token Amount"))
            .when(
                pl.col("raw_disposition_upper").str.contains("SETTLEMENT") &
                (
                    pl.col("raw_disposition_upper").str.contains("PTP") |
                    pl.col("raw_disposition_upper").str.contains("PROMISE") |
                    pl.col("raw_disposition_upper").str.contains("PENDING") |
                    pl.col("raw_disposition_lower").is_in(["futurepay", "ready_to_pay", "willing_to_pay"])
                )
            ).then(pl.lit("PTP (Promise to Pay) - Settlement"))
            .when(pl.col("raw_disposition_upper").str.contains("SETTLEMENT") & (pl.col("call_duration") < 60)).then(pl.lit("Other"))
            .when(pl.col("raw_disposition_lower").is_in(["futurepay", "ready_to_pay", "willing_to_pay"]) | pl.col("raw_disposition_upper").str.contains("WILL PAY AFTER SOMETIME")).then(pl.lit("PTP"))
            .when(
                pl.col("raw_disposition_upper").str.contains("PTP") |
                pl.col("raw_disposition_upper").str.contains("PENDING") |
                pl.col("raw_disposition_upper").str.contains("PROMISE") |
                pl.col("raw_disposition_upper").str.contains("DUE DATE PAYMENT INTENDED") |
                pl.col("raw_disposition_upper").str.contains("NO EARLY PAYMENT") |
                pl.col("raw_disposition_upper").str.contains("CONFIRM") |
                (pl.col("raw_disposition_lower") == "positive")
            ).then(pl.lit("PTP"))
            .when(pl.col("raw_disposition_lower").is_in(["wrongnumber", "wrong number"])).then(pl.lit("Wrong Contact"))
            .when(pl.col("raw_disposition_upper").str.contains("WRONG")).then(pl.lit("Wrong Contact"))
            .when(pl.col("raw_disposition_upper").str.contains("UNAWARE") | pl.col("raw_disposition_upper").str.contains("COMMUNICATION ISSUE") | pl.col("raw_disposition_upper").str.contains("LANGUAGE")).then(pl.lit("Unaware / Communication Issue"))
            .when(pl.col("raw_disposition_upper").str.contains("FAMILY ISSUE") | pl.col("raw_disposition_upper").str.contains("HEALTH")).then(pl.lit("Health / Family Issue"))
            .when(pl.col("raw_disposition_lower").is_in(["collectionofficer missing", "collection officer missing"])).then(pl.lit("Collection Officer Missing"))
            .when(pl.col("raw_disposition_upper").str.contains("OTHER")).then(pl.lit("Other"))
            .when(pl.col("raw_disposition_upper").str.contains("NOISY ENV")).then(pl.lit("Noisy Env"))
            .when(pl.col("raw_disposition_upper").str.contains("INSUFFICIENT CALL DURATION") | pl.col("raw_disposition_upper").str.contains("CALL TOO SHORT")).then(pl.lit("Call too short for categorization"))
            .when(pl.col("raw_disposition_lower") == "neutral").then(pl.lit("Neutral"))
            .otherwise(
                pl.when(pl.col("raw_disposition_trimmed") != "").then(pl.col("raw_disposition_trimmed")).otherwise(pl.lit("No Communication"))
            )
        )

        vb_attempts_lf = (
            voicebot_lm_base_lf
            .join(voicebot_lm_response_lf, on="VoiceBotQueueID", how="left")
            .select([
                pl.col("LoanMstID"),
                vb_disposition_expr.alias("disposition")
            ])
        )

        wbr_has_already_paid = pl.col("has_already_paid").fill_null(False)
        wbr_has_denied = pl.col("has_denied").fill_null(False)
        wbr_has_ptp = pl.col("has_ptp").fill_null(False)

        wa_disposition_expr = (
            pl.when(wbr_has_already_paid).then(pl.lit("Already Paid"))
            .when(wbr_has_denied).then(pl.lit("Refused to pay"))
            .when(wbr_has_ptp).then(pl.lit("PTP - WhatsApp"))
            .otherwise(pl.lit("No Answer"))
        )

        wa_attempts_lf = (
            whatsapp_lm_base_lf
            .join(whatsapp_lm_response_lf, left_on=wa_queue_col_hist, right_on=wa_queue_col, how="left")
            .select([
                pl.col("LoanMstID"),
                wa_disposition_expr.alias("disposition")
            ])
        )

        all_attempts = pl.concat([vb_attempts_lf, wa_attempts_lf])

        attempts_cleaned = (
            all_attempts
            .with_columns(
                pl.col("disposition").str.strip_chars().str.replace_all(r":+[ \t]*$", "").alias("disposition_cleaned")
            )
        )

        from . import business_rules

        alias_df = pl.DataFrame({
            "raw_disposition": [a[0] for a in business_rules.DISPOSITION_ALIASES],
            "normalized_disposition": [a[1] for a in business_rules.DISPOSITION_ALIASES],
            "ranking_disposition": [a[2] for a in business_rules.DISPOSITION_ALIASES],
        })
        alias_df = alias_df.with_columns(
            pl.col("raw_disposition").str.strip_chars().str.to_uppercase()
        ).unique(subset=["raw_disposition"], keep="first")

        ranking_df = pl.DataFrame({
            "disposition": [r[0] for r in business_rules.DISPOSITION_RANKING],
            "rank_val": [r[1] for r in business_rules.DISPOSITION_RANKING],
            "category": [r[2] for r in business_rules.DISPOSITION_RANKING],
        }, schema={"disposition": pl.String, "rank_val": pl.Float64, "category": pl.String})
        ranking_df = ranking_df.with_columns(
            pl.col("disposition").str.strip_chars().str.to_uppercase().alias("ranking_disp_upper")
        ).sort(by=["ranking_disp_upper", "rank_val", "category"]).unique(subset=["ranking_disp_upper"], keep="first")

        ranked = (
            attempts_cleaned
            .with_columns(
                pl.col("disposition_cleaned").str.strip_chars().str.to_uppercase().alias("cleaned_upper")
            )
            .join(alias_df.lazy(), left_on="cleaned_upper", right_on="raw_disposition", how="left")
            .with_columns(
                pl.when(pl.col("disposition_cleaned") == "PTP - Date")
                .then(pl.lit("PTP Date"))
                .otherwise(pl.coalesce([pl.col("ranking_disposition"), pl.col("disposition_cleaned")]))
                .str.strip_chars()
                .str.to_uppercase()
                .alias("ranking_key_upper")
            )
            .join(ranking_df.lazy(), left_on="ranking_key_upper", right_on="ranking_disp_upper", how="left")
            .with_columns([
                pl.coalesce([
                    pl.when(pl.coalesce([pl.col("ranking_disposition"), pl.col("disposition_cleaned")]) == "Denied")
                    .then(pl.lit("Refused to Pay"))
                    .otherwise(pl.col("normalized_disposition")),
                    pl.col("disposition"),
                    pl.lit("Other")
                ]).alias("final_disposition"),
                pl.col("rank_val").fill_null(999).alias("rank_val_filled"),
                pl.col("category").fill_null("Not Categorized").alias("category_filled")
            ])
            .sort(by=["LoanMstID", "rank_val_filled"], descending=[False, False])
            .unique(subset=["LoanMstID"], keep="first")
        )

        self.last_month_best_disposition_df = (
            ranked
            .select([
                pl.col("LoanMstID"),
                pl.col("final_disposition").alias("last_month_best_disposition"),
                pl.col("category_filled").alias("last_month_best_category")
            ])
            .collect()
        )

        # -------------------------------------------------------------------------
        # STEP 2: latest_disposition_mtd_df
        # -------------------------------------------------------------------------
        vb_comb = self.all_call_dispositions_df.select([
            pl.col("LoanMstID"),
            pl.col("disposition"),
            pl.col("category"),
            pl.col("call_date").alias("disposition_date"),
            pl.lit("VoiceBot").alias("source")
        ])

        wa_comb = self.all_wa_dispositions_df.select([
            pl.col("LoanMstID"),
            pl.col("disposition"),
            pl.col("category"),
            pl.col("wa_date").alias("disposition_date"),
            pl.lit("WhatsApp").alias("source")
        ])

        bl_comb = self.all_blaster_dispositions_df.select([
            pl.col("LoanMstID"),
            pl.col("disposition"),
            pl.col("category"),
            pl.col("blaster_date").alias("disposition_date"),
            pl.lit("Blaster").alias("source")
        ])

        combined = pl.concat([vb_comb, wa_comb, bl_comb])

        latest_disposition_mtd_df = (
            combined
            .lazy()
            .filter(
                pl.col("disposition").is_not_null() &
                (pl.col("disposition").str.strip_chars() != "")
            )
            .sort(by=["LoanMstID", "disposition_date", "LoanMstID"], descending=[False, True, False])
            .with_columns(
                pl.int_range(1, pl.len() + 1).over("LoanMstID").alias("rn")
            )
            .filter(pl.col("rn") == 1)
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("latest_disposition"),
                pl.col("category").alias("latest_disposition_category"),
                pl.col("source").alias("latest_disposition_channel"),
                pl.col("disposition_date").alias("latest_disposition_date")
            ])
        )
        self.latest_disposition_mtd_df = latest_disposition_mtd_df.collect()

        # -------------------------------------------------------------------------
        # STEP 3: total_ptp_responses_df
        # -------------------------------------------------------------------------
        vb_max_dur = (
            self.voicebot_mtd_base_df
            .lazy()
            .group_by("VoiceBotQueueID")
            .agg(
                pl.col("CallDuration").fill_null(0).max().alias("max_call_duration")
            )
        )

        response_mtd_base_lf = (
            self.response_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .join(
                vb_max_dur,
                left_on=vb_queue_col,
                right_on="VoiceBotQueueID",
                how="left"
            )
            .filter(
                (pl.col("BankMstID") == 53) &
                (~pl.col("Status").is_in(["", "Terminated"])) &
                (pl.col("ResponseDateTime").cast(pl.Date) >= self.request.start_date) &
                (pl.col("ResponseDateTime").cast(pl.Date) <= self.request.end_date) &
                (
                    pl.col(vb_queue_col).is_null() |
                    pl.col("max_call_duration").is_null() |
                    (pl.col("max_call_duration") > 15)
                )
            )
        )

        total_ptp_responses_df = (
            response_mtd_base_lf
            .filter(
                pl.col("Status").is_in(["Pending", "Fulfilled", "Broken", "Partially Paid"])
            )
            .group_by("LoanMstID")
            .agg([
                (pl.col(vb_queue_col).is_not_null()).cast(pl.Int32).sum().alias("ai_ptp_count"),
                (pl.col(wa_queue_col).is_not_null()).cast(pl.Int32).sum().alias("wa_ptp_count")
            ])
            .select([
                pl.col("LoanMstID"),
                pl.col("ai_ptp_count"),
                pl.col("wa_ptp_count")
            ])
        )
        self.total_ptp_responses_df = total_ptp_responses_df.collect()

        # Validation logging
        dfs = [
            ("last_month_best_disposition_df", self.last_month_best_disposition_df),
            ("latest_disposition_mtd_df", self.latest_disposition_mtd_df),
            ("total_ptp_responses_df", self.total_ptp_responses_df),
        ]
        for name, df in dfs:
            self.logger.info("Validation Log - %s shape: %s", name, df.shape)
            self.logger.info("Validation Log - %s columns: %s", name, df.columns)
            self.logger.info("Validation Log - %s column order: %s", name, df.columns)
            self.logger.info("Validation Log - %s null counts: %s", name, df.null_count().to_dicts()[0])
            self.logger.info("Validation Log - %s unique LoanMstID: %d", name, df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s duplicate LoanMstID: %d", name, len(df) - df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s estimated memory: %d bytes (%.4f KB)", name, df.estimated_size(), df.estimated_size() / 1024.0)

    @timer
    def build_latest_response_tables(self) -> None:
        """
        Builds Phase 15 intermediate tables:
        - latest_response_df
        - latest_voicebot_response_df
        - latest_whatsapp_response_df
        - latest_voicebot_channel_disposition_df
        - latest_whatsapp_channel_disposition_df
        - latest_blaster_channel_disposition_df
        """
        self.logger.info("Starting build_latest_response_tables...")
        
        # Verify required tables exist
        if self.response_df is None:
            raise ValueError("response_df is not available. Ensure load_source_tables has executed.")
        if self.voicebot_mtd_base_df is None:
            raise ValueError("voicebot_mtd_base_df is not available. Ensure build_all_call_dispositions has executed.")
        if self.base_accounts_df is None:
            raise ValueError("base_accounts_df is not available.")
        if self.all_call_dispositions_df is None:
            raise ValueError("all_call_dispositions_df is not available.")
        if self.all_wa_dispositions_df is None:
            raise ValueError("all_wa_dispositions_df is not available.")
        if self.all_blaster_dispositions_df is None:
            raise ValueError("all_blaster_dispositions_df is not available.")

        # Rebuild response_mtd_base_lf exactly as in Phase 14
        vb_max_dur = (
            self.voicebot_mtd_base_df
            .lazy()
            .group_by("VoiceBotQueueID")
            .agg(
                pl.col("CallDuration").fill_null(0).max().alias("max_call_duration")
            )
        )
        
        vb_queue_col = "VoiceBotQueueID" if "VoiceBotQueueID" in self.response_df.columns else "voicebotqueueid"
        wa_queue_col = "WhatsappQueueID" if "WhatsappQueueID" in self.response_df.columns else "whatsappqueueid"

        response_mtd_base_lf = (
            self.response_df
            .lazy()
            .join(
                self.base_accounts_df.select(["LoanMstID"]).lazy(),
                on="LoanMstID",
                how="inner"
            )
            .join(
                vb_max_dur,
                left_on=vb_queue_col,
                right_on="VoiceBotQueueID",
                how="left"
            )
            .filter(
                (pl.col("BankMstID") == 53) &
                (~pl.col("Status").is_in(["", "Terminated"])) &
                (pl.col("ResponseDateTime").cast(pl.Date) >= self.request.start_date) &
                (pl.col("ResponseDateTime").cast(pl.Date) <= self.request.end_date) &
                (
                    pl.col(vb_queue_col).is_null() |
                    pl.col("max_call_duration").is_null() |
                    (pl.col("max_call_duration") > 15)
                )
            )
        )

        # 1. latest_response_df
        latest_response_lf = (
            response_mtd_base_lf
            .sort(by=["LoanMstID", "ResponseDateTime"], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col(vb_queue_col).alias("VoiceBotQueueID"),
                pl.col(wa_queue_col).alias("WhatsappQueueID"),
                pl.col("Status"),
                pl.col("PromiseDateTime").alias("P2P_date"),
                pl.col("ResponseDateTime")
            ])
        )
        self.latest_response_df = latest_response_lf.collect()

        # 2. latest_voicebot_response_df
        latest_voicebot_response_lf = (
            response_mtd_base_lf
            .filter(pl.col(vb_queue_col).is_not_null())
            .sort(by=["LoanMstID", "ResponseDateTime"], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col(vb_queue_col).alias("VoiceBotQueueID"),
                pl.col("Status"),
                pl.col("PromiseDateTime").alias("P2P_date"),
                pl.col("ResponseDateTime")
            ])
        )
        self.latest_voicebot_response_df = latest_voicebot_response_lf.collect()

        # 3. latest_whatsapp_response_df
        latest_whatsapp_response_lf = (
            response_mtd_base_lf
            .filter(pl.col(wa_queue_col).is_not_null())
            .sort(by=["LoanMstID", "ResponseDateTime"], descending=[False, True])
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col(wa_queue_col).alias("WhatsappQueueID"),
                pl.col("Status"),
                pl.col("PromiseDateTime").alias("P2P_date"),
                pl.col("ResponseDateTime")
            ])
        )
        self.latest_whatsapp_response_df = latest_whatsapp_response_lf.collect()

        # 4. latest_voicebot_channel_disposition_df
        latest_voicebot_channel_disposition_lf = (
            self.all_call_dispositions_df
            .lazy()
            .sort(
                by=["LoanMstID", "call_date", "connected_flag", "rank_val", "source_priority", "disposition"],
                descending=[False, True, True, False, False, False]
            )
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("voicebot_latest_disposition"),
                pl.col("category").alias("voicebot_latest_disposition_category"),
                pl.col("call_date").alias("voicebot_latest_disposition_date")
            ])
        )
        self.latest_voicebot_channel_disposition_df = latest_voicebot_channel_disposition_lf.collect()

        # 5. latest_whatsapp_channel_disposition_df
        latest_whatsapp_channel_disposition_lf = (
            self.all_wa_dispositions_df
            .lazy()
            .sort(
                by=["LoanMstID", "wa_date", "connected_flag", "rank_val", "source_priority", "disposition"],
                descending=[False, True, True, False, False, False]
            )
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("whatsapp_latest_disposition"),
                pl.col("category").alias("whatsapp_latest_disposition_category"),
                pl.col("wa_date").alias("whatsapp_latest_disposition_date"),
                pl.col("disposition_ptp_date").alias("whatsapp_latest_disposition_ptp_date")
            ])
        )
        self.latest_whatsapp_channel_disposition_df = latest_whatsapp_channel_disposition_lf.collect()

        # 6. latest_blaster_channel_disposition_df
        latest_blaster_channel_disposition_lf = (
            self.all_blaster_dispositions_df
            .lazy()
            .sort(
                by=["LoanMstID", "blaster_date", "connected_flag", "rank_val", "source_priority", "disposition"],
                descending=[False, True, True, False, False, False]
            )
            .unique(subset=["LoanMstID"], keep="first")
            .select([
                pl.col("LoanMstID"),
                pl.col("disposition").alias("blaster_latest_disposition"),
                pl.col("category").alias("blaster_latest_disposition_category"),
                pl.col("blaster_date").alias("blaster_latest_disposition_date")
            ])
        )
        self.latest_blaster_channel_disposition_df = latest_blaster_channel_disposition_lf.collect()

        # Logging & Assertions for all six DataFrames
        dfs = [
            ("latest_response_df", self.latest_response_df),
            ("latest_voicebot_response_df", self.latest_voicebot_response_df),
            ("latest_whatsapp_response_df", self.latest_whatsapp_response_df),
            ("latest_voicebot_channel_disposition_df", self.latest_voicebot_channel_disposition_df),
            ("latest_whatsapp_channel_disposition_df", self.latest_whatsapp_channel_disposition_df),
            ("latest_blaster_channel_disposition_df", self.latest_blaster_channel_disposition_df),
        ]
        for name, df in dfs:
            self.logger.info("Validation Log - %s shape: %s", name, df.shape)
            self.logger.info("Validation Log - %s columns: %s", name, df.columns)
            self.logger.info("Validation Log - %s column order: %s", name, df.columns)
            self.logger.info("Validation Log - %s null counts: %s", name, df.null_count().to_dicts()[0])
            self.logger.info("Validation Log - %s unique LoanMstID: %d", name, df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s duplicate LoanMstID: %d", name, len(df) - df["LoanMstID"].n_unique())
            self.logger.info("Validation Log - %s estimated memory: %d bytes (%.4f KB)", name, df.estimated_size(), df.estimated_size() / 1024.0)

            # Assert unique LoanMstID constraint
            assert len(df) == df["LoanMstID"].n_unique(), f"Fidelity Check Failed: {name} contains duplicate LoanMstID values!"



    @timer
    def apply_business_rules(self) -> Any:
        """
        Applies disposition ranking, categorization, and settlement overrides.
        """
        logger.info("Applying business rules to master state.")
        # TODO: Connect business_rules module functions.
        raise NotImplementedError("apply_business_rules is not yet implemented.")

    @timer
    def build_daily_pivot(self) -> Any:
        """
        Pivots the consolidated date-wise history columns for reporting.
        """
        logger.info("Building daily activity pivot tables.")
        # TODO: Apply pivot formatting rules.
        raise NotImplementedError("build_daily_pivot is not yet implemented.")

    @timer
    def project_report_columns(self) -> Any:
        """
        Filters and orders columns according to the final schema specification.
        """
        logger.info("Projecting final reporting columns.")
        # TODO: Select and rename final reporting fields.
        raise NotImplementedError("project_report_columns is not yet implemented.")

    @timer
    def paginate(self, data: Any, page: int, page_size: int) -> Any:
        """
        Paginates the final report slice.

        Args:
            data: The processed data structure.
            page: 1-indexed target page number.
            page_size: Maximum record size of a page.
        """
        logger.info("Paginating result set (page=%d, page_size=%d).", page, page_size)
        # TODO: Slice dataset using target bounds.
        raise NotImplementedError("paginate is not yet implemented.")

    @timer
    def generate_report(self, request: ReportRequest) -> Any:
        """
        Triggers the report building workflow for the requested parameter ranges.

        Args:
            request: The ReportRequest object containing date range, paging and export flags.

        Returns:
            JSON-formatted dataframe/records or the full dataframe depending on export flag.
        """
        self.request = request
        self.logger.info(
            "Generating report: start=%s, end=%s, page=%d, page_size=%d, export=%s",
            request.start_date,
            request.end_date,
            request.page,
            request.page_size,
            request.export
        )
        # Step 1: Load raw tables
        self.load_source_tables()

        # Step 2: Build base accounts
        self.build_base_accounts()

        # Step 3: Build voicebot dispositions
        self.build_all_call_dispositions()

        # Step 4: Build WhatsApp dispositions
        self.build_all_wa_dispositions()

        # Step 5: Build Blaster dispositions
        self.build_all_blaster_dispositions()

        # Step 6: Build Best disposition per loan
        self.build_best_disposition_per_loan()

        # Step 7: Build VoiceBot attempt summary
        self.build_voicebot_summary()

        # Step 8: Build WhatsApp attempt summary
        self.build_whatsapp_summary()

        # Step 9: Build Blaster attempt summary
        self.build_blaster_summary()

        # Step 10: Build Communication Summary
        self.build_communication_summary()

        # Step 12: Build Latest Channel State Tables
        self.build_latest_channels()

        # Step 13: Build Collection Tables
        self.build_collections()

        # Step 14: Build Connection Statistics Tables
        self.build_connection_statistics()

        # Step 15: Build Historical Disposition Tables
        self.build_historical_dispositions()

        # Step 16: Build Latest Response & Channel Disposition Tables
        self.build_latest_response_tables()

        # Step 17: Build Enriched Latest Master
        self.build_latest_master()

        # Step 18: Build Daily Call Timeline & Daily Pivot Tables
        self.build_daily_pivots()

        # Step 19: Build Final Report Dataset
        self.build_final_report()

        # Step 20: Export Excel Report
        output_path = self.export_final_report()
        return output_path
